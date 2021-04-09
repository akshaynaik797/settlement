import openpyxl
import xlrd

from movemaster import move_master_to_master_insurer

excel_file = '/home/akshay/temp/4493_892411028018_mediclaim.noble_gmail.com.xlsx'

wbName = 'master.xlsx'
wb = openpyxl.Workbook()
wb.create_sheet('Sheet1')
wb.create_sheet('count')
wb.create_sheet('count_star')
wb.create_sheet('error_sheet')
main_s1 = wb.worksheets[0]
main_s2 = wb.worksheets[1]

sh1 = ['Sno', 'HospitalID', 'InsurerID', 'ALNO', 'ClaimNo', 'MemberID', 'PolicyNo', 'PatientName', 'InsuranceCompany',
       'AccountNo', 'BeneficiaryBank Name', 'Diagnosis', 'UTRNo', 'BilledAmount', 'SettledAmount', 'TDS', 'NetPayable',
       'DiscountAmt', 'COPay', 'PolicyHolder', 'IPNo', 'PrimaryBeneficiary', 'EmployeeID', 'InsurerClaimNo',
       'InsurerMemberID', 'TaxDeductedatSource', 'Netamount payment', 'PaidbythePatient', 'ProrataBasis',
       'PolicyExcessDeductible', 'BeneficiaryName', 'BalanceSumInsuredBeforeClaim', 'NetPayable',
       'BalanceSumInsuredAfterClaim', 'TDS%', 'Remarks', 'PaymentTo', 'DateofAdmission', 'DateofDischarge',
       'AmtPaidtoHospital', 'BillAmt', 'PayableAmt', 'SettledAmt', 'SumInsured', 'ALAmount	Approved', 'Amount',
       'HospitalAmount', 'AmountUtilised', 'FinalAmountSettled', 'DateOfPayment', 'ServiceTax', 'TotalwithServiceTax',
       'InsuredPerson', 'CorporateName', 'DeductibleAmt', 'Transactiondate', 'LOCALAmount', 'ChequeDate',
       'UHCApprovedHospitalAmt', 'InsurerApprovedHospitalAmt', 'InsurerApprovedEmployeeAmt', 'PayableAmount',
       'NEFTTransactionNumber', 'TransactionDate', 'CorporateName', 'Claimed', 'PreHospitalisationPayableAmount',
       'PostHospitalisationPayableAmount', 'AddonBenefit', 'Claimed', 'Paid', 'BillAmount', 'PayableAmount(INR)',
       'BillDate', 'BillNo', 'AmountSettled', 'ApprovedAmount', 'less', 'Excess of Defined / Ailment Limit',
       'policy deduction', 'Limit exceed deduction', 'non payable deduction', 'Bill deduction', 'Other deduction']
sh2 = ['Sr. No.', 'HospitalID', 'InsurerID', 'Claim ID', 'Details', 'Bill amount', 'Payable Amount', 'Deducted Amt',
       'Reason for Deduction', 'Discount']
for i in range(0, len(sh1)):
    # main_s1.cell(row=1, column=i+1).value=i+1
    main_s1.cell(row=1, column=i + 1).value = sh1[i]
    main_s1.cell(row=2, column=i + 1).value = sh1[i]

for i in range(0, len(sh2)):
    main_s2.cell(row=1, column=i + 1).value = sh2[i]
    main_s2.cell(row=2, column=i + 1).value = sh2[i]

wb.save(wbName)
move_master_to_master_insurer('mid')
import pandas as pd
df = pd.read_html(excel_file)[0]

data = []

for index, row in df.iterrows():
    temp = [cell for cell in row]
    data.append(temp)
pass