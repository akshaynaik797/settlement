import os
import subprocess
import PyPDF2
import openpyxl
import sys
import camelot
import pdftotext
import xlrd
from make_log import log_exceptions
from backend import mark_flag
from movemaster import move_master_to_master_insurer

try:
    pdfpath = sys.argv[1]
    onlyfiles = [os.path.split(pdfpath)[1]]
    mypath = os.path.dirname(pdfpath)+'/'

    hosp_name = ''

    with open(pdfpath, "rb") as f:
        pdf = pdftotext.PDF(f)
    with open('temp_files/output.txt', 'w', encoding='utf-8') as f:
        f.write(" ".join(pdf))
    with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
        f = myfile.read()
    if 'has been settled' not in f:
        sys.exit(f'{pdfpath} wrong pdf recieved, so not processed')
    else:
        if 'Balaji Medical' in f:
            op = 'Tpappg@maxhealthcare.com May@2020 outlook.office365.com Max PPT'
            hosp_name = 'Max'
        else:
            op = 'mediclaim@inamdarhospital.org Mediclaim@2019 imap.gmail.com inamdar hospital'
            hosp_name = 'inamdar'
    ###########################################################
    wbkName = 'temp_files/' + 'Medi_Assist' + hosp_name + '.xlsx'
    t, wq =0, 0
    po = []
    wbk = openpyxl.Workbook()
    wbk.create_sheet('Sheet1')
    wbk.create_sheet('Sheet2')
    # wbk.create_sheet('Sheet3')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    s3 = wbk.worksheets[2]
    eu = []
    goble_total = []
    for t in range(0, len(onlyfiles)):
        try:
            with open(mypath + onlyfiles[t], "rb") as f:
                pdf = pdftotext.PDF(f)

            with open('temp_files/output.txt', 'w', encoding='utf-8') as f:
                f.write(" ".join(pdf))
            with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
                f = myfile.read()


            def select_sheet(wks, sheet_name):

                if not sheet_name in wks.sheetnames:
                    wks.create_sheet(sheet_name)
                # print(sheet_name)

                wks.save('temp_files/Medi_Assist' + hosp_name + '.xlsx')


            def select_column(wks, s, ro):
                sheet = wks.worksheets[s]

                max_col = sheet.max_column
                s = []
                for i in range(1, max_col + 1):
                    cell_obj = sheet.cell(row=1, column=i)
                    s.append(cell_obj.value)


            # print(s)

            # print(mypath+'attachments/'+onlyfiles[t])
            CCN = onlyfiles[t].replace('.pdf', '')
            # print(CCN)

            tables = camelot.read_pdf(mypath + onlyfiles[t], pages='1-end')
            a = tables.n

            # print(tables.n)
            # for 2 table case
            if tables.n == 2:
                sh1 = ['Patient Name', 'Insurance Company', 'Medi Assist ID', 'Policy Holder', 'IP No.', 'Policy No.',
                       'Primary Beneficiary', 'Employee ID', 'Insurer Claim No', 'Insurer Member ID', 'diagnosis',
                       'doa',
                       'dod']
                sh2 = ['Settled Amount (INR)', 'Settlement Date', 'UTR Number', 'Account Holder Name', 'Bank Name',
                       'Account Number']

                tables.export('temp_files/foo1.xls', f='excel')
                loc = ("temp_files/foo1.xls")
                wb = xlrd.open_workbook(loc)
                pdfFileObj = open(mypath + onlyfiles[t], 'rb')
                pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
                pageObj = pdfReader.getPage(0)
                f = pageObj.extractText()
                w = f.find('recommended')
                g = f[w:]
                u = g.find('€') + w + 1
                total = f[u:]
                if total.find(' ') != -1:
                    uop = g.find('The') + w
                    total = f[u:uop]

                if total[0] < '0' or total[0] > '9':
                    pageObj = pdfReader.getPage(1)
                    df = pageObj.extractText()
                    w = df.find('payment')
                    g = df[w:]
                    u = df.find('€') + 1
                    uop = df.find('The')
                    total = df[u:uop]
                goble_total.append(total)
                # print(f)
                '''text_file = open("mail.txt", "w")
                n = text_file.write(f)
                text_file.close()'''

                for i in range(0, len(sh1)):
                    s1.cell(row=1, column=i + 3).value = sh1[i]

                x1 = f.find('treatment of') + 12
                g = f[x1:]
                # print(g)
                y1 = g.find(' at ') + x1

                x = f.find('Claimant Name :')
                x3 = f.find('issued by') + 9
                g = f[x3:]
                y3 = g.find(',') + x3
                y = f.find('MAID:')
                x2 = f.find('Relationship')
                x4 = f.find('Policy No:') + 10
                y4 = f.find('Period of Insurance')
                x5 = f.find('Primary Member') + 16
                g = f[x5:]
                y5 = g.find('(') + x5
                y6 = g.find(')') + x5
                x7 = f.find('Member Id') + 10
                y7 = f.find('Address')
                gh = []

                gh.append(f[x + 15:y])
                gh.append(f[x3:y3])
                gh.append(f[y + 5:x2])
                gh.append(' ')  # policy holder not exracted
                gh.append(' ')
                gh.append(f[x4:y4])
                gh.append(f[x5:y5])
                gh.append(f[y5 + 1:y6])
                gh.append(' ')
                gh.append(f[x7:y7])
                gh.append(f[x1:y1])

                w = f.find('from') + 5
                g = f[w:]
                u = g.find('to') + w
                gh.append(f[w:u])

                u1 = g.find('.') + w
                gh.append(f[u + 2:u1])

                for i in range(0, len(gh)):
                    s1.cell(row=t + 2, column=i + 3).value = gh[i]

                for i in range(0, len(sh2)):
                    s2.cell(row=1, column=i + 3).value = sh2[i]

                hg = []
                x8 = f.find('Amount Settled') + 18
                y8 = f.find('Category Break')
                x9 = f.find('Settlement Date') + 16
                y9 = f.find('Insurer')
                x10 = f.find('Transaction Id') + 15
                y10 = f.find('Account Holder Name')
                x11 = f.find('Bank name')
                y11 = f.find('Branch')
                y10 = f.find('Acc No')
                hg.append(f[x8:y8])
                hg.append(f[x9:y9])
                hg.append(f[x10:y10])
                hg.append(' ')
                hg.append(f[x11 + 10:y11])
                hg.append(f[y10 + 7:x11])

                for i in range(0, len(hg)):
                    s2.cell(row=t + 2, column=i + 3).value = hg[i]

                s = []
                d = []
                sheet_3 = wb.sheet_by_index(1)
                sheet_3.cell_value(0, 0)

                for i in range(1, sheet_3.nrows):
                    s.append(sheet_3.cell_value(i, 1))
                    d.append(sheet_3.cell_value(i, 2))
                s = [sub.replace('\n', ' ') for sub in s]
                # print(s,d)
                if (t == 0):
                    for i in range(0, len(s) - 1):
                        s3.cell(row=1, column=i + 3).value = s[i]
                        po.append(s[i])
                        length = i + 3
                        s3.cell(row=t + 2, column=i + 3).value = d[i]
                else:

                    for i in range(0, len(s) - 1):
                        if s[i] not in po:
                            po.append(s[i])
                            s3.cell(row=1, column=length + 1).value = s[i]
                            s3.cell(row=t + 2, column=length + 1).value = d[i]
                            length = length + 1
                        if s[i] in po:
                            # print('hi')
                            u = po.index(s[i])
                            s3.cell(row=t + 2, column=u + 3).value = d[i]
                s_v = []

                sheet_2 = wb.sheet_by_index(0)
                sheet_2.cell_value(0, 0)
                b = []
                p = []
                np = []
                r = []
                for i in range(3, sheet_2.nrows):
                    s_v.append(sheet_2.cell_value(i, 1))
                    b.append(sheet_2.cell_value(i, 2))
                    p.append(sheet_2.cell_value(i, 3))
                    np.append(sheet_2.cell_value(i, 4))
                    r.append(sheet_2.cell_value(i, 5))
                for i in range(len(s_v)):
                    select_sheet(wbk, s_v[i])
                ro = []
                ro.append(sheet_2.row_values(2))
                ro = ro[0][2:]
                # print(ro)
                xls = xlrd.open_workbook('temp_files/Medi_Assist' + hosp_name + '.xlsx', on_demand=True)
                sheet_list = xls.sheet_names()
                # print(sheet_list)
                for i in range(len(s_v)):
                    sheet_name = s_v[i]
                    for y in range(len(sheet_list)):
                        if sheet_name in sheet_list[y]:
                            for k in range(0, len(ro)):
                                sheet = wbk.worksheets[y]
                                # print(ro[k],k,sheet)
                                sheet.cell(row=1, column=k + 3).value = ro[k]
                            sheet.cell(row=t + 2, column=3).value = b[i]
                            sheet.cell(row=t + 2, column=4).value = p[i]
                            sheet.cell(row=t + 2, column=5).value = np[i]
                            sheet.cell(row=t + 2, column=6).value = r[i]
                        # select_column(wbk,y,ro)

            if tables.n == 4 or tables.n == 3:
                if tables.n == 4:
                    tables.export('temp_files/foo1.xls', f='excel')
                    loc = ("temp_files/foo1.xls")
                if tables.n == 3:
                    tables = camelot.read_pdf(mypath + onlyfiles[t], line_scale=100)
                    tables.export('temp_files/foo1.xls', f='excel')
                    loc = ("temp_files/foo1.xls")

                wb = xlrd.open_workbook(loc)
                sheet_0 = wb.sheet_by_index(0)
                sheet_0.cell_value(0, 0)
                k = []
                v = []
                for i in range(1, sheet_0.nrows):
                    k.append(sheet_0.cell_value(i, 1))
                    v.append(sheet_0.cell_value(i, 2))

                pdfFileObj = open(mypath + onlyfiles[t], 'rb')
                pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
                pageObj = pdfReader.getPage(0)
                f = pageObj.extractText()
                w = f.find('recommended')
                g = f[w:]
                u = g.find('€') + w + 1
                total = f[u:]
                if total.find(' ') != -1:
                    uop = g.find('The') + w
                    total = f[u:uop]

                if total[0] < '0' or total[0] > '9':
                    pageObj = pdfReader.getPage(1)
                    df = pageObj.extractText()
                    w = df.find('payment')
                    g = df[w:]
                    u = df.find('€') + 1
                    uop = df.find('The')
                    total = df[u:uop]
                goble_total.append(total)
                x1 = f.find('treatment of') + 12
                g = f[x1:]

                # print(g)
                y1 = g.find(' at ') + x1
                k.append('diagnosis')
                k.append('doa')
                k.append('dod')
                v.append(f[x1:y1])

                w = f.find('from') + 5
                g = f[w:]
                u = g.find('to') + w
                v.append(f[w:u])

                u1 = g.find('.') + w
                v.append(f[u + 2:u1])
                m = []
                l = []
                sheet_1 = wb.sheet_by_index(1)
                sheet_1.cell_value(0, 0)
                for i in range(0, len(k)):
                    s1.cell(row=1, column=i + 3).value = k[i]
                for i in range(1, sheet_1.nrows):
                    m.append(sheet_1.cell_value(i, 1))
                    l.append(sheet_1.cell_value(i, 2))
                for i in range(0, len(m)):
                    s2.cell(row=1, column=i + 3).value = m[i]
                s = []
                d = []
                sheet_3 = wb.sheet_by_index(3)
                sheet_3.cell_value(0, 0)

                for i in range(1, sheet_3.nrows):
                    s.append(sheet_3.cell_value(i, 1))
                    d.append(sheet_3.cell_value(i, 2))
                s = [sub.replace('\n', ' ') for sub in s]
                if (t == 0):
                    for i in range(0, len(s) - 1):
                        s3.cell(row=1, column=i + 3).value = s[i]
                        po.append(s[i])
                        length = i + 3
                        s3.cell(row=t + 2, column=i + 3).value = d[i]
                    # print(po)
                else:

                    for i in range(0, len(s) - 1):
                        for j in range(0, len(po)):
                            if s[i] not in po:
                                # print(s[i])
                                po.append(s[i])
                                s3.cell(row=1, column=length + 1).value = s[i]
                                s3.cell(row=t + 2, column=length + 1).value = d[i]
                                length = length + 1
                            elif s[i] in po:
                                u = po.index(s[i])
                                s3.cell(row=t + 2, column=u + 3).value = d[i]

                # print(len(k))
                # res = {k[i]: v[i] for i in range(len(k))}
                # print (res)

                # if(wbk.worksheets=='<Worksheet "Sheet1">'):

                for i in range(0, len(v)):
                    s1.cell(row=t + 2, column=i + 3).value = v[i]
                for i in range(0, len(l)):
                    s2.cell(row=t + 2, column=i + 3).value = l[i]

                # dynamic method table 3
                s_v = []

                sheet_2 = wb.sheet_by_index(2)
                sheet_2.cell_value(0, 0)
                b = []
                p = []
                np = []
                r = []
                for i in range(2, sheet_2.nrows):
                    s_v.append(sheet_2.cell_value(i, 1))
                    b.append(sheet_2.cell_value(i, 2))
                    p.append(sheet_2.cell_value(i, 3))
                    np.append(sheet_2.cell_value(i, 4))
                    r.append(sheet_2.cell_value(i, 5))
                for i in range(len(s_v)):
                    select_sheet(wbk, s_v[i])
                ro = []
                ro.append(sheet_2.row_values(1))
                ro = ro[0][2:]
                # print(s_v)
                xls = xlrd.open_workbook('temp_files/Medi_Assist' + hosp_name + '.xlsx', on_demand=True)
                sheet_list = xls.sheet_names()
                # print(sheet_list)
                for i in range(len(s_v)):
                    sheet_name = s_v[i]
                    for y in range(len(sheet_list)):
                        if sheet_name in sheet_list[y]:
                            for k in range(0, len(ro)):
                                sheet = wbk.worksheets[y]
                                # print(sheet)
                                sheet.cell(row=1, column=k + 3).value = ro[k]
                            sheet.cell(row=t + 2, column=3).value = b[i]
                            sheet.cell(row=t + 2, column=4).value = p[i]
                            sheet.cell(row=t + 2, column=5).value = np[i]
                            sheet.cell(row=t + 2, column=6).value = r[i]
                        # select_column(wbk,y,ro)


        except Exception as e:
            log_exceptions()
            s1.cell(row=t + 2, column=1).value = 'error'
    for t in range(0, len(onlyfiles)):
        CCN = onlyfiles[t].replace('.pdf', '').split('_')[-1]
        for wd in wbk.worksheets:
            wd.cell(row=1, column=1).value = 'Sr. No.'
            wd.cell(row=1, column=2).value = 'CCN'
            wd.cell(row=t + 2, column=1).value = t + 1
            wd.cell(row=t + 2, column=2).value = CCN
        if t in eu:
            s1.cell(row=t + 2, column=1).fill = 'redFill'
            s1.cell(row=t + 2, column=1).value = 'error'
    if (t != 0):
        u = length + 1
        s3.cell(row=1, column=u).value = 'Net amount recommended for payment'
    for t in range(0, len(onlyfiles)):
        s3.cell(row=t + 2, column=u).value = goble_total[t]

    # print(po)
    print("Done")
    wbk.save(wbkName)
    wbk.close()
    subprocess.run(["python", "make_master.py", 'Medi_Assist', op, '', wbkName])
    ###########################################################
    move_master_to_master_insurer(sys.argv[2], pdfpath=pdfpath)
    mark_flag('X', sys.argv[1])
    print(f'processed {wbkName}')
except SystemExit as e:
    v = e.code
    if 'exit' in v:
        a =1
        os._exit(0)
except:
    log_exceptions()
    pass