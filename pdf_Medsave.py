import os
import subprocess
import PyPDF2
import openpyxl
import sys
import camelot
import pdftotext
import xlrd
from make_log import log_exceptions
from movemaster import move_master_to_master_insurer

try:
    pdfpath = sys.argv[1]
    onlyfiles = [os.path.split(pdfpath)[1]]
    mypath = os.path.dirname(pdfpath)+'/'

    hosp_name = ''

    with open(pdfpath, "rb") as f:
        pdf = pdftotext.PDF(f)
    with open('temp_files/output.txt', 'w', encoding='utf-8') as f:
        f.write(" ".join(pdf))
    with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
        f = myfile.read()
    if 'has been settled' not in f:
        sys.exit(f'{pdfpath} wrong pdf recieved, so not processed')
    else:
        if 'Balaji Medical' in f:
            op = 'Tpappg@maxhealthcare.com May@2020 outlook.office365.com Max PPT'
            hosp_name = 'Max'
        else:
            op = 'mediclaim@inamdarhospital.org Mediclaim@2019 imap.gmail.com inamdar hospital'
            hosp_name = 'inamdar'
    ###########################################################
    wbkName = 'temp_files/' + 'Medsave' + hosp_name + '.xlsx'
    t, wq =0, 0
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    for t in range(0, len(onlyfiles)):
        sh1 = ['Sr No.', 'Claim No', 'Name of Patient', 'Employee ID', 'Employee Name', 'Policy Number',
               'Member Id',
               'UTR NO.', 'insurance company', 'settled amt', 'neft no.', 'account no.', 'transaction date', 'doa',
               'dod', 'billed amt', 'Diagnosis']
        sh2 = ['Sr No.', 'Claim ID', 'category', 'Deduction Amt(Rs)', 'Reason of Deduction (If any)']

        with open(mypath + onlyfiles[t], "rb") as f:
            pdf = pdftotext.PDF(f)

        with open('temp_files/output.txt', 'w', encoding='utf-8') as f:
            f.write(" ".join(pdf))
        with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
            f = myfile.read()

        for i in range(0, len(sh1)):
            s1.cell(row=1, column=i + 1).value = sh1[i]
        for i in range(0, len(sh2)):
            s2.cell(row=1, column=i + 1).value = sh2[i]

        tables = camelot.read_pdf(mypath + onlyfiles[t], pages='all')

        if tables.n != 0:
            tables.export('temp_files/foo1.xls', f='excel')
            loc = ("temp_files/foo1.xls")
            wb = xlrd.open_workbook(loc)
        hg = []
        w = f.find('bearing No.') + 11
        g = f[w:]
        u = g.find('against policy') + w
        hg.append(f[w:u])

        w = f.find('Patient Name') + 11
        g = f[w:]
        x1 = g.find(':') + w
        u = g.find('\n') + w
        hg.append(f[x1 + 1:u])

        w = f.find('Employee Code') + 11
        g = f[w:]
        x1 = g.find(':') + w
        u = g.find('\n') + w
        hg.append(f[x1 + 1:u])

        w = f.find('Proposer Name') + 13
        g = f[w:]
        x1 = g.find(':') + w
        u = g.find('\n') + w
        hg.append(f[x1 + 1:u])

        w = f.find('Policy No.') + 10
        g = f[w:]
        x1 = g.find(':') + w
        u = g.find('\n') + w
        hg.append(f[x1 + 1:u])

        w = f.find('Card No.') + 7
        g = f[w:]
        x1 = g.find(':') + w
        u = g.find('\n') + w
        hg.append(f[x1 + 1:u])

        w = f.find('Payment Float No.') + 17
        g = f[w:]
        x1 = g.find(':') + w
        u = g.find('\n') + w
        hg.append(f[x1 + 1:u])

        w = f.find('issued by') + 9
        g = f[w:]
        u = g.find('has been') + w
        hg.append(f[w:u])

        w = f.find('settled for') + 11
        g = f[w:]
        u = g.find('(') + w
        hg.append(f[w:u])

        w = f.find('ECS/ NEFT') + 9
        g = f[w:]
        u = g.find('in your') + w
        hg.append(f[w:u])

        w = f.find('Account No.') + 11
        g = f[w:]
        u = g.find('on') + w
        hg.append(f[w:u])
        x1 = g.find('against') + w
        hg.append(f[u + 3:x1])

        w = f.find('period from') + 11
        g = f[w:]
        u = g.find('to') + w
        hg.append(f[w:u])
        x1 = g.find('The detail') + w
        hg.append(f[u + 3:x1])

        w = f.find('claimed for Rs.') + 15
        g = f[w:]
        u = g.find('towards') + w
        hg.append(f[w:u])

        w = f.find('treatment of') + 12
        g = f[w:]
        u = g.find('at') + w
        hg.append(f[w:u])

        hg = [sub.replace('  ', '') for sub in hg]
        hg = [sub.replace('\n', ' ') for sub in hg]
        hg = [sub.replace(':', '') for sub in hg]
        hg = [sub.replace('Rs.', '') for sub in hg]
        # print(hg)

        s1.cell(row=t + 2, column=1).value = t + 1
        for i in range(0, len(hg)):
            s1.cell(row=t + 2, column=i + 2).value = hg[i]

        s = []
        gh = []
        h = []
        if tables.n != 0:
            for j in range(0, tables.n):
                sheet_n = wb.sheet_by_index(j)
                sheet_n.cell_value(0, 0)
                for i in range(2, sheet_n.nrows):
                    s.append(sheet_n.cell_value(i, 1))
                    gh.append(sheet_n.cell_value(i, 2))
                    h.append(sheet_n.cell_value(i, 3))
            for i in range(0, len(s)):
                wq += 1
                row_num = s2.max_row
                s2.cell(row=row_num + 1, column=1).value = wq
                s2.cell(row=row_num + 1, column=2).value = hg[0]
                s2.cell(row=row_num + 1, column=3).value = s[i]
                s2.cell(row=row_num + 1, column=4).value = gh[i]
                s2.cell(row=row_num + 1, column=5).value = h[i]
    print("Done")
    wbk.save(wbkName)
    wbk.close()
    subprocess.run(["python", "make_master.py", 'medsave', op, '', wbkName])
    ###########################################################
    move_master_to_master_insurer('')
    print(f'processed {wbkName}')

except:
    log_exceptions()
    pass