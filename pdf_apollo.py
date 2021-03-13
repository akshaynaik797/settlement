import os
import subprocess
import PyPDF2
import openpyxl
import sys
import camelot
import pdftotext
import xlrd
from make_log import log_exceptions
from backend import mark_flag
from movemaster import move_master_to_master_insurer

try:
    pdfpath = sys.argv[1]
    onlyfiles = [os.path.split(pdfpath)[1]]
    mypath = os.path.dirname(pdfpath)+'/'
    ccn = []
    name = []
    uhid = []

    hosp_name = ''

    with open(pdfpath, "rb") as f:
        pdf = pdftotext.PDF(f)
    with open('temp_files/output.txt', 'w', encoding='utf-8') as f:
        f.write(" ".join(pdf))
    with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
        f = myfile.read()
    if 'wanted_text' not in f:
        sys.exit(f'{pdfpath} wrong pdf recieved, so not processed')
    else:
        if 'Balaji Medical' in f:
            op = 'Tpappg@maxhealthcare.com May@2020 outlook.office365.com Max PPT'
            hosp_name = 'Max'
        else:
            op = 'mediclaim@inamdarhospital.org Mediclaim@2019 imap.gmail.com inamdar hospital'
            hosp_name = 'inamdar'
    ###########################################################
    wbkName = 'temp_files/' + 'apollo' + hosp_name + '.xlsx'
    t, wq =0, 0
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    for t in range(0, len(onlyfiles)):
        sh1 = ['Sr No.', 'Preauth Id', 'Hospital Name', 'Claimed Amount', 'Diagnosis', 'Billed Amount', 'Date of Admission',
               'settled Amount', 'Date of Discharge', 'Cheque Amount', 'Cheque Number/NEFT reference', 'Disallowed Amount',
               'Cheque/NEFT date', 'Discount Amount', 'TDS Amount', 'IP Number', 'Bill No', 'ccn', 'uhid', 'patient name']
        sh2 = ['Sr No.', 'Claim ID', 'category', 'Disallowance amount', 'Disallowance Reasons']

        for i in range(0, len(sh1)):
            s1.cell(row=1, column=i + 1).value = sh1[i]
        for i in range(0, len(sh2)):
            s2.cell(row=1, column=i + 1).value = sh2[i]
        tables = camelot.read_pdf(mypath + onlyfiles[t], pages='all', Line_scale=10)
        tables.export('temp_files/foo1.xls', f='excel')
        loc = ("temp_files/foo1.xls")
        wb = xlrd.open_workbook(loc)
        s = []
        sheet_3 = wb.sheet_by_index(0)
        sheet_3.cell_value(0, 0)

        for i in range(1, sheet_3.nrows):
            s.append(sheet_3.cell_value(i, 2))
            s.append(sheet_3.cell_value(i, 4))
        mid = s[-1]
        s.pop(-1)
        s = [sub.replace('\t', ' ') for sub in s]
        s = [sub.replace('Rs.', '') for sub in s]
        # print(s)
        s1.cell(row=t + 2, column=1).value = t + 1
        s1.cell(row=t + 2, column=2).value = mid
        for i in range(0, len(s)):
            s1.cell(row=t + 2, column=i + 3).value = s[i]
        s1.cell(row=t + 2, column=i + 4).value = ccn[t]
        s1.cell(row=t + 2, column=i + 5).value = uhid[t]
        s1.cell(row=t + 2, column=i + 6).value = name[t]
        with open(mypath + onlyfiles[t], "rb") as f:
            pdf = pdftotext.PDF(f)

        with open('temp_files/output.txt', 'w', encoding='utf-8') as f:
            f.write(" ".join(pdf))
        with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
            f = myfile.read()
        hg = []
        w = f.find('Disallowance Reasons :') + 22
        u = f.find('Please note')
        g = f[w:u]
        sy = g.split('\n')
        sy.pop(0)
        sy.pop(-1)
        for i in sy:
            # print(i)
            if (i.find(':') != -1):
                k = i
                k = k.replace(':', '')
                continue
            else:
                w1 = i.find('Rs.') + 3
                g = i[w1:]
                u1 = g.find('.') + w1 + 3
                m = i[w1:u1]
                h = i[u1:]
            row_num = s2.max_row + 1
            wq += 1
            s2.cell(row=row_num, column=1).value = wq
            s2.cell(row=row_num, column=2).value = mid
            s2.cell(row=row_num, column=3).value = k
            s2.cell(row=row_num, column=4).value = m
            s2.cell(row=row_num, column=5).value = h


    print("Done")
    wbk.save(wbkName)
    wbk.close()
    subprocess.run(["python", "make_master.py", 'apollo', op, '', wbkName])
    ###########################################################
    move_master_to_master_insurer('', pdfpath=pdfpath)
    mark_flag('X', sys.argv[1])
    print(f'processed {wbkName}')
except SystemExit as e:
    v = e.code
    if 'exit' in v:
        a =1
        os._exit(0)
except:
    log_exceptions()
    pass