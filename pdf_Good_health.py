import subprocess
import openpyxl
import sys
import pdftotext
import tabula
from make_log import log_exceptions
from movemaster import move_master_to_master_insurer

try:
    # pdfpath = '/home/akshay/PycharmProjects/settelement/backups/fgh_07202020163942/16-FGH-20-3-868412-01.pdf'
    pdfpath = sys.argv[1]
    hosp_name = ''

    with open(pdfpath, "rb") as f:
        pdf = pdftotext.PDF(f)
    with open('temp_files/output.txt', 'w') as f:
        f.write(" ".join(pdf))
    with open('temp_files/output.txt', 'r') as myfile:
        f = myfile.read()
    if 'Hospital Payment' not in f:
        sys.exit(f'{pdfpath} wrong pdf recieved, so not processed')
    else:
        if 'Balaji Medical' in f:
            op = 'Tpappg@maxhealthcare.com May@2020 outlook.office365.com Max PPT'
            hosp_name = 'Max'
        else:
            op = 'mediclaim@inamdarhospital.org Mediclaim@2019 imap.gmail.com inamdar hospital'
            hosp_name = 'inamdar'
    ###########################################################
    wbkName = 'temp_files/' + 'Good_heath' + hosp_name + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    t, wq = 0, 0
    sh1 = ['Sr No.', 'Claim No', 'Claimant/Patient', 'Employee ID', 'Employee Name', 'Policy Number', 'Member Id',
           'Date of Admission', 'Date of Discharge', 'Bill/IP No', 'insurer', 'transaction date', 'Claim Type',
           'Diagnosis']
    sh2 = ['Sr No.', 'Claim ID', 'Service Item', 'Claimed Amt.', 'Disallowed Amt.', 'Amount	Deduction',
           'Remarks']

    for i in range(0, len(sh1)):
        s1.cell(row=1, column=i + 1).value = sh1[i]
    for i in range(0, len(sh2)):
        s2.cell(row=1, column=i + 1).value = sh2[i]

    with open(pdfpath, "rb") as f:
        pdf = pdftotext.PDF(f)

    with open('temp_files/output.txt', 'w') as f:
        f.write(" ".join(pdf))
    with open('temp_files/output.txt', 'r') as myfile:
        f = myfile.read()
    df = tabula.read_pdf(pdfpath, pages='all')
    df['Service Item'] = df['Service Item'].fillna('$$')
    # print(df)
    hg = []
    w = f.find('CCN') + 3
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('\n') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Name of Patient') + 15
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('\n') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Employee ID') + 11
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('\n') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Employee Name') + 13
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('\n') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Policy Number') + 13
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('\n') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Card No') + 7
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('\n') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Date of Admission') + 17
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('\n') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Date of Discharge') + 17
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('\n') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Bill/IP No') + 10
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('\n') + w
    hg.append(f[x1 + 1:u])

    w = f.find('insurer') + 7
    g = f[w:]
    u = g.find('sum of') + w
    hg.append(f[w:u])

    w = f.find('transferred on') + 14
    g = f[w:]
    u = g.find('to your') + w
    hg.append(f[w:u])

    w = f.find('EFT is') + 6
    g = f[w:]
    u = g.find('from') + w
    hg.append(f[w:u])

    w = f.find('Claim Type') + 10
    g = f[w:]
    u = g.find('\n') + w
    hg.append(f[w:u])

    w = f.find('Diagnosis') + 9
    g = f[w:]
    u = g.find(' Service') + w
    hg.append(f[w:u])

    hg = [sub.replace('  ', '') for sub in hg]
    hg = [sub.replace('\n', ' ') for sub in hg]
    hg = [sub.replace(':', '') for sub in hg]
    # print(hg)

    s1.cell(row=t + 2, column=1).value = t + 1
    for i in range(0, len(hg)):
        s1.cell(row=t + 2, column=i + 2).value = hg[i]
    s = df['Service Item']
    gh = df['Claimed Amt.']
    h = df['Disallowed Amt.']
    g = df['Amount']
    hj = df['Deduction Remarks']
    for i in range(0, len(s)):
        wq += 1
        row_num = s2.max_row
        if (s[i] != '$$'):
            # print(s[i])
            s2.cell(row=row_num + 1, column=1).value = wq
            s2.cell(row=row_num + 1, column=2).value = hg[0]
            s2.cell(row=row_num + 1, column=3).value = s[i]
            s2.cell(row=row_num + 1, column=4).value = gh[i]
            s2.cell(row=row_num + 1, column=5).value = h[i]
            s2.cell(row=row_num + 1, column=6).value = g[i]
            s2.cell(row=row_num + 1, column=7).value = hj[i]
        else:
            # print('hi')
            s2.cell(row=row_num, column=7).value = hj[i - 1] + ' ' + hj[i]
    wbk.save(wbkName)
    wbk.close()
    subprocess.run(["python", "make_master.py", 'Good_health', op, '', wbkName])
    ###########################################################
    move_master_to_master_insurer('')
    print(f'processed {wbkName}')

except:
    log_exceptions()
    pass