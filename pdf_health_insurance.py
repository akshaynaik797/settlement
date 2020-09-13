import re
import subprocess
import sys

import openpyxl
import pdftotext

from make_log import log_exceptions
from movemaster import move_master_to_master_insurer

try:
    pdfpath = sys.argv[1]
    hosp_name = ''

    with open(pdfpath, "rb") as f:
        pdf = pdftotext.PDF(f)
    with open('temp_files/output.txt', 'w', encoding='utf-8') as f:
        f.write(" ".join(pdf))
    with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
        f = myfile.read()
    if 'Discharge Voucher cum Claim Settlement Letter' not in f:
        sys.exit(f'{pdfpath} wrong pdf recieved, so not processed')
    else:
        if 'Balaji Medical' in f:
            op = 'Tpappg@maxhealthcare.com May@2020 outlook.office365.com Max PPT'
            hosp_name = 'Max'
        else:
            op = 'mediclaim@inamdarhospital.org Mediclaim@2019 imap.gmail.com inamdar hospital'
            hosp_name = 'inamdar'
    ###########################################################
    wbkName = 'temp_files/' + 'health_insurance' + hosp_name + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    t, wq = 0, 0
    sh1 = ['sr no', 'CCN', 'Insurer', 'Proposer Name', 'diagnosis', 'Policy Number', 'Patient Name', 'doa', 'dod',
           'UHID', 'NEFT Date', 'NEFT No.', 'BilledAmount', 'disallowed', 'DiscountAmt', 'COPay', 'deduction',
           'TDS', 'SettledAmount']
    for i in range(0, len(sh1)):
        s1.cell(row=1, column=i + 1).value = sh1[i]
    sh2 = ['sr no', 'CCN', 'Category', 'Sub Category', 'Requested Amount', 'Approved Amount', 'Deducted Amount',
           'Reason']
    for i in range(0, len(sh2)):
        s2.cell(row=1, column=i + 1).value = sh2[i]
    # tables = camelot.read_pdf(mypath+onlyfiles[t],pages='all',line_scale=50)
    # tables.export('health_insurance/foo1.xls', f='excel')
    # loc = ("health_insurance/foo1.xls")
    with open(pdfpath, "rb") as f:
        pdf = pdftotext.PDF(f)

    with open('temp_files/output.txt', 'w', encoding='utf-8') as f:
        f.write(" ".join(pdf))
    with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
        f = myfile.read()
    hg = []
    w = f.find('claim number') + 12
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('\n') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Insurer') + 7
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('Proposer Name') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Proposer Name') + 13
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('Diagnosis') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Diagnosis') + 9
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('Policy Number') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Policy Number') + 13
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('Patient Name') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Patient Name') + 12
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('Hospital Name') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Period of Hospitalization') + 25
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('to') + w
    hg.append(f[x1 + 1:u])
    x1 = g.find('Patient’s Member UHID') + w
    hg.append(f[u + 2:x1])

    w = f.find('Patient’s Member UHID') + 22
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('\n') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Date') + 5
    g = f[w:]
    u = g.find('\n') + w
    hg.append(f[w:u])

    w = f.find('Neft/Cheque number') + 18
    g = f[w:]
    u = g.find('\n') + w
    hg.append(f[w:u])

    w = f.find('Billed Amount') + 12
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('\n') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Amount Disallowed') + 17
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('\n') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Discount') + 8
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('\n') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Co-payment') + 10
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('\n') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Deductible') + 10
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('\n') + w
    hg.append(f[x1 + 1:u])

    w = f.find('TDS') + 3
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('\n') + w
    hg.append(f[x1 + 1:u])

    w = f.find('Amount Paid') + 11
    g = f[w:]
    x1 = g.find(':') + w
    u = g.find('(') + w
    hg.append(f[x1 + 1:u])

    hg = [sub.replace('\n', ' ') for sub in hg]
    hg = [sub.replace('  ', '') for sub in hg]
    hg = [sub.replace('Rs', '') for sub in hg]
    hg = [sub.replace(':', '') for sub in hg]

    # print(hg)

    for i in range(0, len(hg)):
        s1.cell(row=t + 2, column=1).value = t + 1
        s1.cell(row=t + 2, column=i + 2).value = hg[i]

    w = f.find('Reason') + 5
    g = f[w:]
    x1 = g.find('\n') + w
    u = g.find('Payment Summary') + w
    temp = f[x1:u]
    temp = temp.split('\n')
    so1 = []
    so2 = []
    so3 = []
    so4 = []
    so5 = []
    so6 = []
    for k in temp:
        if k != '':
            temp2 = re.findall("\d+\.\d+", k)
            if len(temp2) != 0:
                k = k.replace('   ', '$')
                while (k.find('$$') != -1):
                    k = k.replace('$$', '$')

                temp1 = k.split('$')
                so1.append(temp1[0])
                so2.append(temp1[1])
                so3.append(temp2[0])
                so4.append(temp2[1])
                so5.append(temp2[2])
                if (len(temp1) == 6):
                    so6.append(temp1[5])
                else:
                    so6.append(' ')
            else:
                continue
    so1 = [sub.replace('  ', '') for sub in so1]
    so2 = [sub.replace('  ', '') for sub in so2]
    so3 = [sub.replace('  ', '') for sub in so3]
    so4 = [sub.replace('  ', '') for sub in so4]
    so5 = [sub.replace('  ', '') for sub in so5]
    so6 = [sub.replace('  ', '') for sub in so6]

    for i in range(0, len(so1)):
        wq += 1
        row_num = s2.max_row
        s2.cell(row=row_num + 1, column=1).value = wq
        s2.cell(row=row_num + 1, column=2).value = hg[0]
        s2.cell(row=row_num + 1, column=3).value = so1[i]
        s2.cell(row=row_num + 1, column=4).value = so2[i]
        s2.cell(row=row_num + 1, column=5).value = so3[i]
        s2.cell(row=row_num + 1, column=6).value = so4[i]
        s2.cell(row=row_num + 1, column=7).value = so5[i]
        s2.cell(row=row_num + 1, column=8).value = so6[i]



    print("Done")
    wbk.save(wbkName)
    wbk.close()
    subprocess.run(["python", "make_master.py", 'health_insurance', op, '', wbkName])
    ###########################################################
    move_master_to_master_insurer('')
    print(f'processed {wbkName}')

except:
    log_exceptions()
    pass