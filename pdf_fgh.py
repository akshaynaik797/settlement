import subprocess
import openpyxl
import sys
import camelot
import pdftotext
import xlrd
from make_log import log_exceptions
from movemaster import move_master_to_master_insurer

try:
    pdfpath = sys.argv[1]
    hosp_name = ''

    with open(pdfpath, "rb") as f:
        pdf = pdftotext.PDF(f)
    with open('temp_files/output.txt', 'w') as f:
        f.write(" ".join(pdf))
    with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
        f = myfile.read()
    if 'Hospital Payment' not in f:
        sys.exit(f'{pdfpath} wrong pdf recieved, so not processed')
    else:
        if 'Balaji Medical' in f:
            op = 'Tpappg@maxhealthcare.com May@2020 outlook.office365.com Max PPT'
            hosp_name = 'Max'
        else:
            op = 'mediclaim@inamdarhospital.org Mediclaim@2019 imap.gmail.com inamdar hospital'
            hosp_name = 'inamdar'
    wbkName = 'temp_files/' + 'fgh' + hosp_name + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('Sheet1')
    # wbk.create_sheet('Sheet3')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    t = 0
    wq = 0
    sh1 = ['Policy Number', 'Reference/UTR No.', 'Payment Date', 'ID Card Number', 'Discount Deduction',
           'Co-Payment']
    for i in range(0, len(sh1)):
        s1.cell(row=1, column=i + 16).value = sh1[i]
    sh2 = ['Sr. No.', 'Claim Number', 'Biling Head', 'Bill No.', 'Claimed', 'Disallowed', 'Approved',
           'Disallowed Reason']
    sh3 = ['Patient Name', 'Hospital Name', 'Date of Admission', 'Date of Discharge', 'Hospital Bill Number',
           'Beneficiary Name', 'Payment Type', 'A/C No', 'Bank Name', 'Bill Amount', 'Disallowed Amount',
           'Service Tax', 'Approved Amount', 'TDS']
    for i in range(0, len(sh2)):
        s2.cell(row=1, column=i + 1).value = sh2[i]
    tables = camelot.read_pdf(pdfpath, pages='all', Line_scale=100)
    tables.export('temp_files/foo1.xls', f='excel')
    loc = ("temp_files/foo1.xls")
    wb = xlrd.open_workbook(loc)
    s = []
    d = []
    sheet_3 = wb.sheet_by_index(0)
    sheet_3.cell_value(0, 0)

    for i in range(1, sheet_3.nrows):
        s.append(sheet_3.cell_value(i, 1))
        d.append(sheet_3.cell_value(i, 2))
    ccn = d[5]
    s.pop(5)
    d.pop(5)
    if (len(s) == 13):
        for i in range(0, len(s)):
            s1.cell(row=1, column=i + 3).value = s[i]
            s1.cell(row=t + 2, column=i + 3).value = d[i]
    else:
        ry = 0
        for i in range(0, len(s)):

            if s[i] in sh3:
                # print(s[i],  d[i])
                s1.cell(row=1, column=ry + 3).value = s[i]
                s1.cell(row=t + 2, column=ry + 3).value = d[i]
                ry = ry + 1
    p = []
    r = []
    ro = []
    po = []
    e = []
    eo = []
    sheet_2 = wb.sheet_by_index(1)
    max_row = sheet_2.nrows
    for i in range(2, max_row):
        p.append(sheet_2.cell_value(i, 1))
        r.append(sheet_2.cell_value(i, 2))
        ro.append(sheet_2.cell_value(i, 3))
        po.append(sheet_2.cell_value(i, 4))
        e.append(sheet_2.cell_value(i, 5))
        eo.append(sheet_2.cell_value(i, 6))
    # print(p)
    for i in range(0, len(p)):
        if (p[i] == 'Discount Deduction'):
            discount = po[i]
            s1.cell(row=t + 2, column=20).value = discount
        # s1.cell(row=t+2, column=12).value =float(s1.cell(row=t+2, column=12).value)-float(discount)
        if (p[i] == 'Co-Payment'):
            CoPayment = po[i]
            s1.cell(row=t + 2, column=21).value = CoPayment
        # s1.cell(row=t+2, column=12).value =float(s1.cell(row=t+2, column=12)).value-float(CoPayment)
        row_num = s2.max_row
        wq += 1
        s2.cell(row=row_num + 1, column=1).value = wq
        s2.cell(row=row_num + 1, column=2).value = ccn
        s2.cell(row=row_num + 1, column=3).value = p[i]
        s2.cell(row=row_num + 1, column=4).value = r[i]
        s2.cell(row=row_num + 1, column=5).value = ro[i]
        s2.cell(row=row_num + 1, column=6).value = po[i]
        s2.cell(row=row_num + 1, column=7).value = e[i]
        s2.cell(row=row_num + 1, column=8).value = eo[i]
    with open(pdfpath, "rb") as f:
        pdf = pdftotext.PDF(f)

    with open('temp_files/output.txt', 'w') as f:
        f.write(" ".join(pdf))
    with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
        f = myfile.read()

    hg = []
    w = f.find('Policy Number') + 13
    g = f[w:]
    u = g.find('Reference/UTR No.') + w
    hg.append(f[w:u])

    w1 = f.find('Reference/UTR No.') + 17
    g = f[w1:]
    u1 = g.find('\n') + w1
    hg.append(f[w1:u1])

    if f.find('Payment Date') != -1:
        w2 = f.find('Payment Date') + 13
        g = f[w2:]
        u2 = g.find('\n') + w2
        hg.append(f[w2:u2])
    else:
        w2 = f.find('Date') + 6
        g = f[w2:]
        u2 = g.find('\n') + w2
        hg.append(f[w2:u2])

    w9 = f.find('ID Card Number :') + 17
    g = f[w9:]
    u9 = g.find('\n') + w9
    hg.append(f[w9:u9])

    hg = [sub.replace('  ', '') for sub in hg]
    hg = [sub.replace(':', '') for sub in hg]
    hg = [sub.replace('\n', ' ') for sub in hg]
    s1.cell(row=1, column=1).value = 'Sr. No.'
    s1.cell(row=1, column=2).value = 'claim number'
    s1.cell(row=t + 2, column=1).value = t + 1
    s1.cell(row=t + 2, column=2).value = ccn
    for i in range(0, len(hg)):
        s1.cell(row=t + 2, column=i + 16).value = hg[i]
    print("Done")
    wbk.save(wbkName)
    wbk.close()
    subprocess.run(["python", "make_master.py", 'fgh', op, '', wbkName])
    move_master_to_master_insurer('')
    print(f'processed {wbkName}')

except:
    log_exceptions()
    pass