import subprocess

import openpyxl
import sys
import camelot
import pdftotext
import xlrd
from make_log import log_exceptions
from backend import mark_flag
from movemaster import move_master_to_master_insurer

try:
    pdfpath = sys.argv[1]
    hosp_name = ''

    with open(pdfpath, "rb") as f:
        pdf = pdftotext.PDF(f)
    with open('temp_files/output.txt', 'w', encoding='utf-8') as f:
        f.write(" ".join(pdf))
    with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
        f = myfile.read()
    if 'Process Sheet' not in f:
        sys.exit(f'{pdfpath} wrong pdf recieved, so not processed')
    else:
        if 'Balaji Medical' in f:
            op = 'Tpappg@maxhealthcare.com May@2020 outlook.office365.com Max PPT'
            hosp_name = 'Max'
        else:
            op = 'mediclaim@inamdarhospital.org Mediclaim@2019 imap.gmail.com inamdar hospital'
            hosp_name = 'inamdar'

    wbkName = 'temp_files/' + 'east_west' + hosp_name + '.xlsx'
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    sh1 = ['Sr No.', 'CCN', 'Name', 'Hospital Address', 'PPN / Non PPN', 'Claim Type', 'Corporate', 'Diagnosis',
           'Policy No.', 'Card No.', 'Disease Code', 'Date Of Admission', 'Date Of Discharge', 'Intimation Date',
           'Sum Insured', 'Domicilary', 'Balance']
    sh2 = ['Sr No.', 'Claim ID', 'category', 'Billed Amt(Rs)', 'Approved Amt(Rs)', 'Deduction Amt(Rs)',
           'Reason of Deduction (If any)']

    for i in range(0, len(sh1)):
        s1.cell(row=1, column=i + 1).value = sh1[i]
    for i in range(0, len(sh2)):
        s2.cell(row=1, column=i + 1).value = sh2[i]
    tables = camelot.read_pdf(pdfpath, pages='all')
    tables.export('temp_files/foo1.xls', f='excel')
    loc = ("temp_files/foo1.xls")
    wb = xlrd.open_workbook(loc)

    jh = []
    gh = []
    h = []
    g = []
    hj = []
    hg = []
    sheet_1 = wb.sheet_by_index(0)
    for k in range(1, len(sh1)):
        c = 0
        for j in [1, 5]:
            for i in range(1, sheet_1.nrows):
                if (sheet_1.cell_value(i, j) == sh1[k]):
                    if (j == 1):
                        c = 1
                        hg.append(sheet_1.cell_value(i, 3))

                    if (j == 5):
                        c = 1
                        hg.append(sheet_1.cell_value(i, 8))
        if c == 0:
            hg.append(' ')
    t = 0
    s1.cell(row=t + 2, column=1).value = t + 1
    x1 = hg[0].find('\n')
    hg[0] = hg[0][:x1]

    temp = 0
    for j in range(0, tables.n):
        sheet_n = wb.sheet_by_index(j)
        sheet_n.cell_value(0, 0)
        for i in range(2, sheet_n.nrows):
            if sheet_n.cell_value(i, 1) == 'Total':
                hg.append(sheet_n.cell_value(i, 2))
                hg.append(sheet_n.cell_value(i, 4))
                hg.append(sheet_n.cell_value(i, 6))
            if temp == 1 and sheet_n.cell_value(i, 1) != 'Net Amt. Paid':
                jh.append(sheet_n.cell_value(i, 1))
                gh.append(sheet_n.cell_value(i, 2))
                h.append(sheet_n.cell_value(i, 4))
                g.append(sheet_n.cell_value(i, 6))
                # hj.append(sheet_n.cell_value(i, 7))
                hj.append('')
            if temp == 0:
                if (sheet_n.cell_value(i, 1) == ''):
                    # print(i,sheet_n.cell_value(i,1))
                    temp = 1
            if sheet_n.cell_value(i, 1) == 'Net Amt. Paid':
                break
    jh = [s.replace('\n', ' ') for s in jh]
    hj = [s.replace('\n', ' ') for s in hj]
    hg = [s.replace('\n', ' ') for s in hg]
    # print(hg)
    for i in range(0, len(hg)):
        s1.cell(row=t + 2, column=i + 2).value = hg[i]
    wq = 0
    for i in range(0, len(jh)):
        wq += 1
        row_num = s2.max_row
        s2.cell(row=row_num + 1, column=1).value = wq
        s2.cell(row=row_num + 1, column=2).value = hg[0]
        s2.cell(row=row_num + 1, column=3).value = jh[i]
        s2.cell(row=row_num + 1, column=4).value = gh[i]
        s2.cell(row=row_num + 1, column=5).value = h[i]
        s2.cell(row=row_num + 1, column=6).value = g[i]
        s2.cell(row=row_num + 1, column=7).value = hj[i]
    print("Done")
    wbk.save(wbkName)
    wbk.close()
    subprocess.run(["python", "make_master.py", 'east_west', op, '', wbkName])
    move_master_to_master_insurer('')
    mark_flag('X', sys.argv[1])
    print(f'processed {wbkName}')
except SystemExit as e:
    v = e.code
    if 'exit' in v:
        a =1
        os._exit(0)
except:
    log_exceptions()
    pass
