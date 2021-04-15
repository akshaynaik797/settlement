import sys
import re

import openpyxl
import pdftotext

from backend import mark_flag
from make_log import log_exceptions
from movemaster import move_master_to_master_insurer
try:
    _, file_path, mid = sys.argv

    with open(file_path, "rb") as f:
        pdf = pdftotext.PDF(f)
    with open('temp_files/output.txt', 'w', encoding='utf-8') as f:
        f.write(" ".join(pdf))
    with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
        f = myfile.read()

    insurer = re.compile(r"(?<=as instructed by).*").search(f)
    if insurer is not None:
        insurer = insurer.group().strip()
    else:
        insurer = ""
    ins_list = (
        ('chola', 'CHOLAMANDALAM'),
        ('bajaj', 'BAJAJ'),
        ('Max_Bupa', 'MAX BUPA')
    )
    for i, j in ins_list:
        if j in insurer:
            insurer = i
            break

    sh1 = ['Sno', 'HospitalID', 'InsurerID', 'ALNO', 'ClaimNo', 'MemberID', 'PolicyNo', 'PatientName', 'InsuranceCompany',
           'AccountNo', 'BeneficiaryBank Name', 'Diagnosis', 'UTRNo', 'BilledAmount', 'SettledAmount', 'TDS', 'NetPayable',
           'DiscountAmt', 'COPay', 'PolicyHolder', 'IPNo', 'PrimaryBeneficiary', 'EmployeeID', 'InsurerClaimNo',
           'InsurerMemberID', 'TaxDeductedatSource', 'Netamount payment', 'PaidbythePatient', 'ProrataBasis',
           'PolicyExcessDeductible', 'BeneficiaryName', 'BalanceSumInsuredBeforeClaim', 'NetPayable',
           'BalanceSumInsuredAfterClaim', 'TDS%', 'Remarks', 'PaymentTo', 'DateofAdmission', 'DateofDischarge',
           'AmtPaidtoHospital', 'BillAmt', 'PayableAmt', 'SettledAmt', 'SumInsured', 'ALAmount	Approved', 'Amount',
           'HospitalAmount', 'AmountUtilised', 'FinalAmountSettled', 'DateOfPayment', 'ServiceTax', 'TotalwithServiceTax',
           'InsuredPerson', 'CorporateName', 'DeductibleAmt', 'Transactiondate', 'LOCALAmount', 'ChequeDate',
           'UHCApprovedHospitalAmt', 'InsurerApprovedHospitalAmt', 'InsurerApprovedEmployeeAmt', 'PayableAmount',
           'NEFTTransactionNumber', 'TransactionDate', 'CorporateName', 'Claimed', 'PreHospitalisationPayableAmount',
           'PostHospitalisationPayableAmount', 'AddonBenefit', 'Claimed', 'Paid', 'BillAmount', 'PayableAmount(INR)',
           'BillDate', 'BillNo', 'AmountSettled', 'ApprovedAmount', 'less', 'Excess of Defined / Ailment Limit',
           'policy deduction', 'Limit exceed deduction', 'non payable deduction', 'Bill deduction', 'Other deduction']

    sh2 = ['Sr. No.', 'HospitalID', 'InsurerID', 'Claim ID', 'Details', 'Bill amount', 'Payable Amount', 'Deducted Amt',
           'Reason for Deduction', 'Discount']

    data_dict = {'Sno': mid, 'HospitalID': 'inamdar', 'InsurerID': insurer}
    regex_dict = {
        'ALNO': [[r"(?<=Claim No).*", r"(?<=Payment Details).*"], [':', 'Claim No']],
        'ClaimNo': [[r"(?<=Claim No).*", r"(?<=Payment Details).*"], [':', 'Claim No']],
        'PolicyNo': [[r"(?<=Policy No :).*"], [':']],
        'UTRNo': [[r"(?<=UTR Reference).*"], []],
        'DiscountAmt': [[r"(?<=Discount).*"], []],
        'NetPayable': [[r".*(?=as instructed by)"], [':', 'Rs', ',']],
        'DateOfPayment': [[r"(?<=We have on).*(?=made)"], []],
        'TDS': [[r"(?<=TDS Amount).*", r"(?<=TDS)\s*\w+", r"(?<=TDS).*(?=\/)"], [':', 'INR']],
        'BilledAmount': [[r"(?<=Bill Amount).*", r"(?<=GROSS AMOUNT)\s*\S+", r"(?<=Billed Amount)\s*\w+"], [':', 'INR']],
    }

    match_dict = {
        "Diagnosis": r"^\w+(?: \w+)*$",
        "BilledAmount": r"^\d+$",
        "TDS": r"^[.\d]+$",
        "DiscountAmt": r"^[.\d]+$"
    }
    for i in regex_dict:
        temp = ""
        for reg in regex_dict[i][0]:
            temp = re.compile(reg).search(f)
            if temp is not None:
                temp = temp.group().strip()
                for j in regex_dict[i][1]:
                    temp = temp.replace(j, '')
                temp = temp.strip()
                if i in match_dict:
                    temp = re.compile(match_dict[i]).match(temp)
                    if temp is not None:
                        temp = temp.group().strip()
                        break
                    temp = ""
            else:
                temp = ""
        data_dict[i] = temp

    wbName = 'master.xlsx'
    wb = openpyxl.Workbook()
    wb.create_sheet('Sheet1')
    wb.create_sheet('count')
    wb.create_sheet('count_star')
    wb.create_sheet('error_sheet')
    main_s1 = wb.worksheets[0]
    main_s2 = wb.worksheets[1]

    for i in range(0, len(sh1)):
        # main_s1.cell(row=1, column=i+1).value=i+1
        main_s1.cell(row=1, column=i + 1).value = sh1[i]
        if sh1[i] in data_dict:
            main_s1.cell(row=2, column=i + 1).value = data_dict[sh1[i]]

    for i in range(0, len(sh2)):
        main_s2.cell(row=1, column=i + 1).value = sh2[i]


    wb.save(wbName)
    move_master_to_master_insurer(sys.argv[2], pdfpath=file_path)
    mark_flag('X', sys.argv[2])
    print(f'processed')

    pass
except:
    log_exceptions()
    pass