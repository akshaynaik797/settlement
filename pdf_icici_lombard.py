import os
import subprocess
import sys

import openpyxl
import pdftotext
import xlrd

from make_log import log_exceptions
from movemaster import move_master_to_master_insurer

try:
    pdfpath = sys.argv[1]
    onlyfiles = [os.path.split(pdfpath)[1]]
    mypath = os.path.dirname(pdfpath)+'/'

    hosp_name = ''

    with open(pdfpath, "rb") as f:
        pdf = pdftotext.PDF(f)
    with open('temp_files/output.txt', 'w') as f:
        f.write(" ".join(pdf))
    with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
        f = myfile.read()
    if 'wanted_text' not in f:
        sys.exit(f'{pdfpath} wrong pdf recieved, so not processed')
    else:
        if 'Balaji Medical' in f:
            op = 'Tpappg@maxhealthcare.com May@2020 outlook.office365.com Max PPT'
            hosp_name = 'Max'
        else:
            op = 'mediclaim@inamdarhospital.org Mediclaim@2019 imap.gmail.com inamdar hospital'
            hosp_name = 'inamdar'
    ###########################################################
    wbkName = 'temp_files/' + 'icici_lombard' + hosp_name + '.xlsx'
    t, wq =0, 0
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    # wbk.create_sheet('Sheet3')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    yu = []
    # print(onlyfiles)
    for i in onlyfiles:
        if i[-1] == '#':
            yu.append(t)
        t += 1
    # print(yu)
    t = 0
    for i in yu:
        onlyfiles.pop(i - t)
        t = t + 1
    # print(pdffiles)
    for t in range(0, len(onlyfiles)):
        with open(pdfpath + onlyfiles[t], "rb") as f:
            pdf = pdftotext.PDF(f)

        with open('temp_files/output.txt', 'w') as f:
            f.write(" ".join(pdf))
        with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
            f = myfile.read()
        eo = []
        w = f.find('amount of Rs.') + 13
        g = f[w:]
        u = u2 = g.find('.') + w
        eo.append(f[w:u])

        w1 = f.find('ref.') + 3
        g = f[w1:]
        x1 = g.find('no.') + 3 + w1
        u1 = g.find('dated') + w1
        eo.append(f[x1:u1])

        w = f.find('dated') + 5
        g = f[w:]
        u = u2 = g.find('towards') + w - 1
        eo.append(f[w:u])

        w1 = f.find('TDS is') + 7
        g = f[w1:]
        u1 = g.find('.') + w1
        eo.append(f[w1:u1])
        # print(eo)
        for i in range(0, len(eo)):
            s1.cell(row=t + 2, column=15 + i).value = eo[i]
    onlyfiles.sort()
    print(onlyfiles)
    for t in range(0, len(onlyfiles)):
        def select_sheet(wks, sheet_name):

            if not sheet_name in wks.sheetnames:
                wks.create_sheet(sheet_name)
            # print(sheet_name)

            wks.save('temp_files/icici_lombard.xlsx')


        def select_column(wks, s, ro):
            sheet = wks.worksheets[s]

            max_col = sheet.max_column
            s = []
            for i in range(1, max_col + 1):
                cell_obj = sheet.cell(row=1, column=i)
                s.append(cell_obj.value)


        # print(s)
        loc = (mypath + onlyfiles[t])
        wb = xlrd.open_workbook(loc)
        wb1 = openpyxl.load_workbook(loc)
        sh1 = ['Claim No', 'UHID NO', 'Name of the Patient', 'Policy Name', 'Requested Amount', 'Final Amount Settled',
               'Diagnosis', 'Date of Admission', 'Date Of Discharge', 'CO-PAYMENT AMOUNT', 'DISALLOWED AMOUNT',
               'DISALLOWED REASONS', 'settled amount', 'cheque/EFT vide ref. no.', 'date of payment', 'TDS']
        for i in range(0, len(sh1)):
            s1.cell(row=1, column=i + 3).value = sh1[i]
        sh2 = ['Charges Details', 'Claimed', 'Deductions', 'Paid', 'Reason for Deductions']
        for i in range(0, len(sh2)):
            s2.cell(row=1, column=i + 3).value = sh2[i]
        sr2 = wb1.worksheets[2]
        sheet_4 = wb.sheet_by_index(2)
        sheet_4.cell_value(0, 0)
        jf = []
        kl = []
        q = []
        w = []
        # row_num=sheet_4.nrows
        # column_num=sheet_4.ncols
        for i in range(1, sheet_4.nrows):
            jf.append(str(sheet_4.cell_value(i, 1)))
            kl.append(str(sheet_4.cell_value(i, 2)))
            q.append(str(sheet_4.cell_value(i, 3)))
            w.append(str(sheet_4.cell_value(i, 4)))
        # print(jf)
        res = [sub.replace('Â\xa0', '') for sub in jf]
        res1 = [sub.replace('Â', '') for sub in res]
        res5 = [sub.replace('Â\xa0', '') for sub in kl]
        res2 = [sub.replace('Â', '') for sub in res5]
        res6 = [sub.replace('Â\xa0', '') for sub in q]
        res3 = [sub.replace('Â', '') for sub in res6]
        res7 = [sub.replace('Â\xa0', '') for sub in w]
        res4 = [sub.replace('Â', '') for sub in res7]
        # print(res1,t)
        for i in range(0, len(res1)):
            sr2.cell(row=i + 2, column=2).value = res1[i]
            sr2.cell(row=i + 2, column=3).value = res2[i]
            sr2.cell(row=i + 2, column=4).value = res3[i]
            sr2.cell(row=i + 2, column=5).value = res4[i]
        # print(t,i)
        sr2 = wb1.worksheets[4]
        sheet_4 = wb.sheet_by_index(4)
        sheet_4.cell_value(0, 0)
        jf = []
        kl = []
        q = []
        w = []
        # row_num=sheet_4.nrows
        # column_num=sheet_4.ncols
        for i in range(1, sheet_4.nrows):
            jf.append(str(sheet_4.cell_value(i, 1)))
            kl.append(str(sheet_4.cell_value(i, 2)))
            q.append(str(sheet_4.cell_value(i, 3)))
            w.append(str(sheet_4.cell_value(i, 4)))
        # print(jf)
        res = [sub.replace('Â\xa0', '') for sub in jf]
        res1 = [sub.replace('Â', '') for sub in res]
        res5 = [sub.replace('Â\xa0', '') for sub in kl]
        res2 = [sub.replace('Â', '') for sub in res5]
        res6 = [sub.replace('Â\xa0', '') for sub in q]
        res3 = [sub.replace('Â', '') for sub in res6]
        res7 = [sub.replace('Â\xa0', '') for sub in w]
        res4 = [sub.replace('Â', '') for sub in res7]
        # print(res1,t)
        for i in range(0, len(res1)):
            sr2.cell(row=i + 2, column=2).value = res1[i]
            sr2.cell(row=i + 2, column=3).value = res2[i]
            sr2.cell(row=i + 2, column=4).value = res3[i]
            sr2.cell(row=i + 2, column=5).value = res4[i]
        # print(t,i)
        wb1.save(loc)
        wb = xlrd.open_workbook(loc)
        sheet_2 = wb.sheet_by_index(2)
        sheet_2.cell_value(0, 0)
        sheet_3 = wb.sheet_by_index(3)
        sheet_3.cell_value(0, 0)
        sheet_4 = wb.sheet_by_index(4)
        sheet_4.cell_value(0, 0)
        jf = []
        jf.append(sheet_2.cell_value(1, 2))
        jf.append(sheet_2.cell_value(2, 2))
        jf.append(sheet_2.cell_value(3, 2))
        jf.append(sheet_2.cell_value(8, 2))
        ccn = sheet_2.cell_value(1, 4)
        jf.append(sheet_2.cell_value(5, 4))
        jf.append(sheet_4.cell_value(2, 4))
        jf.append(sheet_4.cell_value(1, 4))
        jf.append(sheet_2.cell_value(7, 2))
        jf.append(sheet_2.cell_value(7, 4))
        # print(jf)

        m = []
        gh = []
        hg = []
        r = []
        rt = []
        for i in range(2, sheet_3.nrows):
            m.append(sheet_3.cell_value(i, 2))
            gh.append(sheet_3.cell_value(i, 3))
            hg.append(sheet_3.cell_value(i, 4))
            r.append(sheet_3.cell_value(i, 7))
            rt.append(sheet_3.cell_value(i, 8))
        hg = [str(sub).replace('!  td>', '') for sub in hg]
        for wd in wbk.worksheets[:1]:
            wd.cell(row=1, column=1).value = 'Sr. No.'
            wd.cell(row=1, column=2).value = 'AL NO'
            wd.cell(row=t + 2, column=1).value = t + 1
            wd.cell(row=t + 2, column=2).value = ccn
        s2.cell(row=1, column=1).value = 'Sr. No.'
        s2.cell(row=1, column=2).value = 'claim NO'
        for i in range(0, len(m)):
            row_num = s2.max_row
            s2.cell(row=row_num + 1, column=1).value = t + 1
            s2.cell(row=row_num + 1, column=2).value = jf[0]
            s2.cell(row=row_num + 1, column=3).value = m[i]
            s2.cell(row=row_num + 1, column=4).value = gh[i]
            s2.cell(row=row_num + 1, column=5).value = hg[i]
            s2.cell(row=row_num + 1, column=6).value = r[i]
            s2.cell(row=row_num + 1, column=7).value = rt[i]
        sheet_3 = wb.sheet_by_index(1)
        sheet_3.cell_value(0, 0)
        gh = []
        hg = []
        r = []
        rt = []
        for i in range(0, len(jf)):
            s1.cell(row=t + 2, column=i + 3).value = jf[i]
        # for i in range(2,sheet_3.nrows):
        hg.append(sheet_3.cell_value(2, 5))
        r.append(sheet_3.cell_value(2, 6))
        rt.append(sheet_3.cell_value(2, 7))
        sheet_1 = wb.sheet_by_index(0)
        sheet_1.cell_value(0, 0)
        e_id = sheet_1.cell_value(2, 3)
        e_name = sheet_1.cell_value(2, 4)
        for i in range(0, len(hg)):
            s1.cell(row=t + 2, column=len(jf) + 3).value = hg[i]
            s1.cell(row=t + 2, column=len(jf) + 4).value = r[i]
            s1.cell(row=t + 2, column=len(jf) + 5).value = rt[i]
        s1.cell(row=t + 2, column=19).value = e_id
        s1.cell(row=t + 2, column=20).value = e_name


    print("Done")
    wbk.save(wbkName)
    wbk.close()
    subprocess.run(["python", "make_master.py", 'icici_lombard', op, '', wbkName])
    ###########################################################
    move_master_to_master_insurer('')
    print(f'processed {wbkName}')

except:
    log_exceptions()
    pass