from os import path
from pathlib import Path
from shutil import copyfile

import openpyxl
import pandas as pd
import subprocess
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from make_log import log_exceptions


def move_master_to_master_insurer(mail_uid, **kwargs):
    try:
        source, dest = 'master.xlsx','master_insurer.xlsx'
        if 'pdfpath' in kwargs:
            wb = openpyxl.open(source)
            worksheet = wb.active
            temp = []
            temp.extend([worksheet.cell(row=2, column=5).value, worksheet.cell(row=2, column=4).value])
            wb.close()
            f_dst = "../index/Attachments/"
            Path(f_dst).mkdir(parents=True, exist_ok=True)
            r_flag = 0
            for i in temp:
                if i is not None:
                    i = i.strip()
                    f_dst = path.join(f_dst, f"{i}.pdf")
                    copyfile(kwargs['pdfpath'], f_dst)
                    r_flag = 1
                    break
            if r_flag == 0:
                copyfile(kwargs['pdfpath'], path.join(f_dst, path.split(kwargs['pdfpath'])[-1]))
        #subprocess.run(["python", "updation.py","1","max","11",'Yes'])
        if not path.exists(dest):
            wb = Workbook()
            ws1 = wb.create_sheet("Sheet1", 1)
            ws2 = wb.create_sheet("count", 2)
            ws3 = wb.create_sheet("count_star", 3)
            ws4 = wb.create_sheet("error_sheet", 4)
            wb.save(dest)

            book = load_workbook(dest)
            sheetlist = book.get_sheet_names()

            for i in sheetlist:
                excel_data_df = pd.read_excel(source, sheet_name=i)
                sheet = book.get_sheet_by_name(i)
                for r in dataframe_to_rows(excel_data_df, index=False, header=True):
                    sheet.append(r)
                book.save(dest)

        elif path.exists(source):
            book = load_workbook(filename=dest)
            sheetlist = book.get_sheet_names()

            for i in sheetlist:
                excel_data_df = pd.read_excel(source, sheet_name=i)
                sheet = book.get_sheet_by_name(i)
                for r in dataframe_to_rows(excel_data_df, index=False, header=False):
                    sheet.append(r)
                book.save(dest)
        return True
    except:
        log_exceptions()
        return False
if __name__ == '__main__':
    move_master_to_master_insurer('', pdfpath='/home/akshay/temp/19429253_.pdf')