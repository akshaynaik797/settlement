import email
import imaplib
import os
import re
import subprocess
import PyPDF2
import html2text
import openpyxl
import sys
import camelot
import pdftotext
import xlrd
from make_log import log_exceptions
from backend import mark_flag
from movemaster import move_master_to_master_insurer

try:
    pdfpath = sys.argv[1]
    onlyfiles = [os.path.split(pdfpath)[1]]
    mypath = os.path.dirname(pdfpath)+'/'

    hosp_name = ''

    with open(pdfpath, "rb") as f:
        pdf = pdftotext.PDF(f)
    with open('temp_files/output.txt', 'w', encoding='utf-8') as f:
        f.write(" ".join(pdf))
    with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
        f = myfile.read()
    if 'Balaji Medical' in f:
        op = 'Tpappg@maxhealthcare.com May@2020 outlook.office365.com Max PPT'
        hosp_name = 'Max'
    else:
        op = 'mediclaim@inamdarhospital.org Mediclaim@2019 imap.gmail.com inamdar hospital'
        hosp_name = 'inamdar'
    ###########################################################
    wbkName = 'temp_files/' + 'fhpl' + hosp_name + '.xlsx'
    t, wq,f =0, 0,0
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    wbk.create_sheet('Sheet3')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    s3 = wbk.worksheets[2]
    for t in range(0, len(onlyfiles)):
        sh1 = ['Sr No.', 'Claim ID', 'Patient Name', 'Policy No', 'Employee ID', 'Diagnosis', 'Card No.',
               'Date of Admission', 'Date of Discharge', 'NEFT transaction number', 'transaction date',
               'Corporate Name', 'insurance company', 'alno']
        sh2 = ['Sr No.', 'Claim ID', 'Claimed', 'Billed amount', 'Discount', 'Disallowed', 'Settled amount', 'Less TDS',
               'Net Paid Amount']
        sh3 = ['Sr No.', 'Claim ID', 'category', 'Disallowance amount', 'Disallowance Reasons']

        for i in range(0, len(sh1)):
            s1.cell(row=1, column=i + 1).value = sh1[i]
        for i in range(0, len(sh2)):
            s2.cell(row=1, column=i + 1).value = sh2[i]
        for i in range(0, len(sh3)):
            s3.cell(row=1, column=i + 1).value = sh3[i]

        # html = open('temp_files/attachments_' + hosp_name + str(t + 1) + ".html")
        # f = str(html.read())
        # w = open("temp_files/out.txt", "w")
        # w.write(html2text.html2text(f))
        # html.close()
        # w.close()
        # with open('temp_files/out.txt', 'r') as myfile:
        #     f = myfile.read()
        if 'Claim Settled' not in f:
            gh = []
            w2 = f.find('Claim Of')
            g = f[w2:]
            w4 = g.find(':') + w2 + 2
            g = f[w4:]
            u2 = g.find('\n') + w4
            gh.append(f[w4:u2])

            w = f.find('Policy No')
            g = f[w:]
            w5 = g.find(':') + w + 2
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])

            w = f.find('Employee ID')
            g = f[w:]
            w5 = g.find(':') + w + 2
            g = f[w5:]
            u3 = g.find('\n') + w5
            gh.append(f[w5:u3])

            w3 = f.find('Diagnosis')
            g = f[w3:]
            w1 = g.find(':') + w3 + 2
            g = f[w1:]
            u1 = g.find('\n') + w1
            # print(w1,u1)
            gh.append(f[w1:u1])

            w3 = f.find('Card No.')
            g = f[w3:]
            w1 = g.find(':') + w3 + 2
            g = f[w1:]
            u1 = g.find('\n') + w1
            gh.append(f[w1:u1])

            w3 = f.find('Date of Admission')
            g = f[w3:]
            w1 = g.find(':') + w3 + 2
            g = f[w1:]
            u1 = g.find('\n') + w1
            gh.append(f[w1:u1])

            w3 = f.find('Date of Discharge')
            g = f[w3:]
            w1 = g.find(':') + w3 + 2
            g = f[w1:]
            u1 = g.find('\n') + w1
            gh.append(f[w1:u1])
            # print(w1,u1)

            w3 = f.find('number') + 7
            g = f[w3:]
            w1 = g.find('dated') + w3
            gh.append(f[w3:w1])
            g = f[w1:]
            u1 = g.find('.') + w1
            gh.append(f[w1 + 5:w1 + 19])

            w3 = f.find('Corporate Name')
            g = f[w3:]
            w1 = g.find(':') + w3 + 2
            g = f[w1:]
            u1 = g.find('\n') + w1
            gh.append(f[w1:u1])

            w3 = f.find('Good wishes from') + 17
            g = f[w3:]
            u1 = g.find('**') + w3
            gh.append(f[w3:u1])

            w3 = f.find('Claim ID')
            g = f[w3:]
            w1 = g.find(':') + w3 + 2
            g = f[w1:]
            u1 = g.find('\n') + w1
            ccn = f[w1:u1]
            gh = [sub.replace('\n', '') for sub in gh]
            # print(ccn)
            s1.cell(row=t + 2, column=1).value = t + 1
            s1.cell(row=t + 2, column=2).value = ccn
            gh = [sub.replace('**', '') for sub in gh]
            gh = [sub.replace('\n', '') for sub in gh]
            for i in range(0, len(gh)):
                s1.cell(row=t + 2, column=i + 3).value = gh[i]
            # print(gh[4])
            subj = None
            if subj == None:
                subj = ccn
            s1.cell(row=t + 2, column=2).value = ccn
            s1.cell(row=t + 2, column=14).value = subj
            # print(subj)
            gh = []
            w2 = f.find('Claimed')
            g = f[w2:]
            w4 = g.find('Rs.') + w2 + 4
            g = f[w4:]

            u2 = g.find('/-') + w4
            gh.append(f[w4:u2])

            w = f.find('Billed')
            g = f[w:]
            w5 = g.find(':') + w
            u3 = g.find('/-') + w
            gh.append(f[w5:u3])

            w = f.find('Discount')
            g = f[w:]
            w5 = g.find('Rs.') + w
            g = f[w5:]
            u3 = g.find('/-') + w5
            gh.append(f[w5:u3])

            w3 = f.find('Disallowed')
            g = f[w3:]
            w1 = g.find(':') + w3
            u1 = g.find('/-') + w3
            gh.append(f[w1:u1])

            w3 = f.find('Settled')
            g = f[w3:]
            w1 = g.find(':') + w3 + 2
            u1 = g.find('/-') + w3
            gh.append(f[w1:u1])

            w3 = f.find('Less TDS')
            g = f[w3:]
            w1 = g.find('Rs.') + w3
            u1 = g.find('/-') + w3
            gh.append(f[w1:u1])

            w3 = f.find('Net Paid Amount')
            g = f[w3:]
            w1 = g.find(':') + w3 + 2
            u1 = g.find('/-') + w3
            gh.append(f[w1:u1])
            s2.cell(row=t + 2, column=1).value = t + 1
            s2.cell(row=t + 2, column=2).value = ccn
            gh = [sub.replace('Rs.', '') for sub in gh]
            gh = [sub.replace(':', '') for sub in gh]
            gh = [sub.replace(' ', '') for sub in gh]
            for i in range(0, len(gh)):
                s2.cell(row=t + 2, column=2).value = ccn
                s2.cell(row=t + 2, column=i + 3).value = gh[i]
            # print(gh) Rs.

            w3 = f.find('Disallowance Reasons')
            g = f[w3:]
            w1 = g.find('Claimed') + w3
            g = f[w3:w1]
            # mylist = [item for item in g.split('\n')]
            mylist = [item for item in g.split('Rs.')]
            # mylist.remove('')
            # print(mylist)
            x = ''
            hj = 0
            for i in mylist:
                x = i

                if (hj == 0):
                    if x.find('_') != -1:
                        m = x.find('_')
                        l = x[m + 1:].find('_') + m
                        sd = x[m + 1:l]
                    # print(x)
                    hj = 1
                    continue

                else:
                    row_num = s3.max_row + 1
                    if (x != ''):
                        wq += 1
                        m = i.find('\n')
                        i = i[:m]
                        m = i.find('.')
                        s = i[0:m + 3]
                        l = i[m + 3:]
                        s3.cell(row=row_num, column=1).value = wq
                        s3.cell(row=row_num, column=2).value = ccn
                        # s3.cell(row=row_num, column=3).value = sd
                        s3.cell(row=row_num, column=3).value = ''
                        s3.cell(row=row_num, column=4).value = s
                        s3.cell(row=row_num, column=5).value = l
                if x.find('_') != -1:
                    m = x.find('_')
                    l = x[m + 1:].find('_') + m
                    sd = x[m + 1:l]
                # print(x)
            ccn = ccn.replace(' ', '')
        else:
            data = dict()
            if f.find('Employee ID') != -1:
                regex = r'\S+(?= *\|\s*Claim ID)'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['empid'] = x1
                else:
                    data['empid'] = ''
            else:
                data['empid'] = ''

            if f.find('Diagnosis') != -1:
                regex = r'[ \w]+(?=\s*Date of Admission)'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['diagnosis'] = x1
                else:
                    data['diagnosis'] = ''
            else:
                data['diagnosis'] = ''

            if f.find('Claim ID') != -1:
                regex = r'(?<=Claim ID).*'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['claim_id'] = x1
                else:
                    data['claim_id'] = ''
            else:
                data['claim_id'] = ''

            if f.find('Claim Of') != -1:
                regex = r'(?<=Claim Of).*'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['patient_name'] = x1
                else:
                    data['patient_name'] = ''
            else:
                data['patient_name'] = ''

            if f.find('Policy No.') != -1:
                regex = r'\S+(?=\s*Card No.)'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['policy_no'] = x1
                else:
                    data['policy_no'] = ''
            else:
                data['policy_no'] = ''

            if f.find('Card No.') != -1:
                regex = r'\S+(?=\s*Payee Name)'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['card_no'] = x1
                else:
                    data['card_no'] = ''
            else:
                data['card_no'] = ''

            if f.find('Date of Admission') != -1:
                regex = r'(?<=Date of Admission).*(?=Date of Discharge)'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['doa'] = x1
                else:
                    data['doa'] = ''
            else:
                data['doa'] = ''

            if f.find('Date of Discharge') != -1:
                regex = r'( ?\S+){3}(?= *\|?\s*Relation)'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['dod'] = x1
                else:
                    data['dod'] = ''
            else:
                data['dod'] = ''

            if f.find('NEFT') != -1:
                regex = r'\d+(?=\r?\n* *dated)'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['neft'] = x1
                else:
                    data['neft'] = ''
            else:
                data['neft'] = ''

            if f.find('dated') != -1:
                regex = r'(?<=dated)\s*\S+'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['transaction_date'] = x1
                else:
                    data['transaction_date'] = ''
            else:
                data['transaction_date'] = ''

            if f.find('dated') != -1:
                regex = r'(?<=dated)\s*\S+'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['transaction_date'] = x1
                else:
                    data['transaction_date'] = ''
            else:
                data['transaction_date'] = ''

            if f.find('Corporate Name') != -1:
                regex = r'(?<=Corporate Name)\s*\S+[ \S]+(?=\r?\n)'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['corporate_name'] = x1
                else:
                    data['corporate_name'] = ''
            else:
                data['corporate_name'] = ''

            if f.find('instructions of') != -1:
                regex = r'(?<=instructions of)[ \S]+(?=your)'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['insurance_company'] = x1
                else:
                    data['insurance_company'] = ''
            else:
                data['insurance_company'] = ''

            data['alno'] = data['claim_id']

            if f.find('Claimed') != -1:
                regex = r'\S+(?=/-\s*---\|---)'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['claimed_amount'] = x1
                else:
                    data['claimed_amount'] = ''
            else:
                data['claimed_amount'] = ''

            if f.find('Billed') != -1:
                regex = r'\S+(?=/-\s*Discount)'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['billed_amount'] = x1
                else:
                    data['billed_amount'] = ''
            else:
                data['billed_amount'] = ''

            if f.find('Discount') != -1:
                regex = r'\S+(?=/-\s*Disallowed)'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['discount_amount'] = x1
                else:
                    data['discount_amount'] = ''
            else:
                data['discount_amount'] = ''

            if f.find('Disallowed') != -1:
                regex = r'(?<=Rs.)\S+(?=/-\s*Settled)'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['disallowed_amount'] = x1
                else:
                    data['disallowed_amount'] = ''
            else:
                data['disallowed_amount'] = ''

            if f.find('Settled') != -1:
                regex = r'\S+(?=/-\s*Less TDS)'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['settled_amount'] = x1
                else:
                    data['settled_amount'] = ''
            else:
                data['settled_amount'] = ''

            if f.find('Less TDS') != -1:
                regex = r'\S+(?=/-\s*Net Paid Amount:)'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['less_tds'] = x1
                else:
                    data['less_tds'] = ''
            else:
                data['less_tds'] = ''

            if f.find('Net Paid Amount') != -1:
                regex = r'\S+(?=/-\s*Co-Payment Amount)'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip()
                    data['net_paid_amount'] = x1
                else:
                    data['net_paid_amount'] = ''
            else:
                data['net_paid_amount'] = ''

            chars = ['\n', '|', '-']
            for key, value in data.items():
                for i in chars:
                    value = value.replace(i, '')
                    data[key] = value.strip()

            # pprint(data)

            if f.find('Deduction Reason') != -1:
                regex = r'(?<=Deduction Reason\*\*)[\s\S]+(?=\*\*\s*Claimed)'
                x = re.search(regex, f)
                if x:
                    x1 = x.group().strip().replace('---|---', "")

            regex = '(?<=Rs.)\d+'
            amounts = re.findall(regex, x1, re.MULTILINE)

            regex = '[ \S]+(?=\|)'
            catogories = re.findall(regex, x1, re.MULTILINE)

            regex = '(?<=:-)[ \w]+'
            reasons = re.findall(regex, x1, re.MULTILINE)
            if len(reasons) < len(amounts):
                regex = '(?<=:-)[[\s\S][^,\n]*]*'
                reasons = re.findall(regex, x1, re.MULTILINE)
                reasons = [i.replace('\n', '') for i in reasons]



            table = []
            ccn = data['claim_id']
            for i, j in enumerate(amounts):
                try:
                    table.append((catogories[i], j, reasons[i]))
                except IndexError as e:
                    table.append((catogories[-1], j, reasons[i]))

            gh = [data['claim_id'], data['patient_name'], data['policy_no'], data['empid'], data['diagnosis'],
                  data['card_no'], data['doa'], data['dod'], data['neft'], data['transaction_date'],
                  data['corporate_name'], data['insurance_company'], data['alno'], ]
            s1.cell(row=t + 2, column=1).value = s1.max_row
            for i in range(0, len(gh)):
                s1.cell(row=t + 2, column=i + 2).value = gh[i]
            gh = [data['claim_id'], data['claimed_amount'], data['billed_amount'], data['discount_amount'],
                  data['discount_amount'], data['settled_amount'], data['less_tds'],  data['net_paid_amount']]
            s2.cell(row=t + 2, column=1).value = s2.max_row
            for i in range(0, len(gh)):
                s2.cell(row=t + 2, column=i + 2).value = gh[i]
            rowno = s3.max_row
            for i in table:
                rowno = s3.max_row + 1
                s3.cell(row=rowno, column=1).value = rowno-1
                s3.cell(row=rowno, column=2).value = data['claim_id']
                s3.cell(row=rowno, column=3).value = i[0]
                s3.cell(row=rowno, column=4).value = i[1]
                s3.cell(row=rowno, column=5).value = i[2]



    print("Done")
    wbk.save(wbkName)
    wbk.close()
    subprocess.run(["python", "make_master.py", 'fhpl', op, '', wbkName])
    ###########################################################
    move_master_to_master_insurer(sys.argv[2], pdfpath=pdfpath)
    mark_flag('X', sys.argv[2])
    print(f'processed {wbkName}')
except SystemExit as e:
    v = e.code
    if 'exit' in v:
        a =1
        os._exit(0)
except:
    log_exceptions()
    pass
