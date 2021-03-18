import os
import re
import subprocess
import PyPDF2
import openpyxl
import sys
import camelot
import pdftotext
import xlrd
from make_log import log_exceptions
from backend import mark_flag
from movemaster import move_master_to_master_insurer

try:
    pdfpath = sys.argv[1]
    onlyfiles = [os.path.split(pdfpath)[1]]
    mypath = os.path.dirname(pdfpath)+'/'

    hosp_name = ''

    with open(pdfpath, "rb") as f:
        pdf = pdftotext.PDF(f)
    with open('temp_files/output.txt', 'w', encoding='utf-8') as f:
        f.write(" ".join(pdf))
    with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
        f = myfile.read()
    if 'wanted_text' not in f:
        sys.exit(f'{pdfpath} wrong pdf recieved, so not processed')
    else:
        if 'Balaji Medical' in f:
            op = 'Tpappg@maxhealthcare.com May@2020 outlook.office365.com Max PPT'
            hosp_name = 'Max'
        else:
            op = 'mediclaim@inamdarhospital.org Mediclaim@2019 imap.gmail.com inamdar hospital'
            hosp_name = 'inamdar'
    ###########################################################
    wbkName = 'temp_files/' + 'aditya' + hosp_name + '.xlsx'
    t, wq =0, 0
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    wbk.create_sheet('Sheet3')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    s3 = wbk.worksheets[2]
    for t in range(0, len(onlyfiles)):
        tables = camelot.read_pdf(mypath+onlyfiles[t],pages = 'all', line_scale=100)
        tables.export('temp_files/foo1.xls', f='excel')
        loc = ("temp_files/foo1.xls")
        with open(mypath + onlyfiles[t], "rb") as f:
            pdf = pdftotext.PDF(f)

        with open('temp_files/output.txt', 'w', encoding='utf-8') as f:
            f.write(" ".join(pdf))
        with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
            f = myfile.read()
        f = f.replace('\n', '$$ ')
        # print(f)

        wb = xlrd.open_workbook(loc)
        sh1 = ['Sum Insured', 'Claimed amount (Rs.)', 'AL Amount(in case of cashless)', 'Approved amount (Rs.)',
               'Deduction amount (Rs.)', 'Hospital Amount', 'TDS', 'Discount Amount', 'Reason for deduction',
               'Amount Utilised', 'CoPay']
        for i in range(0, len(sh1)):
            s2.cell(row=1, column=i + 3).value = sh1[i]
        sh2 = ['Policy Number', 'Member id', 'Patient Name', 'Policy Holder', 'Payee Bank name', 'Payee account number',
               'Amount of transfer', 'UTR number', 'Diagnosis', 'doa', 'dod', 'Deduction amount', 'discount',
               'transaction date']
        for i in range(0, len(sh2)):
            s1.cell(row=1, column=i + 3).value = sh2[i]
        sh3 = ['Sr', 'CCN', 'Deduction amount', 'Deduction reason']
        for i in range(0, len(sh3)):
            s3.cell(row=1, column=i + 1).value = sh3[i]
        hg = []
        w = f.find('Policy Number') + 15
        g = f[w:]
        u = g.find('$$') + w
        hg.append(f[w:u])

        w1 = f.find('Member id') + 11
        g = f[w1:]
        u1 = g.find('$$') + w1
        hg.append(f[w1:u1])

        w2 = f.find('Patient Name') + 14
        g = f[w2:]
        u2 = g.find('$$') + w2
        hg.append(f[w2:u2])

        w3 = f.find('Policy Holder') + 14
        g = f[w3:]
        u3 = g.find('$$') + w3
        hg.append(f[w3:u3])

        w4 = f.find('Payee Bank name') + 17
        g = f[w4:]
        u4 = g.find('$$') + w4
        hg.append(f[w4:u4])

        w5 = f.find('Payee account number') + 22
        g = f[w5:]
        u5 = g.find('$$') + w5
        hg.append(f[w5:u5])

        w6 = f.find('Amount of transfer') + 25
        g = f[w6:]
        u6 = g.find('$$') + w6
        hg.append(f[w6:u6])

        w7 = f.find('UTR number') + 12
        g = f[w7:]
        u7 = g.find('$$') + w7
        hg.append(f[w7:u7])

        w8 = f.find('Ailment Name') + 14
        g = f[w8:]
        u8 = g.find('Please note') + w8
        hg.append(f[w8:u8])

        w8 = f.find('Date of Admission:') + 19
        g = f[w8:]
        u8 = g.find('$$') + w8
        hg.append(f[w8:u8])

        w8 = f.find('Date of Discharge:') + 19
        g = f[w8:]
        u8 = g.find('$$') + w8
        hg.append(f[w8:u8])

        w8 = f.find('Deduction amount') + 23
        g = f[w8:]
        u8 = g.find('$$') + w8
        hg.append(f[w8:u8])

        w8 = f.find('Date of transfer :') + 19
        g = f[w8:]
        u8 = g.find('$$') + w8
        hg.append(f[w8:u8])

        w9 = f.find('Claim registration number') + 26
        g = f[w9:]
        u9 = g.find('$$') + w9
        ccn = (f[w9:u9])
        hg = [sub.replace('  ', '') for sub in hg]
        hg = [sub.replace('$$ ', '') for sub in hg]
        # print(hg)
        transaction_date = hg[12]
        for i in range(0, len(hg)):
            s1.cell(row=t + 2, column=i + 3).value = hg[i]

        for wd in wbk.worksheets[:2]:
            wd.cell(row=1, column=1).value = 'Sr. No.'
            wd.cell(row=1, column=2).value = 'CCN'
            wd.cell(row=t + 2, column=1).value = t + 1
            wd.cell(row=t + 2, column=2).value = ccn

        # wb = xlrd.open_workbook('foo1.xls', on_demand=True)
        sheetno = len(wb.sheet_names())
        sheet_with_details = 0
        for i in range(sheetno):
            worksheet = wb.sheet_by_index(i)
            a = worksheet.cell_value(1, 1)
            if 'Details' in a:
                sheet_contains_details = i
                break

        sheet_2 = wb.sheet_by_index(sheet_contains_details)
        sheet_2.cell_value(0, 0)
        b = []

        for i in range(2, sheet_2.nrows):
            b.append(sheet_2.cell_value(i, 2))

    ############################################################<
    # code block returns list of amounts = clean[] in Approved amount (Rs.) column

        for i, elem in enumerate(b):
            if 'MOU' in elem or 'mou' in elem:
                text = b[i]

        templist, clean = [], []
        temp = text.split('\t')
        for i in temp:
            if '/-' in i:
                templist.append(i)
        for i in templist:
            num = i.split('/')[0]
            if 'rs' in num:
                num = num.strip(',rs.')
            clean.append(int(num))
        rd = text
    ###########################################################>
        rd = rd.replace('Rs', 'rs')
        rd = rd.replace('RS', 'rs')
        rd = rd.replace('\t', ' ')

        reason = rd.split(',rs')
        deduct = []
        deduct_res = []
        for i in reason:
            w8 = i.find('rs') + 2
            g = i[w8:]
            u8 = g.find('/-') + w8
            # deduct.append(i[w8:u8])
            w8 = i.find('/-') + 2
            deduct_res.append(i[w8:])
            if '' in deduct_res:
                deduct_res.remove('')
        # print(deduct,deduct_res,reason)
    #############################################################<
        # send clean array to deducts coloumn
        deduct.extend(clean)
    #############################################################>

    #######################################################<
    #fixing s2 sheet

        with open(sys.argv[0].strip('.py')+'/output.txt') as f:
            txt = f.read()
        x = re.search(r"Copay +\d+", txt)
        x1 = x.group()
        x2 = re.search(r"\d+", x1)
        cp = x2.group()

        temp = b
        temp[6] = temp[11]
        temp[7] = temp[5]
        temp[8] = temp[9]
        temp[9] = temp[13]
        temp[5] = temp[12]
        temp[10] = cp

    ########################################################>
        for i in range(0, len(b)):
            # s2 is foo1.xls -> 1
            s2.cell(row=t + 2, column=i + 3).value = temp[i]
    #########################################################<
    # fixing s1 sheet
        hg[11] = temp[4]
        hg[12] = clean[0]
        hg.append(transaction_date)
        for i in range(0, len(hg)):
            s1.cell(row=t + 2, column=i + 3).value = hg[i]
    #########################################################>
        for i in range(0, len(deduct)):
            if (deduct_res[i].find('MOU Discount') != -1 or deduct_res[i].find('MOU discount') != -1):
                discount = deduct[i]
                s1.cell(row=t + 2, column=14).value = float(hg[-2]) - float(discount)
                s1.cell(row=t + 2, column=15).value = discount
            if (deduct[i] != ''):
                #########################################################<
                # start srno of s3 with 2
                # row = s3.max_row
                #########################################################>
                row = s3.max_row + 1
                s3.cell(row=row, column=1).value = row + 1
                s3.cell(row=row, column=2).value = ccn
                s3.cell(row=row, column=3).value = deduct[i]
                s3.cell(row=row, column=4).value = deduct_res[i]
        s1.cell(row=t + 2, column=16).value = hg[-1]


    print("Done")
    wbk.save(wbkName)
    wbk.close()
    subprocess.run(["python", "make_master.py", 'aditya', op, '', wbkName])
    ###########################################################
    move_master_to_master_insurer(sys.argv[2], pdfpath=pdfpath)
    mark_flag('X', sys.argv[1])
    print(f'processed {wbkName}')
except SystemExit as e:
    v = e.code
    if 'exit' in v:
        a =1
        os._exit(0)
except:
    log_exceptions()
    pass