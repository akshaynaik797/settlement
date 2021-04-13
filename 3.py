import os
import subprocess
import sys
import re

import camelot
import openpyxl
import pdftotext
import xlrd

from make_log import log_exceptions
from backend import mark_flag
from movemaster import move_master_to_master_insurer

try:
    pdfpath = sys.argv[1]
    onlyfiles = [os.path.split(pdfpath)[1]]
    mypath = os.path.dirname(pdfpath)+'/'

    hosp_name = ''

    with open(pdfpath, "rb") as f:
        pdf = pdftotext.PDF(f)
    with open('temp_files/output.txt', 'w', encoding='utf-8') as f:
        f.write(" ".join(pdf))
    with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
        f = myfile.read()
    if 'Balaji Medical' in f:
        op = 'Tpappg@maxhealthcare.com May@2020 outlook.office365.com Max PPT'
        hosp_name = 'Max'
    else:
        op = 'mediclaim@inamdarhospital.org Mediclaim@2019 imap.gmail.com inamdar hospital'
        hosp_name = 'inamdar'
    ###########################################################
    wbkName = os.path.join('temp_files', 'icici_lombard' + hosp_name + '.xlsx')
    t, wq =0, 0
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]

    tables = camelot.read_pdf(pdfpath, line_scale=40, pages='2')
    tables.export('temp_files/foo1.xlsx', f='excel')
    twb = openpyxl.load_workbook('temp_files/foo1.xlsx', read_only=True)
    tws = twb.active

    ccn = ''
    temp = re.compile(r"(?<=Claim No).*(?=AL)").search(f)
    if temp is not None:
        temp = temp.group()
        for i in [':']:
            temp = temp.replace(i, '')
        ccn = temp.strip()

    sh1 = ['Sr No', 'Claim No', 'UHID NO', 'Name of the Patient', 'Policy Name', 'Requested Amount', 'Final Amount Settled',
           'Diagnosis', 'Date of Admission', 'Date Of Discharge', 'CO-PAYMENT AMOUNT', 'DISALLOWED AMOUNT',
           'DISALLOWED REASONS', 'settled amount', 'cheque/EFT vide ref. no.', 'date of payment', 'TDS']
    sh2 = ['Sr No', 'Claim No', 'Charges Details', 'Claimed', 'Deductions', 'Paid', 'Reason for Deductions']

    sh1_regex = [r"(?<=UHID NO :).*(?=Relationship)", r"(?<=Name of the Patient).*(?=Policy)",
                 r"(?<=Name of the Patient).*(?=Policy)", r"(?<=Requested Amount).*",
                 r"(?<=Final Amount Settled in Rs.).*", r"(?<=Diagnosis).*", r"(?<=Date Of Admission).*(?=Date)",
                 r"(?<=Date Of Discharge).*", r"(?<=Date Of Discharge).*", r"(?<=Final Amount Settled in Rs.).*",
                 r"(?<=Final Amount Settled in Rs.).*", r"(?<=Final Amount Settled in Rs.).*", r"(?<=ref. no.).*(?=dated)",
                 r"(?<=dated).*(?=towards)", r"(?<=TDS is) *\d+"]

    for i in range(0, len(sh1)):
        s1.cell(row=1, column=i + 1).value = sh1[i]

    sh1_data = ['1', ccn]
    for regex in sh1_regex:
        clean = ''
        temp = re.compile(regex).search(f)
        if temp is not None:
            temp = temp.group()
            for i in [':']:
                temp = temp.replace(i, '')
            clean = temp.strip()
        sh1_data.append(clean)

    for i in range(0, len(sh1_data)):
        s1.cell(row=2, column=i + 1).value = sh1_data[i]

    for i in range(0, len(sh2)):
        s2.cell(row=1, column=i + 1).value = sh2[i]

    table = []
    for row in tws.rows:
        for cell in row:
            for sno in range(1, tws.max_row):
                if cell.value == str(sno):
                    t_row = []
                    for ele in row[1:]:
                        ele = ele.value
                        if ele is not None:
                            ele = re.sub(r"\s", ' ', ele, re.MULTILINE)
                            t_row.append(ele)
                    table.append(t_row[1:])

    table2 = []
    for cnt, row in enumerate(table):
        row = [str(cnt+1), ccn] + row
        table2.append(row)

    for i in range(0, len(sh2)):
        s2.cell(row=1, column=i + 1).value = sh2[i]

    for row, record in enumerate(table2):
        for col, ele in enumerate(record):
            s2.cell(row=row+2, column=col+1).value = ele
    print("Done")
    wbk.save(wbkName)
    wbk.close()
    subprocess.run(["python", "make_master.py", 'icici_lombard', op, '', wbkName])
    ###########################################################
    move_master_to_master_insurer(sys.argv[2], pdfpath=pdfpath)
    mark_flag('X', sys.argv[2])
    print(f'processed {wbkName}')
except SystemExit as e:
    v = e.code
    if 'exit' in v:
        a =1
        os._exit(0)
except:
    log_exceptions()
    pass