import os
import glob
import os.path
import sys
from os import listdir
from os import path
import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl import load_workbook
from make_log import log_exceptions

try:

	ins = str(sys.argv[1])

	redFill = PatternFill(start_color='FFFF0000',
					   end_color='FFFF0000',
					   fill_type='solid')
	#with open('config.txt', 'r') as myfile:
	#	f = myfile.read()
	op=sys.argv[2]
	#op.pop(-1)
	wbName = 'master.xlsx'
	wb = openpyxl.Workbook()
	wb.create_sheet('Sheet1')
	wb.create_sheet('count')
	wb.create_sheet('count_star')
	wb.create_sheet('error_sheet')
	main_s1=wb.worksheets[0]
	main_s2=wb.worksheets[1]
	main_s5=wb.worksheets[4]
	main_s5.cell(row=1, column=1).value='ins_id'
	main_s5.cell(row=1, column=2).value='claim_id'
	main_s5.cell(row=1, column=3).value='attachment'
	main_s5.cell(row=1, column=4).value='Need assistance'
	sh1=['Sno','HospitalID','InsurerID','ALNO','ClaimNo','MemberID','PolicyNo','PatientName','InsuranceCompany','AccountNo','BeneficiaryBank Name','Diagnosis','UTRNo','BilledAmount','SettledAmount','TDS','NetPayable','DiscountAmt','COPay','PolicyHolder','IPNo','PrimaryBeneficiary','EmployeeID','InsurerClaimNo','InsurerMemberID','TaxDeductedatSource','Netamount payment','PaidbythePatient','ProrataBasis','PolicyExcessDeductible','BeneficiaryName','BalanceSumInsuredBeforeClaim','NetPayable','BalanceSumInsuredAfterClaim','TDS%','Remarks','PaymentTo','DateofAdmission','DateofDischarge','AmtPaidtoHospital','BillAmt','PayableAmt','SettledAmt','SumInsured','ALAmount	Approved','Amount','HospitalAmount','AmountUtilised','FinalAmountSettled','DateOfPayment','ServiceTax','TotalwithServiceTax','InsuredPerson','CorporateName','DeductibleAmt','Transactiondate','LOCALAmount','ChequeDate','UHCApprovedHospitalAmt','InsurerApprovedHospitalAmt','InsurerApprovedEmployeeAmt','PayableAmount','NEFTTransactionNumber','TransactionDate','CorporateName','Claimed','PreHospitalisationPayableAmount','PostHospitalisationPayableAmount','AddonBenefit','Claimed','Paid','BillAmount','PayableAmount(INR)','BillDate','BillNo','AmountSettled','ApprovedAmount','less','Excess of Defined / Ailment Limit','policy deduction','Limit exceed deduction','non payable deduction','Bill deduction','Other deduction']
	sh2=['Sr. No.','HospitalID','InsurerID','Claim ID','Details','Bill amount','Payable Amount','Deducted Amt','Reason for Deduction','Discount']
	for i in range(0,len(sh1)):
		#main_s1.cell(row=1, column=i+1).value=i+1
		main_s1.cell(row=1, column=i+1).value=sh1[i]
	for i in range(0,len(sh2)):
		main_s2.cell(row=1, column=i+1).value=sh2[i]
	for i in range(0,1):
		k = op.split(' ')

	#aditya_birla
		if ins == 'aditya_birla':
			wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			s3=wbk.worksheets[2]
			row_count = s1.max_row

			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				b=[]
				bp=[]
				be=[]
				pe=[]
				dt=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value)
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					b.append(s1.cell(row=t, column=12).value)
					bp.append(s1.cell(row=t, column=13).value)
					be.append(s1.cell(row=t, column=14).value)
					pe.append(s1.cell(row=t, column=15).value)
					dt.append(s1.cell(row=t, column=16).value)
				main_row_count = main_s1.max_row
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=7).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=6).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=20).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=11).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=10).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=17).value=e[i]
					main_s1.cell(row=i+main_row_count+1, column=13).value=eo[i]
					main_s1.cell(row=i+main_row_count+1, column=12).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=bp[i]
					main_s1.cell(row=i+main_row_count+1, column=55).value=be[i]
					main_s1.cell(row=i+main_row_count+1, column=18).value=pe[i]
					# check the format here before inserting the value...it should be DD-MM-YYYY varun
					main_s1.cell(row=i+main_row_count+1, column=64).value=dt[i].replace('--', '-')
				row_count = s2.max_row
				b=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				for t in range(2,row_count+1):
					p.append(s2.cell(row=t, column=3).value)
					np.append(s2.cell(row=t, column=4).value)
					r.append(s2.cell(row=t, column=5).value)
					s_v.append(s2.cell(row=t, column=6).value)
					we.append(s2.cell(row=t, column=7).value)
					b.append(s2.cell(row=t, column=8).value)
					ew.append(s2.cell(row=t, column=9).value)
					e.append(s2.cell(row=t, column=10).value)
					eo.append(s2.cell(row=t, column=11).value)
					ro.append(s2.cell(row=t, column=12).value)
				#print(eo)
				for i in range(0,len(p)):
					main_s1.cell(row=i+main_row_count+1, column=44).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=14).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=46).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=47).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=16).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=18).value=e[i]
					main_s1.cell(row=i+main_row_count+1, column=48).value=ro[i]
				we=[]
				eo=[]
				ccn=[]
				row_count = s3.max_row
				for t in range(2,row_count+1):
					ccn.append(s3.cell(row=t, column=2).value)
					we.append(s3.cell(row=t, column=3).value)
					eo.append(s3.cell(row=t, column=4).value)
				main_row_count = main_s2.max_row
				for i in range(0,len(we)):
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s2.cell(row=i+main_row_count+1, column=8).value=we[i]
					main_s2.cell(row=i+main_row_count+1, column=9).value=eo[i]

	#apollo_munich
		if ins=='apollo_munich':
			wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				b=[]
				bp=[]
				be=[]
				pe=[]
				qw=[]
				re=[]
				mid=[]
				uhid=[]
				p_name=[]
				for t in range(2,row_count+1):
					mid.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=mid[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					b.append(s1.cell(row=t, column=12).value)
					bp.append(s1.cell(row=t, column=13).value )
					be.append(s1.cell(row=t, column=14).value)
					pe.append(s1.cell(row=t, column=15).value)
					qw.append(s1.cell(row=t, column=16).value)
					re.append(s1.cell(row=t, column=17).value)
					ccn.append(s1.cell(row=t, column=18).value)
					uhid.append(s1.cell(row=t, column=19).value)
					p_name.append(s1.cell(row=t, column=20).value)
				#print(re)
		#insert
				main_row_count = main_s1.max_row
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=5).value=mid[i]
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=6).value=uhid[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=p_name[i]
					main_s1.cell(row=i+main_row_count+1, column=66).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=12).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=14).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=e[i]
					main_s1.cell(row=i+main_row_count+1, column=17).value=eo[i]
					main_s1.cell(row=i+main_row_count+1, column=63).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=13).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=55).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=bp[i]
					main_s1.cell(row=i+main_row_count+1, column=18).value=be[i]
					main_s1.cell(row=i+main_row_count+1, column=16).value=pe[i]
					main_s1.cell(row=i+main_row_count+1, column=21).value=qw[i]
					main_s1.cell(row=i+main_row_count+1, column=75).value=re[i]


				row_count = s2.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				for t in range(2,row_count+1):
					ccn.append(s2.cell(row=t, column=2).value )
					p.append(s2.cell(row=t, column=3).value)
					np.append(s2.cell(row=t, column=4).value)
					r.append(s2.cell(row=t, column=5).value)
				#print(r,ccn)

		#insert
				main_row_count = main_s2.max_row
				for i in range(0,len(r)):
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s2.cell(row=i+main_row_count+1, column=5).value=p[i]
					main_s2.cell(row=i+main_row_count+1, column=8).value=np[i]
					main_s2.cell(row=i+main_row_count+1, column=9).value=r[i]
	#fgh
		if ins=='fgh':
			wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				b=[]
				bp=[]
				be=[]
				pe=[]
				qw=[]
				re=[]
				ks=[]
				sk=[]
				dis=[]
				cop=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					b.append(s1.cell(row=t, column=12).value)
					bp.append(s1.cell(row=t, column=13).value )
					be.append(s1.cell(row=t, column=14).value)
					pe.append(s1.cell(row=t, column=15).value)
					qw.append(s1.cell(row=t, column=16).value)
					re.append(s1.cell(row=t, column=17).value)
					ks.append(s1.cell(row=t, column=18).value)
					sk.append(s1.cell(row=t, column=19).value)
					dis.append(s1.cell(row=t, column=20).value)
					cop.append(s1.cell(row=t, column=21).value)
				#print(sk)
		#insert
				stt=[]
				main_row_count = main_s1.max_row
				for i in range(0,len(ccn)):
					x=[b[i],pe[i],be[i],dis[i],cop[i]]
					x=['0.00' if v is None else v for v in x]
					#x=[sub.replace('None','0.0') for sub in x]
					fur=float(x[0])+float(x[1])+float(x[2])+float(x[3])+float(x[4])
					stt.append(str(fur))
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=p[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=s_v[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=22).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=10).value=e[i]
					main_s1.cell(row=i+main_row_count+1, column=11).value=eo[i]
					main_s1.cell(row=i+main_row_count+1, column=14).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=55).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=51).value=bp[i]
					main_s1.cell(row=i+main_row_count+1, column=17).value=be[i]
					main_s1.cell(row=i+main_row_count+1, column=16).value=pe[i]
					main_s1.cell(row=i+main_row_count+1, column=7).value=qw[i]
					main_s1.cell(row=i+main_row_count+1, column=13).value=re[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=ks[i]
					main_s1.cell(row=i+main_row_count+1, column=6).value=sk[i]
					main_s1.cell(row=i+main_row_count+1, column=18).value=dis[i]
					main_s1.cell(row=i+main_row_count+1, column=19).value=cop[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=stt[i]
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				row_count = s2.max_row
				for t in range(2,row_count+1):
					ccn.append(s2.cell(row=t, column=2).value )
					p.append(s2.cell(row=t, column=3).value)
					np.append(s2.cell(row=t, column=5).value)
					r.append(s2.cell(row=t, column=6).value)
					s_v.append(s2.cell(row=t, column=7).value)
					we.append(s2.cell(row=t, column=8).value)
				#print(ccn)
		#insert
				main_row_count = main_s2.max_row
				for i in range(0,len(ccn)):
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s2.cell(row=i+main_row_count+1, column=5).value=p[i]
					main_s2.cell(row=i+main_row_count+1, column=6).value=np[i]
					main_s2.cell(row=i+main_row_count+1, column=8).value=r[i]
					main_s2.cell(row=i+main_row_count+1, column=7).value=s_v[i]
					main_s2.cell(row=i+main_row_count+1, column=9).value=we[i]

	#Fhpl
		if ins=='fhpl':
			wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			s3=wbk.worksheets[2]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				b=[]
				cop=[]
				al=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					b.append(s1.cell(row=t, column=12).value)
					cop.append(s1.cell(row=t, column=13).value)
					al.append(s1.cell(row=t, column=14).value)
				#print(b)
		#insert
				bp=[]
				be=[]
				pe=[]
				qw=[]
				re=[]
				ks=[]
				sk=[]
				row_count=s2.max_row
				for t in range(2,row_count+1):
					bp.append(s2.cell(row=t, column=3).value )
					be.append(s2.cell(row=t, column=4).value)
					pe.append(s2.cell(row=t, column=5).value)
					qw.append(s2.cell(row=t, column=6).value)
					re.append(s2.cell(row=t, column=7).value)
					ks.append(s2.cell(row=t, column=8).value)
					sk.append(s2.cell(row=t, column=9).value)
				#print(sk)
		#insert
				main_row_count = main_s1.max_row
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=5).value=al[i]
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=7).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=23).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=12).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=6).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=e[i]
					main_s1.cell(row=i+main_row_count+1, column=13).value=eo[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=65).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=9).value=cop[i]
					main_s1.cell(row=i+main_row_count+1, column=66).value=bp[i]
					main_s1.cell(row=i+main_row_count+1, column=14).value=be[i]
					main_s1.cell(row=i+main_row_count+1, column=18).value=pe[i]
					main_s1.cell(row=i+main_row_count+1, column=55).value=qw[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=re[i]
					main_s1.cell(row=i+main_row_count+1, column=43).value=re[i]
					main_s1.cell(row=i+main_row_count+1, column=16).value=ks[i]
					main_s1.cell(row=i+main_row_count+1, column=17).value=sk[i]

				row_count = s3.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				for t in range(2,row_count+1):
					ccn.append(s3.cell(row=t, column=2).value )
					p.append(s3.cell(row=t, column=3).value)
					np.append(s3.cell(row=t, column=4).value)
					r.append(s3.cell(row=t, column=5).value)
				#print(ccn,r)
		#insert
				main_row_count = main_s2.max_row
				for i in range(0,len(np)):
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s2.cell(row=i+main_row_count+1, column=5).value=p[i]
					main_s2.cell(row=i+main_row_count+1, column=8).value=np[i]
					main_s2.cell(row=i+main_row_count+1, column=9).value=r[i]


	#health_heritage

		if ins=='health_heritage':
			wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				b=[]
				bp=[]
				be=[]
				pe=[]
				qw=[]
				st_amt=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					b.append(s1.cell(row=t, column=12).value)
					bp.append(s1.cell(row=t, column=13).value)
					be.append(s1.cell(row=t, column=14).value)
					pe.append(s1.cell(row=t, column=15).value)
					qw.append(s1.cell(row=t, column=16).value)
					st_amt.append(s1.cell(row=t, column=17).value)
				#print(bp)
		#insert
				main_row_count = main_s1.max_row
				temp_row=main_row_count
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=6).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=7).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=12).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=66).value=e[i]
					main_s1.cell(row=i+main_row_count+1, column=14).value=e[i]
					main_s1.cell(row=i+main_row_count+1, column=17).value=eo[i]
					main_s1.cell(row=i+main_row_count+1, column=13).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=16).value=bp[i]
					main_s1.cell(row=i+main_row_count+1, column=9).value=be[i]
					main_s1.cell(row=i+main_row_count+1, column=22).value=pe[i]
					main_s1.cell(row=i+main_row_count+1, column=11).value=qw[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=st_amt[i]
				re=[]
				ks=[]
				sk=[]
				row_count=s2.max_row
				for t in range(2,row_count+1):
					re.append(s2.cell(row=t, column=2).value )
					ks.append(s2.cell(row=t, column=3).value)
					sk.append(s2.cell(row=t, column=4).value)
				#print(re,sk)
		#insert
				num_ko=main_s1.max_row
				main_row_count = main_s2.max_row
				for i in range(0,len(sk)):
					'''if sk[i]=='Co-Payment()':
						for ko in range(1,num_ko+1):
							if(re[i]==main_s1.cell(row=ko, column=5).value):
								main_s1.cell(row=ko, column=19).value=ks[i]
					if sk[i].find('discount')!=-1 or sk[i].find('Discount')!=-1:
						
						for ko in range(1,num_ko+1):
							if(re[i]==main_s1.cell(row=ko, column=5).value):
								main_s1.cell(row=ko, column=18).value=ks[i]'''
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=re[i]
					main_s2.cell(row=i+main_row_count+1, column=8).value=ks[i]
					main_s2.cell(row=i+main_row_count+1, column=9).value=sk[i]
				for i in range(0,len(we)):
					det=0.00
					for ip in range(0,len(ks)):
						if(re[ip]==main_s1.cell(row=i+temp_row+1, column=4).value):
							#print(ip)
							det=det+float(ks[ip])

					lp=[main_s1.cell(row=i+temp_row+1, column=19).value]
					lk=[main_s1.cell(row=i+temp_row+1, column=18).value]
					iu=['0.00' if v is None else v for v in lp]
					up=['0.00' if h is None else h for h in lk]
					#print(det)
					#print(iu,up,det-float(iu[0])-float(up[0]))
					#print(float(det-float(iu[0])-float(up[0])))
					main_s1.cell(row=i+temp_row+1, column=55).value=float(det-float(iu[0])-float(up[0]))

	#health_india
		if ins=='health_india':
			wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				b=[]
				bp=[]
				be=[]
				pe=[]
				qw=[]
				re=[]
				ks=[]
				sk=[]
				dis=[]
				cop=[]
				dod=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)

					bp.append(s1.cell(row=t, column=13).value )
					be.append(s1.cell(row=t, column=14).value)
					pe.append(s1.cell(row=t, column=15).value)
					qw.append(s1.cell(row=t, column=16).value)
					re.append(s1.cell(row=t, column=17).value)
					ks.append(s1.cell(row=t, column=18).value)
					sk.append(s1.cell(row=t, column=19).value)
					dis.append(s1.cell(row=t, column=20).value)
					cop.append(s1.cell(row=t, column=21).value)
					ty=s1.cell(row=t, column=12).value
					if(ty!=None):
						hu=ty.find('-')
						b.append(ty[:hu])
						dod.append(ty[hu+1:])
					else:
						dod.append(' ')
				#print(sk)
		#insert
				main_row_count = main_s1.max_row
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=9).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=7).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=54).value=r[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=23).value=we[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=e[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=eo[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=dod[i]
					main_s1.cell(row=i+main_row_count+1, column=12).value=bp[i]
					main_s1.cell(row=i+main_row_count+1, column=13).value=be[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=pe[i]
					main_s1.cell(row=i+main_row_count+1, column=14).value=qw[i]
					main_s1.cell(row=i+main_row_count+1, column=55).value=re[i]
					main_s1.cell(row=i+main_row_count+1, column=18).value=ks[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=sk[i]
					main_s1.cell(row=i+main_row_count+1, column=16).value=dis[i]
					main_s1.cell(row=i+main_row_count+1, column=17).value=cop[i]
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				row_count = s2.max_row
				for t in range(2,row_count+1):
					ccn.append(s2.cell(row=t, column=2).value )
					p.append(s2.cell(row=t, column=5).value)
					np.append(s2.cell(row=t, column=6).value)
					r.append(s2.cell(row=t, column=7).value)
					s_v.append(s2.cell(row=t, column=8).value)
					we.append(s2.cell(row=t, column=9).value)
				#print(ccn)
		#insert
				main_row_count = main_s2.max_row
				for i in range(0,len(ccn)):
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s2.cell(row=i+main_row_count+1, column=6).value=p[i]
					main_s2.cell(row=i+main_row_count+1, column=7).value=np[i]
					main_s2.cell(row=i+main_row_count+1, column=8).value=r[i]
					main_s2.cell(row=i+main_row_count+1, column=9).value=s_v[i]
					main_s2.cell(row=i+main_row_count+1, column=5).value=we[i]


	#HDFC
		if ins=='hdfc':
			wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			s3=wbk.worksheets[2]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				b=[]
				bp=[]
				bo=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					b.append(s1.cell(row=t, column=12).value)
					bp.append(s1.cell(row=t, column=13).value)
					bo.append(s1.cell(row=t, column=14).value)
				#print(e)
		#insert
				main_row_count = main_s1.max_row
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=6).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=7).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=10).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=11).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=12).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=e[i]
					main_s1.cell(row=i+main_row_count+1, column=22).value=eo[i]
					main_s1.cell(row=i+main_row_count+1, column=13).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=bp[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=bo[i]
				ro=[]
				b=[]
				re=[]
				ks=[]
				sk=[]
				dis=[]
				ded=[]
				sett=[]
				row_count=s3.max_row
				for t in range(2,row_count+1):
					re.append(s3.cell(row=t, column=3).value )
					ks.append(s3.cell(row=t, column=4).value)
					sk.append(s3.cell(row=t, column=5).value)
					ro.append(s3.cell(row=t, column=6).value)
					b.append(s3.cell(row=t, column=7).value)
					dis.append(s3.cell(row=t, column=8).value)
					ded.append(s3.cell(row=t, column=9).value)
					sett.append(s3.cell(row=t, column=10).value)
				#print(b)
		#insert
				#main_row_count = main_s1.max_row
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=51).value=re[i]
					main_s1.cell(row=i+main_row_count+1, column=52).value=ks[i]
					main_s1.cell(row=i+main_row_count+1, column=16).value=sk[i]
					main_s1.cell(row=i+main_row_count+1, column=19).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=17).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=18).value=dis[i]
					main_s1.cell(row=i+main_row_count+1, column=55).value=ded[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=sett[i]
				ro=[]
				b=[]
				re=[]
				ks=[]
				sk=[]
				pe=[]
				qw=[]
				row_count=s2.max_row
				for t in range(2,row_count+1):
					pe.append(s2.cell(row=t, column=2).value )
					re.append(s2.cell(row=t, column=3).value )
					ks.append(s2.cell(row=t, column=4).value)
					b.append(s2.cell(row=t, column=7).value)
					sk.append(s2.cell(row=t, column=5).value)
					qw.append(s2.cell(row=t, column=8).value )
					ro.append(s2.cell(row=t, column=6).value)

				#print(pe,qw)
		#insert
				main_row_count = main_s2.max_row
				for i in range(0,len(re)):
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=pe[i]
					main_s2.cell(row=i+main_row_count+1, column=5).value=re[i]
					main_s2.cell(row=i+main_row_count+1, column=6).value=ks[i]
					main_s2.cell(row=i+main_row_count+1, column=7).value=b[i]
					main_s2.cell(row=i+main_row_count+1, column=8).value=sk[i]
					main_s2.cell(row=i+main_row_count+1, column=9).value=qw[i]
					main_s2.cell(row=i+main_row_count+1, column=10).value=ro[i]
	#Religare
		if ins=='religare':
			wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				b=[]
				bp=[]
				bo=[]
				e_id=[]
				e_n=[]
				ded=[]
				dis=[]
				al=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					b.append(s1.cell(row=t, column=12).value)
					bp.append(s1.cell(row=t, column=13).value)
					bo.append(s1.cell(row=t, column=14).value)
					e_id.append(s1.cell(row=t, column=15).value)
					e_n.append(s1.cell(row=t, column=16).value)
					dis.append(s1.cell(row=t, column=17).value)
					ded.append(s1.cell(row=t, column=19).value)
					al.append(s1.cell(row=t, column=18).value)
				#print(eo)
		#insert
				stt=[]
				main_row_count = main_s1.max_row
				for i in range(0,len(ccn)):
					x=[b[i],bp[i],ded[i],dis[i]]
					x=['0.00' if v is None else v for v in x]
					x=[sub.replace(' ','') for sub in x]
					x=['0.00' if v is '' else v for v in x]
					#print(x)
					fur=float(x[0])+float(x[2])+float(x[1])+float(x[3])#+float(x[4])
					stt.append(str(fur))
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=4).value=al[i]
					main_s1.cell(row=i+main_row_count+1, column=7).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=11).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=22).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=13).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=e[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=eo[i]
					main_s1.cell(row=i+main_row_count+1, column=14).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=16).value=round(float(b[i])/10)
					main_s1.cell(row=i+main_row_count+1, column=17).value=float(b[i])-round(float(b[i])/10)
					main_s1.cell(row=i+main_row_count+1, column=19).value=bp[i]
					main_s1.cell(row=i+main_row_count+1, column=18).value=dis[i]
					main_s1.cell(row=i+main_row_count+1, column=55).value=ded[i]
					main_s1.cell(row=i+main_row_count+1, column=23).value=e_id[i]
					main_s1.cell(row=i+main_row_count+1, column=22).value=e_n[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=stt[i]
				ro=[]
				b=[]
				re=[]
				ks=[]
				sk=[]
				pe=[]
				row_count=s2.max_row
				for t in range(2,row_count+1):
					pe.append(s2.cell(row=t, column=2).value )
					re.append(s2.cell(row=t, column=3).value )
					ks.append(s2.cell(row=t, column=6).value)
					sk.append(s2.cell(row=t, column=7).value)
					ro.append(s2.cell(row=t, column=8).value)
					b.append(s2.cell(row=t, column=9).value)
				#print(pe,b)
		#insert
				main_row_count = main_s2.max_row
				for i in range(0,len(re)):
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=pe[i]
					main_s2.cell(row=i+main_row_count+1, column=5).value=re[i]
					main_s2.cell(row=i+main_row_count+1, column=6).value=ks[i]
					main_s2.cell(row=i+main_row_count+1, column=7).value=sk[i]
					main_s2.cell(row=i+main_row_count+1, column=8).value=ro[i]
					main_s2.cell(row=i+main_row_count+1, column=9).value=b[i]
	#star
		if ins=='star':
			wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				b=[]
				bp=[]
				ks=[]
				sk=[]
				dis=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					b.append(s1.cell(row=t, column=12).value)
					bp.append(s1.cell(row=t, column=13).value)
					ks.append(s1.cell(row=t, column=14).value)
					sk.append(s1.cell(row=t, column=15).value)
					dis.append(s1.cell(row=t, column=16).value)
				#print(ks)
		#insert
				main_row_count = main_s1.max_row
				star_row=main_row_count
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=7).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=12).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=we[i]
					#main_s1.cell(row=i+main_row_count+1, column=12).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=14).value=e[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=eo[i]
					if(eo[i]!=None):
						main_s1.cell(row=i+main_row_count+1, column=16).value=round(float(eo[i])/10)
					main_s1.cell(row=i+main_row_count+1, column=67).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=68).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=69).value=bp[i]
					main_s1.cell(row=i+main_row_count+1, column=55).value=sk[i]
					main_s1.cell(row=i+main_row_count+1, column=18).value=dis[i]
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				row_count = s2.max_row
				for t in range(2,row_count+1):
					ccn.append(s2.cell(row=t, column=2).value )
					p.append(s2.cell(row=t, column=3).value)
					np.append(s2.cell(row=t, column=4).value)
					r.append(s2.cell(row=t, column=5).value)
					s_v.append(s2.cell(row=t, column=6).value)
				#print(s_v)

		#insert
				main_row_count = main_s2.max_row
				for i in range(0,len(r)):
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s2.cell(row=i+main_row_count+1, column=5).value=p[i]
					main_s2.cell(row=i+main_row_count+1, column=6).value=np[i]
					main_s2.cell(row=i+main_row_count+1, column=7).value=r[i]
					if(np[i]!=None):
						mf=int(np[i])
					else:
						mf_row = main_s5.max_row
						if(main_s5.cell(row=mf_row, column=2).value!=ccn[i]):
							main_s5.cell(row=mf_row+1, column=1).value=ins
							main_s5.cell(row=mf_row+1, column=2).value=ccn[i]
							main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
						continue
					if(r[i])!=None:
						dg=int(r[i])
					else:
						dg=0
					main_s2.cell(row=i+main_row_count+1, column=8).value=mf-dg
					main_s2.cell(row=i+main_row_count+1, column=9).value=s_v[i]
	#small_star
		if ins=='small_star':
			wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				da=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value)
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					da.append(s1.cell(row=t, column=7).value)
				#print(s_v)
		#insert
				ms=[]
				main_row_count = main_s1.max_row
				for i in range(1,main_row_count+1):
					ms.append(main_s1.cell(row=i, column=5).value)
				#print(ms,ccn)
				df=[]
				fd=[]
				for i in range(0,len(ccn)):
					for y in range(0,len(ms)):
						if(ms[y]==ccn[i]):
							#print(y)
							df.append(i)
							fd.append(y)
				for i in range(0,len(ccn)):
					if i in df:
						j=df.index(i)
						y=fd[j]
						#print(y)
						main_s1.cell(row=y+1, column=1).value=sys.argv[3]
						main_s1.cell(row=y+1, column=2).value=k[3]
						main_s1.cell(row=y+1, column=3).value=ins
						main_s1.cell(row=y+1, column=4).value=ccn[i]
						main_s1.cell(row=y+1, column=5).value=ccn[i]
						main_s1.cell(row=y+1, column=13).value=p[i]
						main_s1.cell(row=y+1, column=17).value=np[i]
						#main_s1.cell(row=y+1, column=).value=r[i]
						main_s1.cell(row=y+1, column=10).value=s_v[i]
						main_s1.cell(row=y+1, column=64).value=da[i]
					else:
						main_row_count = main_s1.max_row
						main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
						main_s1.cell(row=main_row_count+1, column=2).value=k[3]
						main_s1.cell(row=main_row_count+1, column=3).value=ins
						main_s1.cell(row=main_row_count+1, column=4).value=ccn[i]
						main_s1.cell(row=main_row_count+1, column=5).value=ccn[i]
						main_s1.cell(row=main_row_count+1, column=13).value=p[i]
						main_s1.cell(row=main_row_count+1, column=17).value=np[i]
						#main_s1.cell(row=i+main_row_count+1, column=).value=r[i]
						main_s1.cell(row=main_row_count+1, column=10).value=s_v[i]
						main_s1.cell(row=main_row_count+1, column=64).value=da[i]

			'''	main_row_count = main_s1.max_row
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=63).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=17).value=np[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=10).value=s_v[i]'''
	#united
		if ins=='united':
			wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			s3=wbk.worksheets[2]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				b=[]
				bp=[]
				be=[]
				pe=[]
				qw=[]
				re=[]
				ks=[]
				sk=[]
				of=[]
				oe=[]
				pri_b=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					b.append(s1.cell(row=t, column=12).value)
					bp.append(s1.cell(row=t, column=13).value)
					be.append(s1.cell(row=t, column=14).value)
					pri_b.append(s1.cell(row=t, column=15).value)
				#print(be)
				row_count = s2.max_row
				for t in range(2,row_count+1):
					pe.append(s2.cell(row=t, column=3).value)
					qw.append(s2.cell(row=t, column=4).value)
					re.append(s2.cell(row=t, column=5).value)
					ks.append(s2.cell(row=t, column=6).value)
					sk.append(s2.cell(row=t, column=7).value)
					of.append(s2.cell(row=t, column=8).value)
					oe.append(s2.cell(row=t, column=9).value)
				#print(oe)
		#insert
				#print(p)
				main_row_count = main_s1.max_row
				temp_row=main_row_count
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=7).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=23).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=9).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=13).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=e[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=eo[i]
					main_s1.cell(row=i+main_row_count+1, column=12).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=57).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=14).value=bp[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=be[i]
					main_s1.cell(row=i+main_row_count+1, column=59).value=pe[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=qw[i]
					main_s1.cell(row=i+main_row_count+1, column=60).value=re[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=re[i]
					main_s1.cell(row=i+main_row_count+1, column=61).value=ks[i]
					main_s1.cell(row=i+main_row_count+1, column=62).value=sk[i]
					main_s1.cell(row=i+main_row_count+1, column=17).value=sk[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=of[i]
					main_s1.cell(row=i+main_row_count+1, column=16).value=oe[i]
					main_s1.cell(row=i+main_row_count+1, column=22).value=pri_b[i]
				row_count = s3.max_row
				ccn=[]
				p=[]
				np=[]
				for t in range(2,row_count+1):
					ccn.append(s3.cell(row=t, column=2).value )
					p.append(s3.cell(row=t, column=3).value)
					np.append(s3.cell(row=t, column=4).value)
				#print(np)
		#insert
				num_ko=main_s1.max_row
				main_row_count = main_s2.max_row
				for i in range(0,len(np)):
					'''if np[i]=='Co-Payment()':
						for ko in range(1,num_ko+1):
							if(ccn[i]==main_s1.cell(row=ko, column=5).value):
								main_s1.cell(row=ko, column=19).value=p[i]
					if np[i]=='LOC Limit Exhausted()':
						for ko in range(1,num_ko+1):
							if(ccn[i]==main_s1.cell(row=ko, column=5).value):
								main_s1.cell(row=ko, column=80).value=p[i]
					if np[i]=='Advance()':
						for ko in range(1,num_ko+1):
							if(ccn[i]==main_s1.cell(row=ko, column=5).value):
								main_s1.cell(row=ko, column=81).value=p[i]
					if np[i].find('discount')!=-1 or np[i].find('Discount')!=-1:
						
						for ko in range(1,num_ko+1):
							if(ccn[i]==main_s1.cell(row=ko, column=5).value):
								main_s1.cell(row=ko, column=18).value=p[i]'''
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s2.cell(row=i+main_row_count+1, column=8).value=p[i]
					main_s2.cell(row=i+main_row_count+1, column=9).value=np[i]
				for i in range(0,len(oe)):
					det=0.00
					for ip in range(0,len(np)):
						if(ccn[ip]==main_s1.cell(row=i+temp_row+1, column=4).value):
							det=det+float(p[ip])

					lp=[main_s1.cell(row=i+temp_row+1, column=19).value]
					lk=[main_s1.cell(row=i+temp_row+1, column=18).value]
					jp=[main_s1.cell(row=i+temp_row+1, column=80).value]
					jk=[main_s1.cell(row=i+temp_row+1, column=81).value]
					ju=['0.00' if qw is None else qw for qw in jp]
					jl=['0.00' if eq is None else eq for eq in jk]
					iu=['0.00' if v is None else v for v in lp]
					up=['0.00' if h is None else h for h in lk]
					#print(det)
					#print(iu,up,det-float(iu[0])-float(up[0]))
					main_s1.cell(row=i+temp_row+1, column=55).value=float(det-float(iu[0])-float(up[0])-float(ju[0])-float(jl[0]))
	#vidal
		if ins=='vidal':
			wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				b=[]
				bp=[]
				be=[]
				pe=[]
				qw=[]
				re=[]
				ks=[]
				sk=[]
				of=[]
				oe=[]
				c_f_n=[]
				ic=[]
				ded_amt=[]
				net_amt=[]
				set_amt=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					b.append(s1.cell(row=t, column=12).value)
					bp.append(s1.cell(row=t, column=13).value )
					be.append(s1.cell(row=t, column=14).value)
					pe.append(s1.cell(row=t, column=15).value)
					qw.append(s1.cell(row=t, column=16).value)
					re.append(s1.cell(row=t, column=17).value)
					ks.append(s1.cell(row=t, column=18).value)
					sk.append(s1.cell(row=t, column=19).value)
					of.append(s1.cell(row=t, column=20).value)
					oe.append(s1.cell(row=t, column=21).value)
					c_f_n.append(s1.cell(row=t, column=22).value)
					ic.append(s1.cell(row=t, column=23).value)
					ded_amt.append(s1.cell(row=t, column=24).value)
					net_amt.append(s1.cell(row=t, column=25).value)
					set_amt.append(s1.cell(row=t, column=26).value)
				#print(oe)
		#insert
				main_row_count = main_s1.max_row
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=6).value=c_f_n[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=22).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=7).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=54).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=9).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=23).value=e[i]
					main_s1.cell(row=i+main_row_count+1, column=12).value=eo[i]
					main_s1.cell(row=i+main_row_count+1, column=16).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=19).value=b[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=bp[i]
					main_s1.cell(row=i+main_row_count+1, column=18).value=be[i]
					main_s1.cell(row=i+main_row_count+1, column=17).value=pe[i]
					main_s1.cell(row=i+main_row_count+1, column=63).value=qw[i]
					main_s1.cell(row=i+main_row_count+1, column=13).value=qw[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=re[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=ks[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=sk[i]
					main_s1.cell(row=i+main_row_count+1, column=21).value=of[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=oe[i]
					main_s1.cell(row=i+main_row_count+1, column=9).value=ic[i]
					main_s1.cell(row=i+main_row_count+1, column=55).value=ded_amt[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=net_amt[i]
					main_s1.cell(row=i+main_row_count+1, column=14).value=set_amt[i]
				row_count = s2.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				v=[]

				for t in range(2,row_count+1):
					ccn.append(s2.cell(row=t, column=2).value )
					p.append(s2.cell(row=t, column=5).value)
					np.append(s2.cell(row=t, column=6).value)
					r.append(s2.cell(row=t, column=7).value)
					s_v.append(s2.cell(row=t, column=8).value)
					v.append(s2.cell(row=t, column=9).value)
				#print(v)
		#insert
				main_row_count = main_s2.max_row
				for i in range(0,len(np)):
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s2.cell(row=i+main_row_count+1, column=5).value=p[i]
					main_s2.cell(row=i+main_row_count+1, column=6).value=np[i]
					main_s2.cell(row=i+main_row_count+1, column=8).value=r[i]
					main_s2.cell(row=i+main_row_count+1, column=7).value=s_v[i]
					main_s2.cell(row=i+main_row_count+1, column=9).value=v[i]

	#Medi_Assist
		if ins=='Medi_Assist':
			wbkName  = sys.argv[4]
			mypath=os.getcwd()+wbkName
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			s3=wbk.worksheets[2]
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row
				ccn=[]
				p=[]
				b=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				oe=[]
				pe=[]
				doa=[]
				dod=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					oe.append(s1.cell(row=t, column=12).value)
					pe.append(s1.cell(row=t, column=13).value)
					doa.append(s1.cell(row=t, column=14).value)
					dod.append(s1.cell(row=t, column=15).value)
				#print(pe)
				main_row_count = main_s1.max_row
			#insert
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=9).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=6).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=7).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=12).value=pe[i]
					main_s1.cell(row=i+main_row_count+1, column=20).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=21).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=22).value=e[i]
					main_s1.cell(row=i+main_row_count+1, column=23).value=eo[i]
					main_s1.cell(row=i+main_row_count+1, column=24).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=25).value=eo[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=doa[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=dod[i]

				p=[]
				b=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				for t in range(2,row_count+1):

					p.append(s2.cell(row=t, column=3).value)
					np.append(s2.cell(row=t, column=4).value)
					r.append(s2.cell(row=t, column=5).value)
					s_v.append(s2.cell(row=t, column=6).value)
					we.append(s2.cell(row=t, column=7).value)
					b.append(s2.cell(row=t, column=8).value)
				#print(b)
			#insert
				for i in range(0,len(p)):
					#main_s1.cell(row=i+main_row_count+1, column=15).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=np[i] #date
					main_s1.cell(row=i+main_row_count+1, column=13).value=r[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=s_v[i]#acount holder
					main_s1.cell(row=i+main_row_count+1, column=11).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=10).value=b[i]
				row_count = s3.max_row
				p=[]
				np=[]
				r=[]
				s_v=[]
				ed=[]
				we=[]
				z=[]
				col_count = s3.max_column
				l=lhs=lc=ltd=led=cun=100
				for i in range(2,col_count+1):
					if s3.cell(row=1, column=i).value=='(LESS)':
						l=i
					if s3.cell(row=1, column=i).value=='(LESS)Hospital Discount':
						lhs=i
					if s3.cell(row=1, column=i).value=='(LESS)Copay':
						lc=i
					if s3.cell(row=1, column=i).value=='(LESS)Tax Deducted at Source':
						ltd=i
					if s3.cell(row=1, column=i).value=='(LESS)Excess of Defined / Ailment Limit':
						led=i
					if s3.cell(row=1, column=i).value=='Net amount recommended for payment':
						cun=i
				for t in range(2,row_count+1):
					p.append(s3.cell(row=t, column=l).value)
					np.append(s3.cell(row=t, column=lhs).value)
					r.append(s3.cell(row=t, column=lc).value)
					s_v.append(s3.cell(row=t, column=ltd).value)
					ed.append(s3.cell(row=t, column=led).value)
					we.append(s3.cell(row=t, column=cun).value)
				#print(p)

		#insert
				for i in range(0,len(b)):
					main_s1.cell(row=i+main_row_count+1, column=78).value=p[i]  #less
					main_s1.cell(row=i+main_row_count+1, column=18).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=19).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=16).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=79).value=ed[i]
					main_s1.cell(row=i+main_row_count+1, column=17).value=we[i]
				m=3
				b=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				kl=[]
				for wd in wbk.worksheets[3:]:
					row_count = wd.max_row

					yu=str(wbk.worksheets[m])
					g=yu.find('"')
					yu=yu[g+1:]
					g=yu.find('"')
					yu=yu[:g]
					for t in range(2,row_count+1):
						kl.append(wd.cell(row=t, column=2).value)
						b.append(yu)
						p.append(wd.cell(row=t, column=3).value)
						np.append(wd.cell(row=t, column=4).value)
						r.append(wd.cell(row=t, column=5).value)
						s_v.append(wd.cell(row=t, column=6).value)
					m+=1
				#print(s_v)
			#insert
				num_k=main_row_count
				it=0
				temp_er=0
				main_row_count = main_s2.max_row
				for i in range(0,len(b)):
					if(p[i]!=None):
						if(b[i]=='Total'):
							num_k+=1
							main_s1.cell(row=num_k, column=14).value=p[i]
							main_s1.cell(row=num_k, column=15).value=np[i]
							temp_ded=0
							for gh in range(3,cun):
								omp=s3.cell(row=temp_er+2, column=gh).value
								if omp!=None:
									temp_ded=temp_ded+int(omp)
								#print(s3.cell(row=2, column=gh).value)
							dis_r=s3.cell(row=temp_er+2, column=lhs).value
							if dis_r!=None:
								dis_r=int(dis_r)
							else:
								dis_r=0
							co_r=s3.cell(row=temp_er+2, column=lc).value
							if co_r!=None:
								co_r=int(co_r)
							else:
								co_r=0
							tax=s3.cell(row=temp_er+2, column=ltd).value
							if tax!=None:
								tax=int(tax)
							else:
								tax=0
							#print(temp_ded,dis_r,co_r)
							main_s1.cell(row=num_k, column=55).value=int(r[i])+temp_ded-dis_r-co_r-tax
							#print(int(r[i])+temp_ded-dis_r-co_r-tax)

							temp_er+=1
						main_s2.cell(row=it+main_row_count+1, column=2).value=k[3]
						main_s2.cell(row=it+main_row_count+1, column=3).value=ins
						main_s2.cell(row=it+main_row_count+1, column=4).value=kl[i]
						main_s2.cell(row=it+main_row_count+1, column=5).value=b[i]
						main_s2.cell(row=it+main_row_count+1, column=6).value=p[i]
						main_s2.cell(row=it+main_row_count+1, column=7).value=np[i]
						main_s2.cell(row=it+main_row_count+1, column=8).value=r[i]
						main_s2.cell(row=it+main_row_count+1, column=9).value=s_v[i]
						it+=1

	#icici lombard
		if ins=='icici_lombard':
			wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				b=[]
				bp=[]
				be=[]
				pe=[]
				qw=[]
				re=[]
				ks=[]
				sk=[]
				e_id=[]
				e_name=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					b.append(s1.cell(row=t, column=12).value)
					bp.append(s1.cell(row=t, column=13).value )
					be.append(s1.cell(row=t, column=14).value)
					pe.append(s1.cell(row=t, column=15).value)
					qw.append(s1.cell(row=t, column=16).value)
					re.append(s1.cell(row=t, column=17).value)
					ks.append(s1.cell(row=t, column=18).value)
					e_id.append(s1.cell(row=t, column=19).value)
					e_name.append(s1.cell(row=t, column=20).value)
				#print(ccn)
		#insert
				main_row_count = main_s1.max_row
				temp_row=main_row_count
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=6).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=7).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=14).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=12).value=e[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=eo[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=19).value=b[i]
					#main_s1.cell(row=i+main_row_count+1, column=55).value=bp[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=be[i]
					main_s1.cell(row=i+main_row_count+1, column=17).value=pe[i]
					main_s1.cell(row=i+main_row_count+1, column=13).value=qw[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=re[i]
					main_s1.cell(row=i+main_row_count+1, column=16).value=ks[i]
					main_s1.cell(row=i+main_row_count+1, column=23).value=e_id[i]
					main_s1.cell(row=i+main_row_count+1, column=22).value=e_name[i]
				row_count = s2.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				v=[]

				for t in range(2,row_count+1):
					ccn.append(s2.cell(row=t, column=2).value )
					p.append(s2.cell(row=t, column=3).value)
					np.append(s2.cell(row=t, column=4).value)
					r.append(s2.cell(row=t, column=5).value)
					s_v.append(s2.cell(row=t, column=6).value)
					v.append(s2.cell(row=t, column=7).value)
				#print(v)
		#insert
				num_ko=main_s1.max_row
				main_row_count = main_s2.max_row
				for i in range(0,len(np)):
					'''if v[i]!=None:
						if v[i].find('discount')!=-1 or v[i].find('Discount')!=-1:	
							for ko in range(1,num_ko+1):
								if(ccn[i]==main_s1.cell(row=ko, column=5).value):
									main_s1.cell(row=ko, column=18).value=r[i]
									#print(r[i])'''
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s2.cell(row=i+main_row_count+1, column=5).value=p[i]
					main_s2.cell(row=i+main_row_count+1, column=6).value=np[i]
					main_s2.cell(row=i+main_row_count+1, column=8).value=r[i]
					main_s2.cell(row=i+main_row_count+1, column=7).value=s_v[i]
					main_s2.cell(row=i+main_row_count+1, column=9).value=v[i]

				for i in range(0,len(bp)):
					#lp=[main_s1.cell(row=i+temp_row+1, column=19).value]
					lk=[main_s1.cell(row=i+temp_row+1, column=18).value]
					#iu=['0.00' if v is None else v for v in lp]
					up=['0.00' if h is None else h for h in lk]
					#iu[0]=iu[0].replace(',','')
					up[0]=up[0].replace(',','')
					#print(det)
					#print(iu,up,det-float(iu[0])-float(up[0]))
					kop=str(bp[i])

					kop.replace(',','')
					#kop=['0.00' if h is None else h for h in kop]
					if kop=='None':
						kop='0.00'
					#print(kop, up[0])
					#print(float(float(kop)-float(up[0])))
					main_s1.cell(row=i+temp_row+1, column=55).value=float(float(kop)-float(up[0]))



	#paramount

		if ins=='Paramount':
			wbkName  = sys.argv[4]
			mypath=os.getcwd()+wbkName
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row

				ccn=[]
				b=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				oe=[]
				pe=[]
				cb=[]
				g_n=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					oe.append(s1.cell(row=t, column=12).value)
					pe.append(s1.cell(row=t, column=13).value)
					b.append(s1.cell(row=t, column=14).value)
					cb.append(s1.cell(row=t, column=16).value)
					g_n.append(s1.cell(row=t, column=17).value)
				#print(b,cb)
		#insert
				main_row_count = main_s1.max_row
				temp_row=main_row_count
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=6).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=9).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=7).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=23).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=11).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=13).value=e[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=eo[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=66).value=oe[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=pe[i]
					main_s1.cell(row=i+main_row_count+1, column=40).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=cb[i]

				b=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				for t in range(2,row_count+1):
					p.append(s2.cell(row=t, column=3).value)
					np.append(s2.cell(row=t, column=4).value)
					r.append(s2.cell(row=t, column=5).value)
					s_v.append(s2.cell(row=t, column=6).value)
					we.append(s2.cell(row=t, column=7).value)
					b.append(s2.cell(row=t, column=8).value)
					ew.append(s2.cell(row=t, column=9).value)
				for i in range(0,len(b)):
					main_s1.cell(row=i+main_row_count+1, column=14).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=41).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=42).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=55).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=18).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=43).value=we[i]
					#main_s1.cell(row=i+main_row_count+1, column=16).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=16).value=round(float(ew[i])/10)  #TDS
					main_s1.cell(row=i+main_row_count+1, column=17).value=float(ew[i])-round(float(ew[i])/10) #netpayable

				'''for i in range(0,len(np)):
					lp=[main_s1.cell(row=i+temp_row+1, column=19).value]
					lk=[main_s1.cell(row=i+temp_row+1, column=18).value]
					iu=['0.00' if v is None else v for v in lp]
					up=['0.00' if h is None else h for h in lk]
					iu[0]=iu[0].replace(',','')
					up[0]=up[0].replace(',','')
					#print(det)
					#print(iu,up,det-float(iu[0])-float(up[0]))
					kop=str(np[i])
					kop.replace(',','')
					main_s1.cell(row=i+temp_row+1, column=55).value=float(float(kop)-float(iu[0])-float(up[0]))
					print(np)'''
				p=[]
				np=[]
				r=[]
				s_v=[]
				kl=[]
				for wd in wbk.worksheets[2:]:
					row_count = wd.max_row
					for t in range(2,row_count+1):
						kl.append(wd.cell(row=t, column=2).value)
						p.append(wd.cell(row=t, column=4).value)
						np.append(wd.cell(row=t, column=5).value)
						r.append(wd.cell(row=t, column=6).value)
						s_v.append(wd.cell(row=t, column=7).value)
				main_row_count = main_s2.max_row
				for i in range(0,len(p)):
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=kl[i]
					main_s2.cell(row=i+main_row_count+1, column=5).value=p[i]
					main_s2.cell(row=i+main_row_count+1, column=6).value=np[i]
					main_s2.cell(row=i+main_row_count+1, column=8).value=r[i]
					main_s2.cell(row=i+main_row_count+1, column=9).value=s_v[i]


	#MDINDIA 11 23
		if ins=='MDINDIA':
			wbkName  = sys.argv[4]
			#wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)

			s1=wbk.worksheets[0]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				s2=wbk.worksheets[1]
				s3=wbk.worksheets[2]
				s4=wbk.worksheets[3]
				row_count = s1.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				b=[]
				bp=[]
				be=[]
				pe=[]
				qw=[]
				re=[]
				ks=[]
				sk=[]
				e_id=[]
				e_name=[]
				mo=[]
				ko=[]
				po=[]
				jo=[]
				utr=[]
				tran=[]
				doa=[]
				dod=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					b.append(s1.cell(row=t, column=12).value)
					utr.append(s1.cell(row=t, column=13).value)
					tran.append(s1.cell(row=t, column=14).value)
					doa.append(s1.cell(row=t, column=15).value)
					dod.append(s1.cell(row=t, column=16).value)
					bp.append(s2.cell(row=t, column=3).value )
					be.append(s2.cell(row=t, column=4).value)
					pe.append(s2.cell(row=t, column=5).value)
					qw.append(s2.cell(row=t, column=6).value)
					re.append(s2.cell(row=t, column=7).value)
					ks.append(s2.cell(row=t, column=8).value)
					e_id.append(s2.cell(row=t, column=9).value)
					e_name.append(s2.cell(row=t, column=10).value)
					mo.append(s2.cell(row=t, column=11).value)
					ko.append(s2.cell(row=t, column=12).value)
					po.append(s2.cell(row=t, column=13).value)
					jo.append(s2.cell(row=t, column=14).value)

		#insert
				main_row_count = main_s1.max_row
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=7).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=9).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=6).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=23).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=22).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=12).value=e[i]
					main_s1.cell(row=i+main_row_count+1, column=11).value=eo[i]
					main_s1.cell(row=i+main_row_count+1, column=10).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=31).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=13).value=utr[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=tran[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=doa[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=dod[i]
					main_s1.cell(row=i+main_row_count+1, column=32).value=bp[i]
					main_s1.cell(row=i+main_row_count+1, column=14).value=be[i]
					main_s1.cell(row=i+main_row_count+1, column=55).value=pe[i]
					main_s1.cell(row=i+main_row_count+1, column=18).value=qw[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=re[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=ks[i]
					main_s1.cell(row=i+main_row_count+1, column=17).value=e_id[i]
					main_s1.cell(row=i+main_row_count+1, column=34).value=e_name[i]
					main_s1.cell(row=i+main_row_count+1, column=35).value=mo[i]
					main_s1.cell(row=i+main_row_count+1, column=16).value=ko[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=po[i]
					main_s1.cell(row=i+main_row_count+1, column=36).value=jo[i]

				row_count = s3.max_row
				ccn=[]
				p=[]
				np=[]

				for t in range(2,row_count+1):
					ccn.append(s3.cell(row=t, column=2).value )
					p.append(s3.cell(row=t, column=3).value)
					np.append(s3.cell(row=t, column=4).value)
				#print(v)
		#insert
				main_row_count = main_s2.max_row
				for i in range(0,len(np)):
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s2.cell(row=i+main_row_count+1, column=8).value=p[i]
					main_s2.cell(row=i+main_row_count+1, column=9).value=np[i]

				row_count = s4.max_row
				ccn=[]
				p=[]
				np=[]
				rt=[]
				for t in range(2,row_count+1):
					ccn.append(s4.cell(row=t, column=2).value )
					p.append(s4.cell(row=t, column=3).value)
					np.append(s4.cell(row=t, column=4).value)
					rt.append(s4.cell(row=t, column=5).value)
				#print(v)
		#insert
				main_row_count = main_s2.max_row
				for i in range(0,len(np)):
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s2.cell(row=i+main_row_count+1, column=5).value=p[i]
					main_s2.cell(row=i+main_row_count+1, column=10).value=np[i]
					main_s2.cell(row=i+main_row_count+1, column=9).value=rt[i]
	#universal_sompo
		if ins=='Universal_Sompo':
			wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				b=[]
				bp=[]
				be=[]
				pe=[]
				qw=[]
				re=[]
				ks=[]
				sk=[]
				e_id=[]
				e_name=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					b.append(s1.cell(row=t, column=12).value)
					bp.append(s1.cell(row=t, column=13).value )
					be.append(s1.cell(row=t, column=14).value)
					pe.append(s1.cell(row=t, column=15).value)
					qw.append(s1.cell(row=t, column=16).value)
					re.append(s1.cell(row=t, column=17).value)
					ks.append(s1.cell(row=t, column=18).value)
					e_id.append(s1.cell(row=t, column=19).value)
					e_name.append(s1.cell(row=t, column=20).value)
				#print(ccn)
		#insert
				main_row_count = main_s1.max_row
				temp_row=main_row_count
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=21).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=12).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=22).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=10).value=e[i]
					main_s1.cell(row=i+main_row_count+1, column=11).value=eo[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=13).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=bp[i]
					main_s1.cell(row=i+main_row_count+1, column=14).value=be[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=pe[i]
					main_s1.cell(row=i+main_row_count+1, column=17).value=qw[i]
					main_s1.cell(row=i+main_row_count+1, column=18).value=re[i]
					main_s1.cell(row=i+main_row_count+1, column=19).value=ks[i]
					main_s1.cell(row=i+main_row_count+1, column=55).value=e_id[i]
					main_s1.cell(row=i+main_row_count+1, column=16).value=qw[i]
					main_s1.cell(row=i+main_row_count+1, column=17).value=re[i]
					main_s1.cell(row=i+main_row_count+1, column=18).value=ks[i]
					main_s1.cell(row=i+main_row_count+1, column=19).value=0
					#main_s1.cell(row=i+main_row_count+1, column=).value=e_name[i]
				row_count = s2.max_row
				ccn=[]
				p=[]
				np=[]
				po=[]
				for t in range(2,row_count+1):
					ccn.append(s2.cell(row=t, column=2).value )
					p.append(s2.cell(row=t, column=3).value)
					np.append(s2.cell(row=t, column=4).value)
					po.append(s2.cell(row=t, column=5).value)
				#print(v)
		#insert
				main_row_count = main_s2.max_row
				for i in range(0,len(np)):
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s2.cell(row=i+main_row_count+1, column=5).value=p[i]
					main_s2.cell(row=i+main_row_count+1, column=8).value=np[i]
					main_s2.cell(row=i+main_row_count+1, column=9).value=po[i]
	#vipul
		if ins=='vipul':
			wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				b=[]
				bp=[]
				be=[]
				pe=[]
				qw=[]
				km=[]
				mk=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					b.append(s1.cell(row=t, column=12).value)
					bp.append(s1.cell(row=t, column=13).value )
					be.append(s1.cell(row=t, column=14).value)
					pe.append(s1.cell(row=t, column=15).value)
					qw.append(s1.cell(row=t, column=16).value)
					km.append(s1.cell(row=t, column=17).value)
					mk.append(s1.cell(row=t, column=18).value)
				#print(ccn)
		#insert
				main_row_count = main_s1.max_row
				temp_row=main_row_count
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=23).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=14).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=55).value=float(e[i])+float(ro[i])
					main_s1.cell(row=i+main_row_count+1, column=19).value=eo[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=ro[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=18).value=bp[i]
					#main_s1.cell(row=i+main_row_count+1, column=14).value=be[i]
					main_s1.cell(row=i+main_row_count+1, column=16).value=pe[i]
					main_s1.cell(row=i+main_row_count+1, column=17).value=qw[i]
					main_s1.cell(row=i+main_row_count+1, column=13).value=km[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=mk[i]
				row_count = s2.max_row
				ccn=[]
				p=[]
				np=[]
				po=[]
				for t in range(2,row_count+1):
					ccn.append(s2.cell(row=t, column=2).value )
					p.append(s2.cell(row=t, column=3).value)
					np.append(s2.cell(row=t, column=4).value)
				#print(v)
		#insert
				main_row_count = main_s2.max_row
				for i in range(0,len(np)):
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s2.cell(row=i+main_row_count+1, column=8).value=p[i]
					main_s2.cell(row=i+main_row_count+1, column=9).value=np[i]

	#reliance
		if ins=='reliance':
			wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				b=[]
				bp=[]
				be=[]
				pe=[]
				qw=[]
				re=[]
				ks=[]
				sk=[]
				e_id=[]
				e_name=[]
				mp=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					b.append(s1.cell(row=t, column=12).value)
					bp.append(s1.cell(row=t, column=13).value )
					be.append(s1.cell(row=t, column=14).value)
					pe.append(s1.cell(row=t, column=15).value)
					qw.append(s1.cell(row=t, column=16).value)
					re.append(s1.cell(row=t, column=17).value)
					ks.append(s1.cell(row=t, column=18).value)
					e_id.append(s1.cell(row=t, column=19).value)
					e_name.append(s1.cell(row=t, column=20).value)
					mp.append(s1.cell(row=t, column=21).value)
				#print(ccn)
		#insert
				main_row_count = main_s1.max_row
				temp_row=main_row_count
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=7).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=23).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=6).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=12).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=e[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=eo[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=b[i]
					#main_s1.cell(row=i+main_row_count+1, column=64).value=bp[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=be[i]
					main_s1.cell(row=i+main_row_count+1, column=13).value=pe[i]
					main_s1.cell(row=i+main_row_count+1, column=14).value=qw[i]
					main_s1.cell(row=i+main_row_count+1, column=19).value=re[i]
					main_s1.cell(row=i+main_row_count+1, column=18).value=ks[i]
					main_s1.cell(row=i+main_row_count+1, column=17).value=mp[i]
					main_s1.cell(row=i+main_row_count+1, column=16).value=e_name[i]
				row_count = s2.max_row
				ccn=[]
				p=[]
				np=[]
				po=[]
				ep=[]
				eo=[]
				for t in range(2,row_count+1):
					ccn.append(s2.cell(row=t, column=2).value )
					p.append(s2.cell(row=t, column=3).value)
					ep.append(s2.cell(row=t, column=4).value)
					eo.append(s2.cell(row=t, column=5).value)
					np.append(s2.cell(row=t, column=6).value)
					po.append(s2.cell(row=t, column=7).value)
				#print(v)
		#insert
				cpy=[]
				ded=[]
				main_row_count = main_s2.max_row
				for i in range(0,len(np)):
					if p[i]=='Total':
						cpy.append(ccn[i])
						ded.append(np[i])
					else:
						main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
						main_s2.cell(row=i+main_row_count+1, column=3).value=ins
						main_s2.cell(row=i+main_row_count+1, column=4).value=ccn[i]
						main_s2.cell(row=i+main_row_count+1, column=5).value=p[i]
						main_s2.cell(row=i+main_row_count+1, column=6).value=ep[i]
						main_s2.cell(row=i+main_row_count+1, column=7).value=eo[i]
						main_s2.cell(row=i+main_row_count+1, column=8).value=np[i]
						main_s2.cell(row=i+main_row_count+1, column=9).value=po[i]

				for i in range(0,len(ded)):
					for j in range(temp_row,main_s1.max_row+1):
						#print((cpy[i],main_s1.cell(row=j, column=4).value))
						if(cpy[i]==main_s1.cell(row=j, column=4).value):
							main_s1.cell(row=j, column=55).value=ded[i]
							print(j,ded[i])
	#raksha
		if ins=='raksha':
			wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				b=[]
				bp=[]
				be=[]
				pe=[]
				qw=[]
				re=[]
				ks=[]
				sk=[]
				e_id=[]
				e_name=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					b.append(s1.cell(row=t, column=12).value)
					bp.append(s1.cell(row=t, column=13).value )
					be.append(s1.cell(row=t, column=16).value)
					pe.append(s1.cell(row=t, column=17).value)
					qw.append(s1.cell(row=t, column=18).value)
					re.append(s1.cell(row=t, column=19).value)
					ks.append(s1.cell(row=t, column=20).value)
					e_id.append(s1.cell(row=t, column=22).value)
					e_name.append(s1.cell(row=t, column=23).value)
				#print(ccn)
		#insert
				main_row_count = main_s1.max_row
				temp_row=main_row_count
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=6).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=23).value=np[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=14).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=17).value=e[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=eo[i]
					main_s1.cell(row=i+main_row_count+1, column=12).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=9).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=7).value=bp[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=be[i]
					main_s1.cell(row=i+main_row_count+1, column=55).value=pe[i]
					main_s1.cell(row=i+main_row_count+1, column=16).value=qw[i]
					main_s1.cell(row=i+main_row_count+1, column=13).value=re[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=ks[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=e_id[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=e_name[i]
				row_count = s2.max_row
				ccn=[]
				p=[]
				np=[]
				po=[]
				ep=[]
				eo=[]
				for t in range(2,row_count+1):
					ccn.append(s2.cell(row=t, column=2).value )
					p.append(s2.cell(row=t, column=3).value)
					ep.append(s2.cell(row=t, column=4).value)
					eo.append(s2.cell(row=t, column=5).value)
					np.append(s2.cell(row=t, column=6).value)
					po.append(s2.cell(row=t, column=7).value)
				#print(v)
		#insert
				main_row_count = main_s2.max_row
				for i in range(0,len(np)):
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s2.cell(row=i+main_row_count+1, column=5).value=p[i]
					main_s2.cell(row=i+main_row_count+1, column=6).value=ep[i]
					main_s2.cell(row=i+main_row_count+1, column=8).value=eo[i]
					main_s2.cell(row=i+main_row_count+1, column=7).value=np[i]
					main_s2.cell(row=i+main_row_count+1, column=9).value=po[i]
			wb.save(wbName)

	#medsave
		if ins=='medsave':
			wbkName  = sys.argv[4]
			mypath=os.getcwd()+wbkName
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row

				ccn=[]
				b=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				oe=[]
				pe=[]
				cb=[]
				g_n=[]
				md=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					oe.append(s1.cell(row=t, column=12).value)
					pe.append(s1.cell(row=t, column=13).value)
					b.append(s1.cell(row=t, column=14).value)
					cb.append(s1.cell(row=t, column=15).value)
					g_n.append(s1.cell(row=t, column=16).value)
					md.append(s1.cell(row=t, column=17).value)
				main_row_count = main_s1.max_row
				temp_row=main_row_count
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=p[i]
					main_s1.cell(row=i+main_row_count+1, column=23).value=np[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=7).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=6).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=13).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=9).value=e[i]
					main_s1.cell(row=i+main_row_count+1, column=17).value=eo[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=10).value=oe[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=pe[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=cb[i]
					main_s1.cell(row=i+main_row_count+1, column=14).value=g_n[i]
					main_s1.cell(row=i+main_row_count+1, column=12).value=md[i]

				row_count = s2.max_row
				ccn=[]
				p=[]
				np=[]
				po=[]
				for t in range(2,row_count+1):
					ccn.append(s2.cell(row=t, column=2).value )
					p.append(s2.cell(row=t, column=3).value)
					np.append(s2.cell(row=t, column=4).value)
					po.append(s2.cell(row=t, column=5).value)
				#print(np)
		#insert
				cpy=[]
				ded=[]
				main_row_count = main_s2.max_row
				for i in range(0,len(np)):
					if np[i]=='TOTAL:':
						cpy.append(ccn[i])
						ded.append(po[i])
					else:
						main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
						main_s2.cell(row=i+main_row_count+1, column=3).value=ins
						main_s2.cell(row=i+main_row_count+1, column=4).value=ccn[i]
						main_s2.cell(row=i+main_row_count+1, column=5).value=p[i]
						main_s2.cell(row=i+main_row_count+1, column=8).value=np[i]
						main_s2.cell(row=i+main_row_count+1, column=9).value=po[i]
				for i in range(0,len(ded)):
					for j in range(temp_row,main_s1.max_row+1):
						#print((cpy[i],main_s1.cell(row=j, column=4).value))
						if(cpy[i]==main_s1.cell(row=j, column=4).value):
							main_s1.cell(row=j, column=55).value=ded[i]
							#print(j,ded[i])
	#hitpa
		if ins=='health_insurance':
			wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				b=[]
				bp=[]
				be=[]
				pe=[]
				qw=[]
				re=[]
				ks=[]
				sk=[]
				e_id=[]
				e_name=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					b.append(s1.cell(row=t, column=12).value)
					bp.append(s1.cell(row=t, column=13).value )
					be.append(s1.cell(row=t, column=14).value)
					pe.append(s1.cell(row=t, column=15).value)
					qw.append(s1.cell(row=t, column=16).value)
					re.append(s1.cell(row=t, column=17).value)
					ks.append(s1.cell(row=t, column=18).value)
					e_id.append(s1.cell(row=t, column=19).value)
				#print(ccn)
		#insert
				main_row_count = main_s1.max_row
				temp_row=main_row_count
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=9).value=p[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=np[i]
					main_s1.cell(row=i+main_row_count+1, column=12).value=r[i]
					main_s1.cell(row=i+main_row_count+1, column=7).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=e[i]
					main_s1.cell(row=i+main_row_count+1, column=6).value=eo[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=13).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=14).value=bp[i]
					main_s1.cell(row=i+main_row_count+1, column=55).value=float(be[i])+float(re[i])
					main_s1.cell(row=i+main_row_count+1, column=18).value=pe[i]
					main_s1.cell(row=i+main_row_count+1, column=19).value=qw[i]
					main_s1.cell(row=i+main_row_count+1, column=16).value=ks[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=e_id[i]
				row_count = s2.max_row
				ccn=[]
				p=[]
				np=[]
				po=[]
				ep=[]
				eo=[]
				for t in range(2,row_count+1):
					ccn.append(s2.cell(row=t, column=2).value )
					p.append(s2.cell(row=t, column=3).value)
					ep.append(s2.cell(row=t, column=5).value)
					eo.append(s2.cell(row=t, column=6).value)
					np.append(s2.cell(row=t, column=7).value)
					po.append(s2.cell(row=t, column=8).value)
				#print(v)
		#insert
				main_row_count = main_s2.max_row
				for i in range(0,len(np)):
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s2.cell(row=i+main_row_count+1, column=5).value=p[i]
					main_s2.cell(row=i+main_row_count+1, column=6).value=ep[i]
					main_s2.cell(row=i+main_row_count+1, column=7).value=eo[i]
					main_s2.cell(row=i+main_row_count+1, column=8).value=np[i]
					main_s2.cell(row=i+main_row_count+1, column=9).value=po[i]
	#Ease_West
		if ins=='east_west':
			wbkName  = sys.argv[4]
			wbk= openpyxl.load_workbook(wbkName)
			s1=wbk.worksheets[0]
			s2=wbk.worksheets[1]
			row_count = s1.max_row
			if s1.cell(row=1, column=1).value!=None:
				row_count = s1.max_row
				ccn=[]
				p=[]
				np=[]
				r=[]
				s_v=[]
				we=[]
				ew=[]
				e=[]
				eo=[]
				ro=[]
				b=[]
				bp=[]
				be=[]
				pe=[]
				qw=[]
				re=[]
				ks=[]
				sk=[]
				e_id=[]
				e_name=[]
				for t in range(2,row_count+1):
					ccn.append(s1.cell(row=t, column=2).value )
					if(s1.cell(row=t, column=1).value=='error'):
						mf_row = main_s5.max_row
						main_s5.cell(row=mf_row+1, column=1).value=ins
						main_s5.cell(row=mf_row+1, column=2).value=ccn[-1]
						main_s5.cell(row=mf_row+1, column=4).value='Need assistance'
					p.append(s1.cell(row=t, column=3).value)
					np.append(s1.cell(row=t, column=4).value)
					r.append(s1.cell(row=t, column=5).value)
					s_v.append(s1.cell(row=t, column=6).value)
					we.append(s1.cell(row=t, column=7).value )
					ew.append(s1.cell(row=t, column=8).value)
					e.append(s1.cell(row=t, column=9).value)
					eo.append(s1.cell(row=t, column=10).value)
					ro.append(s1.cell(row=t, column=11).value)
					b.append(s1.cell(row=t, column=12).value)
					bp.append(s1.cell(row=t, column=13).value )
					be.append(s1.cell(row=t, column=14).value)
					pe.append(s1.cell(row=t, column=15).value)
					qw.append(s1.cell(row=t, column=16).value)
					re.append(s1.cell(row=t, column=17).value)
					ks.append(s1.cell(row=t, column=18).value)
					e_id.append(s1.cell(row=t, column=19).value)
					e_name.append(s1.cell(row=t, column=20).value)
				#print(ccn)
		#insert
				main_row_count = main_s1.max_row
				temp_row=main_row_count
				for i in range(0,len(ccn)):
					main_s1.cell(row=i+main_row_count+1, column=1).value=sys.argv[3]
					main_s1.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s1.cell(row=i+main_row_count+1, column=3).value=ins
					main_s1.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=5).value=ccn[i]
					main_s1.cell(row=i+main_row_count+1, column=8).value=p[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=np[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=r[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=s_v[i]
					main_s1.cell(row=i+main_row_count+1, column=9).value=we[i]
					main_s1.cell(row=i+main_row_count+1, column=12).value=ew[i]
					main_s1.cell(row=i+main_row_count+1, column=7).value=e[i]
					main_s1.cell(row=i+main_row_count+1, column=6).value=eo[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=ro[i]
					main_s1.cell(row=i+main_row_count+1, column=38).value=b[i]
					main_s1.cell(row=i+main_row_count+1, column=39).value=bp[i]
					main_s1.cell(row=i+main_row_count+1, column=64).value=be[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=pe[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=qw[i]
					#main_s1.cell(row=i+main_row_count+1, column=).value=re[i]
					main_s1.cell(row=i+main_row_count+1, column=14).value=ks[i]
					main_s1.cell(row=i+main_row_count+1, column=15).value=e_id[i]
					main_s1.cell(row=i+main_row_count+1, column=55).value=e_name[i]
				row_count = s2.max_row
				ccn=[]
				p=[]
				np=[]
				po=[]
				ep=[]
				eo=[]
				for t in range(2,row_count+1):
					ccn.append(s2.cell(row=t, column=2).value )
					p.append(s2.cell(row=t, column=3).value)
					ep.append(s2.cell(row=t, column=4).value)
					eo.append(s2.cell(row=t, column=5).value)
					np.append(s2.cell(row=t, column=6).value)
					po.append(s2.cell(row=t, column=7).value)
				#print(v)
		#insert
				main_row_count = main_s2.max_row
				for i in range(0,len(np)):
					main_s2.cell(row=i+main_row_count+1, column=2).value=k[3]
					main_s2.cell(row=i+main_row_count+1, column=3).value=ins
					main_s2.cell(row=i+main_row_count+1, column=4).value=ccn[i]
					main_s2.cell(row=i+main_row_count+1, column=5).value=p[i]
					main_s2.cell(row=i+main_row_count+1, column=6).value=ep[i]
					main_s2.cell(row=i+main_row_count+1, column=8).value=eo[i]
					main_s2.cell(row=i+main_row_count+1, column=7).value=np[i]
					main_s2.cell(row=i+main_row_count+1, column=9).value=po[i]

		if ins == 'bajaj':
			# wbkName  = sys.argv[4]
			# wbk= openpyxl.load_workbook(wbkName)

			# book = load_workbook(wbkName)
			# sheetlist = book.get_sheet_names()
			# sheet = book.get_sheet_by_name(sheetlist[0])
			# row_count = sheet.max_row

			# utr_no = sheet.cell(row=2, column=8).value
			# alno = claim_no = sheet.cell(row=2, column=2).value
			# patient_name = sheet.cell(row=2, column=3).value
			# tds = sheet.cell(row=2, column=6).value
			# netpayable = sheet.cell(row=2, column=5).value
			# tran_date = sheet.cell(row=2, column=7).value

			# main_row_count = main_s1.max_row
			# main_s1.cell(row=i + main_row_count + 1, column=2).value = k[3]
			# main_s1.cell(row=i + main_row_count + 1, column=3).value = ins
			# main_s1.cell(row=i + main_row_count + 1, column=4).value = alno
			# main_s1.cell(row=i + main_row_count + 1, column=5).value = claim_no
			# main_s1.cell(row=i + main_row_count + 1, column=8).value = patient_name
			# main_s1.cell(row=i + main_row_count + 1, column=13).value = utr_no
			# main_s1.cell(row=i + main_row_count + 1, column=16).value = tds
			# main_s1.cell(row=i + main_row_count + 1, column=17).value = netpayable
			# main_s1.cell(row=i + main_row_count + 1, column=64).value = tran_date


			wbkName  = sys.argv[4]
			wbk = openpyxl.load_workbook(wbkName)
			book = load_workbook(wbkName)
			sheetlist = book.get_sheet_names()
			sheet = book.worksheets[0]
			sheet1 = book.worksheets[1]
			a = sheet1.max_row
			b = main_s2.max_column
			c = sheet1[sheet1.max_row][3].value
			a = main_s1[2][1].value
			for i in range(2, sheet.max_row+1):
				rowno = main_s1.max_row
				dtrow = sheet[i]
				main_s1[rowno][0].value=sys.argv[3]
				main_s1[rowno][1].value = k[3]
				main_s1[rowno][2].value = ins
				main_s1[rowno][3].value = dtrow[2].value
				main_s1[rowno][4].value = dtrow[3].value
				main_s1[rowno][5].value = dtrow[1].value
				main_s1[rowno][7].value = dtrow[0].value
				main_s1[rowno][8].value = ins
				main_s1[rowno][9].value = dtrow[1].value
				main_s1[rowno][12].value = dtrow[7].value
				main_s1[rowno][13].value = dtrow[8].value
				main_s1[rowno][14].value = dtrow[9].value
				main_s1[rowno][15].value = dtrow[11].value
				main_s1[rowno][16].value = dtrow[9].value
				main_s1[rowno][17].value = sheet1[sheet1.max_row][3].value
				main_s1[rowno][37].value = dtrow[4].value
				main_s1[rowno][38].value = dtrow[5].value
				# main_s1[rowno][54].value = int(dtrow[10].value)-int(sheet1[sheet1.max_row][3].value)
				main_s1[rowno][54].value = int(0)
				main_s1[rowno][63].value = dtrow[12].value
				a = main_s1[rowno][6].value

			for i in range(2, sheet1.max_row+1):
				rowno = main_s2.max_row+1
				dtrow = sheet1[i]
				main_s2[rowno][0].value = ''#rowno-1
				main_s2[rowno][1].value = 'inamdar'#k[3]
				main_s2[rowno][2].value = 'bajaj'#ins
				main_s2[rowno][3].value = main_s1[main_s1.max_row][4].value
				main_s2[rowno][4].value = dtrow[1].value
				main_s2[rowno][5].value = dtrow[2].value
				main_s2[rowno][6].value = dtrow[4].value
				main_s2[rowno][7].value = dtrow[3].value
				main_s2[rowno][8].value = dtrow[5].value
				main_s2[rowno][9].value = ''#discout amt



		row_count = main_s1.max_row
		l={'01':'Jan','02':'Feb','03':'Mar','04':'Apr', '05':'May', '06':'Jun', '07':'Jul', '08':'Aug', '09':'Sep', '10':'Oct', '11':'Nov', '12':'Dec'}
		x=[38,39,64]
		for j in x:
			for i in range(2,row_count+1):
				try:
					d=str(main_s1.cell(row=i, column=j).value)
					if(d!='None'):
						#print([d])
						d=d.replace('  ','')
						if d.startswith(' '):
							d=d[1:]
						#print(d,i,j)
						v=d[1]
						if(v>='a' and v<='z' or v>='A' and v<='Z'):
							d='0'+d[:1]+'-'+d[1:]

						n=d[2]
						if(n>='a' and n<='z' or n>='A' and n<='Z' and v!=' '):
							d=d[:2]+'-'+d[2:]
						d=d.replace(' ','-')
						d=d.replace(' ','-')
						d=d.replace('/','-')
						h=d.find('-')
						if(h==1):
							d='0'+d
						h=d.find('-')
						g=d[h+1:]
						#print(g)
						k=g.find('-')
						g=g[k+1:]
						#print(g)
						m=g.find('-')
						#print(m)
						if m!=-1:
							d=d[:m+k+h+2]
						for val,key in l.items():
							d = d.replace(key, val)
							d = d.replace(key.lower(),val)
							d = d.replace(key.upper(),val)

						n=d[5]
						if(n!='-'):
							d=d[:5]+'-'+d[5:]
						if(d[-3]=='-'):
							d=d[:-2]+'20'+d[-2:]
						d=d[:11]
						if(d[3:5]>'12'):
							d=d[3:5]+'-'+d[0:2]+'-'+d[6:10]
						if(d[-1]=='-'):
								d=d[:-1]
						main_s1.cell(row=i, column=j).value=d
				except Exception as e:
					#s1.cell(row=i, column=1).value = 'error'
					print(i,s1.cell(row=i, column=4).value)

	#seaching copay and discount
	'''
	def search_copay(i,d,r,c):
		pos=[]
		amt=[]
		gh=0
		co=['co-pay','copay','Co-Payment()','Co-Payment','Paid by Patient','CoPayment','Member Paid']
		for x in range(0,len(r)):
			#print(r[x])
			if r[x]!=None:
				#print(r[x])
				for j in co:	
					if r[x].find(j)!=-1:
						amt.append(d[x])
						pos.append(i[x]) 
						break
		#print(amt)			
		for x in range (0,len(pos)):
			k=0
			for j in range (0,len(c)):
				try:
					if pos[x] in c[j]:
						k=1
						if(main_s1.cell(row=j+3, column=19).value!=amt[x]):
							if(main_s1.cell(row=j+3, column=19).value!=None):
								#print(amt[x])
								main_s1.cell(row=j+3, column=19).value=float(main_s1.cell(row=j+3, column=19).value)+float(amt[x])
							else:
								#print(amt[x])
								main_s1.cell(row=j+3, column=19).value=amt[x]
								gh+=1
							main_s1.cell(row=j+3, column=55).value=float(main_s1.cell(row=j+3, column=55).value)-float(amt[x])
						else:
							main_s1.cell(row=j+3, column=55).value=float(main_s1.cell(row=j+3, column=55).value)-float(amt[x])		
				except Exception as e:
					continue
		print(gh,len(pos))
	
	def search_discount(i,d,r,c):
		gh=0
		pos=[]
		amt=[]
		print('dis')
		co=['discount','Discount']
		for x in range(0,len(r)):
			#print(r[x])
			if r[x]!=None:
				#print(r[x])
				for j in co:	
					if r[x].find(j)!=-1:
						amt.append(d[x])
						pos.append(i[x]) 
		for x in range (0,len(pos)):
			for j in range (0,len(c)):
				try:
	
					if pos[x] in c[j]:
						if(main_s1.cell(row=j+3, column=18).value!=amt[x]):
							main_s1.cell(row=j+3, column=18).value=amt[x]
							gh+=1
						main_s1.cell(row=j+3, column=55).value=float(main_s1.cell(row=j+3, column=55).value)-float(amt[x])
				except Exception as e:
					continue					
		print(gh,len(pos))
	'''
	mol=[]
	import re
	def str_num_s1(jk):
		for i in range(3,main_s1.max_row+1):
			if main_s1.cell(row=i, column=jk).value!=None:
				string=str(main_s1.cell(row=i, column=jk).value)
				try:
					jok=[float(re.search(r"\d+\.\d+", string).group())]
					main_s1.cell(row=i, column=jk).value=jok[0]
					#print('right',jok[0])
				except AttributeError:
					try:
						#print(i)
						koj=[int(re.search(r'\d+', string).group())]
						main_s1.cell(row=i, column=jk).value=koj[0]
						#print('right',koj[0])
					except AttributeError:
						main_s1.cell(row=i, column=jk).value=0
						#print('wrong',string)
			else:
				main_s1.cell(row=i, column=jk).value=0
	def str_num_s2(jk):
		for i in range(2,main_s2.max_row+1):
			if main_s2.cell(row=i, column=jk).value!=None:
				string=str(main_s2.cell(row=i, column=jk).value)
				try:
					jok=[float(re.search(r"\d+\.\d+", string).group())]
					main_s2.cell(row=i, column=jk).value=jok[0]
					#print('right',jok[0])
				except AttributeError:
					try:
						#print(i)
						koj=[int(re.search(r'\d+', string).group())]
						main_s2.cell(row=i, column=jk).value=koj[0]
						#print('right',koj[0])
					except AttributeError:
						main_s2.cell(row=i, column=jk).value=None
			else:
				main_s2.cell(row=i, column=jk).value=None
						#print('wrong',string)
	def deduction_none():
		kol=0
		for i in range(1,main_s2.max_row+1):
			if main_s2.cell(row=i, column=8).value==None:
					#main_s2.delete_rows(i, 1)
				mol.append(i-kol)
				kol+=1
			elif main_s2.cell(row=i, column=8).value==0 or main_s2.cell(row=i, column=8).value==0.00 or main_s2.cell(row=i, column=5).value=='Total' or main_s2.cell(row=i, column=5).value=='Final Approved Amount' or str(main_s2.cell(row=i, column=8).value)=='' or  main_s2.cell(row=i, column=5).value=='Total Payble Amount' or main_s2.cell(row=i, column=5).value=='Miscellaneous':
					#main_s2.delete_rows(i, 1)
				mol.append(i-kol)
				kol+=1
	bill=[]
	pay=[]
	deduction=[]
	sheet1_var=[]
	'''
	k=0
	for j in range(14,20):
		sheet1_var.append([])
		for i in range(3,main_s1.max_row+1):
			sheet1_var[k].append(str(main_s1.cell(row=i, column=j).value))
		k+=1
	#print(len(sheet1_var))
	for i in range(2,main_s2.max_row+1):
		bill.append(str(main_s2.cell(row=i, column=6).value))
		pay.append(str(main_s2.cell(row=i, column=7).value))
		deduction.append(str(main_s2.cell(row=i, column=8).value))
	'''
	str_num_s1(14)
	str_num_s1(15)
	str_num_s1(16)
	str_num_s1(17)
	str_num_s1(18)
	str_num_s1(19)
	str_num_s1(55)
	str_num_s2(6)
	str_num_s2(7)
	str_num_s2(8)
	deduction_none()
	for i in range(0,len(mol)):
		#print(main_s2.cell(row=mol[i], column=5).value,main_s2.cell(row=mol[i], column=8).value)
		main_s2.delete_rows(mol[i], 1)

	#print(deduction)
	max_row=main_s2.max_row

	cate=[]
	reason=[]
	for i in range(2,max_row+1):
		cate.append(main_s2.cell(row=i, column=5).value)
		reason.append(main_s2.cell(row=i, column=9).value)
		#print(str(deduction[1297:1302]))
	max_row=main_s1.max_row

	dg=['Final\xa0Approved\xa0Amount','Total','Total Payble Amount','Total Payble Amount.','Miscellaneous Charges','Miscellaneous\nCharges','Miscellaneous charges']
	for i in range(0,len(reason)):
		if(reason[i]==None or reason[i]==' ' and cate[i] not in dg):
			reason[i]=cate[i]
			main_s2.cell(row=i+2, column=9).value=cate[i]
	'''cvn=[]
	for i in range(3,max_row+1):
		cvn.append(main_s1.cell(row=i, column=5).value)
	search_copay(insurance,deduction,reason,cvn)
	search_discount(insurance,deduction,reason,cvn)
	'''
	temp=[]
	temp_d=[]
	for i in range (0,len(reason)):
		if reason[i] not in temp:
			temp.append(reason[i])
			temp_d.append(cate[i])

	'''
	'''

	dict_cat={}
	for k_num in range (0,len(temp)):
		i=temp[k_num]
		l_cat=temp_d[k_num]
		temp_i=i
		if i!=None:
			i=temp_i.replace('\n',' ')
		#print(i)
		if i==None and l_cat==None:
			dict_cat[temp_i]=''
		elif i.lower().find('co-pay')!=-1 or i.lower().find('copay')!=-1 or i.lower().find('co pay')!=-1 or  i.lower().find('co-payment' )!=-1:
			dict_cat[temp_i]='1'

		elif i.lower().find('nme')!=-1 or i.lower().find('non-consumable')!=-1 or i.lower().find('non payable')!=-1 or i.lower().find('not payable')!=-1:
			dict_cat[temp_i]='5'

		elif i.lower().find('disc')!=-1 and i.lower().find('discharge')==-1:
			dict_cat[temp_i]='6'

		elif i.lower().find('per soc')!=-1 or i.lower().find('tariff')!=-1 or i.lower().find('do not collect')!=-1 or i.lower().find('not to be collect')!=-1 or i.lower().find('not to collect')!=-1 or i.lower().find('as per mou')!=-1 or i.lower().find('as per ppn')!=-1 or i.lower().find('as per mini soc')!=-1 or i.lower().find('as per hospital')!=-1:
			dict_cat[temp_i]='7'


		elif i.lower().find('as per patient')!=-1 or i.lower().find('as per policy')!=-1 or i.lower().find('as per authorization')!=-1 or i.lower().find('excess room')!=-1:
			dict_cat[temp_i]='2'

		elif i.lower().find('paid by patient')!=-1 or  i.lower().find('room')!=-1:
			if l_cat!=None:
				if l_cat.lower().find('package')!=-1 or l_cat.lower().find('pkg')!=-1 or l_cat.lower().find('icu')!=-1 or l_cat.lower().find('hospital')!=-1 or l_cat.lower().find('consultant')!=-1 or l_cat.lower().find('room')!=-1:
					dict_cat[temp_i]='2'
				else:
					dict_cat[temp_i]='5'
			else:
				dict_cat[temp_i]='5'


		elif i.lower().find('member paid')!=-1 or i.lower().find('admin')!=-1 or i.lower().find('micro')!=-1 or i.lower().find('casu')!=-1 or i.lower().find('TPA')!=-1 or i.lower().find('payable')!=-1 and i.lower().find('not')!=-1 or i.lower().find('non')!=-1:
			dict_cat[temp_i]='5'

		elif i.lower().find('non')!=-1 and i.lower().find('med')!=-1 or i.lower().find('adm')!=-1:
			dict_cat[temp_i]='5'

		elif i.lower().find('exhausted')!=-1 or i.lower().find('gipsa')!=-1 or i.lower().find('ppn')!=-1 or i.lower().find('pkg')!=-1 or i.lower().find('over')!=-1 or i.lower().find('excess')!=-1 or i.lower().find('room')!=-1 or i.lower().find('limit')!=-1 or i.lower().find('exceed')!=-1:
			dict_cat[temp_i]='3'


		elif i.lower().find('tax')!=-1 or i.lower().find('tds')!=-1 or i.lower().find('gst')!=-1:
			dict_cat[temp_i]='8'

		elif i.lower().find('mou ')!=-1:
			dict_cat[temp_i]='6'


		else:
			if l_cat!=None and l_cat.lower().find('other')!=-1:
				dict_cat[temp_i]='4'
				print(l_cat,i)
			else:
				dict_cat[temp_i]='5'


	temp_k=1
	main_s2.cell(row=1, column=11).value='deduction category'
	main_row_count = main_s2.max_row
	for i in reason:
		main_s2.cell(row=temp_k+1, column=11).value=dict_cat[i]
		temp_k+=1
	print(len(dict_cat),len(temp))
	#print(temp,len(temp))
	d_cat=[]
	deduction=[]
	d_ccn=[]
	pol={}
	lim={}
	oth={}
	npay={}
	cop={}
	dis={}
	bil={}
	Tax={}

	max_row=main_s2.max_row
	for i in range(2,max_row+1):   #varun
		d_ccn.append(main_s2.cell(row=i, column=4).value)
		d_cat.append(main_s2.cell(row=i, column=11).value)
		deduction.append(main_s2.cell(row=i, column=8).value)
		pol[d_ccn[-1]]=0
		lim[d_ccn[-1]]=0
		oth[d_ccn[-1]]=0
		npay[d_ccn[-1]]=0
		cop[d_ccn[-1]]=0
		dis[d_ccn[-1]]=0
		bil[d_ccn[-1]]=0
		Tax[d_ccn[-1]]=0
	for i in range(0,len(d_cat)):
		if d_cat[i]=='2':
			pol[d_ccn[i]]+=deduction[i]
		if d_cat[i]=='3':
			lim[d_ccn[i]]+=deduction[i]
		if d_cat[i]=='4':
			oth[d_ccn[i]]+=deduction[i]
		if d_cat[i]=='5':
			npay[d_ccn[i]]+=deduction[i]
		if d_cat[i]=='1':
			cop[d_ccn[i]]+=deduction[i]
		if d_cat[i]=='6':
			dis[d_ccn[i]]+=deduction[i]
		if d_cat[i]=='7':
			bil[d_ccn[i]]+=deduction[i]
		if d_cat[i]=='8':
			Tax[d_ccn[i]]+=deduction[i]

	for i in range(2,main_s1.max_row+1):
		if main_s1.cell(row=i, column=4).value in d_ccn:
			main_s1.cell(row=i, column=80).value=pol[main_s1.cell(row=i, column=4).value]
			main_s1.cell(row=i, column=81).value=lim[main_s1.cell(row=i, column=4).value]
			main_s1.cell(row=i, column=82).value=npay[main_s1.cell(row=i, column=4).value]
			main_s1.cell(row=i, column=83).value=bil[main_s1.cell(row=i, column=4).value]
			main_s1.cell(row=i, column=84).value=oth[main_s1.cell(row=i, column=4).value]
			if main_s1.cell(row=i, column=19).value==None:
				main_s1.cell(row=i, column=19).value=cop[main_s1.cell(row=i, column=4).value]
			elif main_s1.cell(row=i, column=19).value!=cop[main_s1.cell(row=i, column=4).value]:
				#####################################################akshay
				# main_s1.cell(row=i, column=19).value=main_s1.cell(row=i, column=19).value+cop[main_s1.cell(row=i, column=4).value]
				if main_s1.cell(row=i, column=19).value is not None:
					main_s1.cell(row=i, column=19).value=float(main_s1.cell(row=i, column=19).value)+cop[main_s1.cell(row=i, column=4).value]
			#####################################################akshay
			# main_s1.cell(row=i, column=55).value=main_s1.cell(row=i, column=55).value-cop[main_s1.cell(row=i, column=4).value]
			if main_s1.cell(row=i, column=55).value is not None:
				main_s1.cell(row=i, column=55).value=float(main_s1.cell(row=i, column=55).value)-cop[main_s1.cell(row=i, column=4).value]
			#####################################################akshayend
			if main_s1.cell(row=i, column=18).value==None:
				main_s1.cell(row=i, column=18).value=dis[main_s1.cell(row=i, column=4).value]

			elif main_s1.cell(row=i, column=18).value!=dis[main_s1.cell(row=i, column=4).value] and main_s1.cell(row=i, column=3).value!='hdfc':
				#####################################################akshay
				# main_s1.cell(row=i, column=18).value=main_s1.cell(row=i, column=18).value+dis[main_s1.cell(row=i, column=4).value]
				if main_s1.cell(row=i, column=18).value is not None:
					main_s1.cell(row=i, column=18).value=float(main_s1.cell(row=i, column=18).value)+dis[main_s1.cell(row=i, column=4).value]
				#####################################################akshayend

			if main_s1.cell(row=i, column=3).value!='hdfc':
				if main_s1.cell(row=i, column=55).value is not None:
					main_s1.cell(row=i, column=55).value=float(main_s1.cell(row=i, column=55).value)-dis[main_s1.cell(row=i, column=4).value]

			if main_s1.cell(row=i, column=16).value==None:
				main_s1.cell(row=i, column=16).value=Tax[main_s1.cell(row=i, column=4).value]

			elif main_s1.cell(row=i, column=16).value!=Tax[main_s1.cell(row=i, column=4).value]:
				##################################################akshay
				# main_s1.cell(row=i, column=16).value=main_s1.cell(row=i, column=16).value+Tax[main_s1.cell(row=i, column=4).value]
				if main_s1.cell(row=i, column=16).value is not None:
					main_s1.cell(row=i, column=16).value=float(main_s1.cell(row=i, column=16).value)+Tax[main_s1.cell(row=i, column=4).value]
				##################################################akshayend
				#main_s1.cell(row=i, column=55).value=main_s1.cell(row=i, column=55).value-Tax[main_s1.cell(row=i, column=4).value]

		elif main_s1.cell(row=i, column=5).value in d_ccn:
			main_s1.cell(row=i, column=80).value=pol[main_s1.cell(row=i, column=5).value]
			main_s1.cell(row=i, column=81).value=lim[main_s1.cell(row=i, column=5).value]
			main_s1.cell(row=i, column=82).value=npay[main_s1.cell(row=i, column=5).value]
			main_s1.cell(row=i, column=83).value=bil[main_s1.cell(row=i, column=5).value]
			main_s1.cell(row=i, column=84).value=oth[main_s1.cell(row=i, column=5).value]
			if main_s1.cell(row=i, column=19).value==None:
				main_s1.cell(row=i, column=19).value=cop[main_s1.cell(row=i, column=5).value]
			elif main_s1.cell(row=i, column=19).value!=cop[main_s1.cell(row=i, column=5).value]:
				##################################################akshay
				# main_s1.cell(row=i, column=19).value=main_s1.cell(row=i, column=19).value+cop[main_s1.cell(row=i, column=5).value]
				if main_s1.cell(row=i, column=19).value is not None:
					main_s1.cell(row=i, column=19).value=float(main_s1.cell(row=i, column=19).value.replace(',',''))+cop[main_s1.cell(row=i, column=5).value]
				##################################################akshayend
			##################################################akshay
			# main_s1.cell(row=i, column=55).value=main_s1.cell(row=i, column=55).value-cop[main_s1.cell(row=i, column=5).value]
			if main_s1.cell(row=i, column=55).value is not None:
				main_s1.cell(row=i, column=55).value=float(main_s1.cell(row=i, column=55).value)-cop[main_s1.cell(row=i, column=5).value]
			##################################################akshayend
			if main_s1.cell(row=i, column=55).value is not None:
				main_s1.cell(row=i, column=55).value=float(main_s1.cell(row=i, column=55).value)-cop[main_s1.cell(row=i, column=5).value]

			if main_s1.cell(row=i, column=18).value==None:
				main_s1.cell(row=i, column=18).value=dis[main_s1.cell(row=i, column=5).value]

			elif main_s1.cell(row=i, column=18).value!=dis[main_s1.cell(row=i, column=5).value] and main_s1.cell(row=i, column=3).value!='hdfc':
				##############################ak
				# main_s1.cell(row=i, column=18).value=main_s1.cell(row=i, column=18).value+dis[main_s1.cell(row=i, column=5).value]
				if main_s1.cell(row=i, column=18) is not None:
					main_s1.cell(row=i, column=18).value=float(str(main_s1.cell(row=i, column=18).value).replace(',','').strip())+dis[main_s1.cell(row=i, column=5).value]
				##############################akend
			if main_s1.cell(row=i, column=3).value!='hdfc':
				if main_s1.cell(row=i, column=55).value is not None:
					main_s1.cell(row=i, column=55).value=float(main_s1.cell(row=i, column=55).value)-dis[main_s1.cell(row=i, column=5).value]

			if main_s1.cell(row=i, column=16).value==None:
				main_s1.cell(row=i, column=16).value=Tax[main_s1.cell(row=i, column=5).value]

			elif main_s1.cell(row=i, column=16).value!=Tax[main_s1.cell(row=i, column=5).value]:
				###########################ak
				# main_s1.cell(row=i, column=16).value=main_s1.cell(row=i, column=16).value+Tax[main_s1.cell(row=i, column=5).value]
				if main_s1.cell(row=i, column=16).value is not None:
					main_s1.cell(row=i, column=16).value=float(main_s1.cell(row=i, column=16).value)+Tax[main_s1.cell(row=i, column=5).value]
				#############################akend
				#main_s1.cell(row=i, column=55).value=main_s1.cell(row=i, column=55).value-Tax[main_s1.cell(row=i, column=5).value]
	#error count
	wb1 = openpyxl.load_workbook('count/count.xlsx')
	ws1 = wb1.worksheets[0]

	main_s3=wb.worksheets[2]

	mr = ws1.max_row
	mc = ws1.max_column
	temp_mr=main_s1.max_row
	ins_id=['','aditya_birla','bajaj','apollo_munich','fgh','fhpl','health_india','health_hertige','icici_lombard','MDINDIA','Medi_Assist','Paramount','religare','united','hdfc','vidal','universal_sompo','vipul','Ease_West','Good_health','medsave','raksha','reliance','health_insurance']
	t_mr = len(ins_id)
	for i in range (1, mr + 1):
		for j in range (1, mc + 1):
			c = ws1.cell(row = i, column = j)
			main_s3.cell(row = i, column = j).value = c.value
	s=[main_s1.cell(row = i, column = 3).value for i in range(1,temp_mr+1)]
	main_s3.cell(row =1, column =5).value ='Sheet_count'
	for i in range (1, t_mr):
		ch=s.count(ins_id[i])
		main_s3.cell(row = i+1, column =5).value =ch
	ws1 = wb1.worksheets[1]

	main_s4=wb.worksheets[3]

	mr = ws1.max_row
	mc = ws1.max_column

	for i in range (1, mr + 1):
		for j in range (1, mc + 1):
			c = ws1.cell(row = i, column = j)
			main_s4.cell(row = i, column = j).value = c.value


	print("Done")
	wb.save(wbName)
	wb.close()
except:
	log_exceptions()