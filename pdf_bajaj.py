import os
import re
import sqlite3
import subprocess
import sys
import time

import openpyxl
import pdftotext

from backend import mark_flag
from make_log import log_exceptions
from movemaster import move_master_to_master_insurer

try:
    pdfpath = sys.argv[1]
    onlyfiles = [os.path.split(pdfpath)[1]]
    mypath = os.path.dirname(pdfpath)+'/'

    hosp_name = ''

    with open(pdfpath, "rb") as f:
        pdf = pdftotext.PDF(f)
    with open('temp_files/output.txt', 'w', encoding='utf-8') as f:
        f.write(" ".join(pdf))
    with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
        f = myfile.read()
    if 'made a payment' not in f:
        sys.exit(f'{pdfpath} wrong pdf recieved, so not processed')
    else:
        if 'Balaji Medical' in f:
            op = 'Tpappg@maxhealthcare.com May@2020 outlook.office365.com Max PPT'
            hosp_name = 'Max'
        else:
            op = 'mediclaim@inamdarhospital.org Mediclaim@2019 imap.gmail.com inamdar hospital'
            hosp_name = 'inamdar'
    ###########################################################
    wbkName = 'temp_files/' + 'bajaj' + hosp_name + '.xlsx'
    t, wq =0, 0
    with sqlite3.connect("database1.db") as con:
        cur = con.cursor()
        q = f'select id, password from bajaj_credentials where hospital="{hosp_name}";'
        # print(q)
        cur.execute(q)
        r = cur.fetchone()

        if r:
            hosp_mail = r[0]
            hosp_pass = r[1]
        else:
            hosp_mail = ''
            hosp_pass = ''

    for t in range(0, len(onlyfiles)):
        with open(mypath + onlyfiles[t], "rb") as f:
            pdf = pdftotext.PDF(f)
        data = "\n\n".join(pdf)
        with open('temp_files/temppdf.txt', "w") as f:
            f.write(data)
        with open('temp_files/temppdf.txt', "r") as f:
            data = f.readlines()
        with open('temp_files/temppdf.txt', "r") as f:
            f = f.read()

        regex = r'(?<=UTR Reference).*'
        x = re.search(regex, f)
        if x:
            utr = x.group().strip()
        else:
            utr = ''

        regex = r'\d+ +\d{2}/\d{2}/.+'
        list1 = re.findall(regex, f)
        list1 = [re.split(r' {2,}', i) for i in list1]
        regex = r'\d{4} +\d{4}.*'
        list2 = re.findall(regex, f)
        list2 = [re.split(r' {2,}', i) for i in list2]

        table = []
        if len(list1) == len(list2):
            for i, j in zip(list1, list2):
                datadict = {}
                datadict['appr_no'] = i[0] + j[0]
                datadict['date'] = i[1] + j[1]
                if j[2].isdigit() is True:
                    datadict['name'] = i[2]
                else:
                    datadict['patientname'] = i[2] + ' ' + j[2]
                if len(j) > 3:
                    datadict['claim_no'] = i[3] + j[3]
                else:
                    datadict['claim_no'] = i[3] + j[2]
                datadict['bill_amount'] = i[4]
                datadict['paid_amount'] = i[-1].replace(',', '')
                datadict['tds_amount'] = i[-2]
                datadict['utr_no'] = utr
                table.append(datadict)
            for index, row in enumerate(table):
                params = (
                    ('patientname', r'(?<=Name Of The Patient)[ \S]+'),
                    ('idcardno', r'(?<=ID Card No)[ \S]+'),
                    ('claim_id', r'(?<=Claim ID)[ \S]+'),
                    ('claim_no', r'(?<=Claim Number)[ \S]+'),
                    ('doa', r'(?<=DOA:) ?\S+'),
                    ('dod', r'(?<=DOD:) ?\S+'),
                    ('appr_no', r'(?<=Approval Number)[ \S]+'),
                    ('utr_no', r'(?<=UTR No)[ \S]+'),
                    ('bill_amount', r'\d+(?=\s+Paid Amount)'),
                    ('paid_amount', r'\d+(?=\s+Disallowed Amount)'),
                    ('disallowed_amount', r'\d+(?=\s+TDS Amount)'),
                    ('tds_amount', r'\d+(?=\s+Hospital Service Tax No)'),
                )

                for i, j in params:
                    if i in row:
                        datadict[i] = row[i]
                    else:
                        datadict[i] = ''

                regex = r'\w+ ?Charges[\s\S]+(?=\n[\s\S]+Payment Details)'
                x = re.search(regex, f)
                # keyname = i[0]
                if x:
                    data = x.group().split('\n')
                    data = [re.split(r' {2,}', i) for i in data]
                    for i, j in enumerate(data):
                        if len(j) > 2:
                            for x, y in enumerate(data[i + 1:]):
                                if len(y) <= 2:
                                    j[-1] = j[-1] + ' ' + y[-1]
                                else:
                                    break
                    clean = []
                    for i, j in enumerate(data):
                        if len(j) > 2:
                            clean.append(j)
                else:
                    data = ''

                tempdata = table[index]

                mylist = [i[0] for i in params]
                mylist.append('date')
                mydata = []
                for i, j in params:
                    if i in datadict:
                        mydata.append(datadict[i])
                    else:
                        mydata.append("")
                mydata.append(tempdata['date'].replace('/','-'))
                wbk = openpyxl.Workbook()
                wbk.create_sheet('1')
                s1 = wbk.worksheets[0]
                s2 = wbk.worksheets[1]
                rowno = s1.max_row+1
                for i, j in enumerate(mylist):
                    s1.cell(row=1, column=i+1).value = j
                    s1.cell(row=rowno, column=i+1).value = mydata[i]

                mylist = ['Sr no','Particular', 'Bill Amount', 'Disallowed Amount', 'Approved Amount', 'Disallowance Reason']
                for i, j in enumerate(mylist):
                    s2.cell(row=1, column=i+1).value = j
                clean = []
                for i, j in enumerate(clean):
                    for x, y in enumerate(j):
                        s2.cell(row=i+2, column=1).value = i+1
                        s2.cell(row=i+2, column=x + 2).value = y

                wbname = 'bajaj'+hosp_name+'.xlsx'
                wbk.save('temp_files/'+wbname)
                wbk.close()

                print("Done")
                subprocess.run(["python", "make_master.py", 'bajaj', op, '', wbkName])
                ###########################################################
                move_master_to_master_insurer('')
                mark_flag('X', sys.argv[1])
                print(f'processed {wbkName}')
except SystemExit as e:
    v = e.code
    if 'exit' in v:
        a =1
        os._exit(0)
except:
    log_exceptions()
    pass