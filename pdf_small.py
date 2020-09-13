import os
import subprocess
import sys

import PyPDF2
import openpyxl
import pdftotext

from make_log import log_exceptions
from movemaster import move_master_to_master_insurer

try:
    pdfpath = sys.argv[1]
    onlyfiles = [os.path.split(pdfpath)[1]]
    mypath = os.path.dirname(pdfpath) + '/'

    hosp_name = ''

    with open(pdfpath, "rb") as f:
        pdf = pdftotext.PDF(f)
    with open('temp_files/output.txt', 'w', encoding='utf-8') as f:
        f.write(" ".join(pdf))
    with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
        f = myfile.read()
    if 'Bene Code' not in f:
        if 'Intimation No' in f:
            subprocess.run(["python", 'pdf_' + 'big' + ".py", sys.argv[1]])
            os._exit(1)
        sys.exit(f'{pdfpath} wrong pdf recieved, so not processed')
    else:
        if 'Balaji Medical' in f:
            op = 'Tpappg@maxhealthcare.com May@2020 outlook.office365.com Max PPT'
            hosp_name = 'Max'
        else:
            op = 'mediclaim@inamdarhospital.org Mediclaim@2019 imap.gmail.com inamdar hospital'
            hosp_name = 'inamdar'
    ###########################################################
    wbkName = 'temp_files/' + 'small' + hosp_name + '.xlsx'
    t, wq = 0, 0
    wbk = openpyxl.Workbook()
    for t in range(0, len(onlyfiles)):
        s1 = wbk.worksheets[0]
        sh2 = ['Chq /DD/Ft No', 'Amount', 'IFSC Code', 'Credit A/c No.', 'transaction date']
        for i in range(0, len(sh2)):
            s1.cell(row=1, column=i + 3).value = sh2[i]
        pdfFileObj = open(mypath + onlyfiles[t], 'rb')
        pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
        pageObj = pdfReader.getPage(0)
        # f = pageObj.extractText()
        # f = f.replace('\n', '$$ ')
        # # print(f)
        # text_file = open('temp_files/mail1.txt', "w")
        # n = text_file.write(f)
        # text_file.close()
        # pdfFileObj.close()

        gh = []
        x1 = f.find('INTIMATION NO') + 14
        g = f[x1:]
        y1 = g.find('_') + x1
        cli = (f[x1:y1])
        cli = cli.replace(' ', '')
        x2 = f.find('Chq /DD/Ft No') + 13
        g = f[x2:]
        x3 = g.find(':') + x2 + 1
        y2 = g.find('Value') + x2
        gh.append(f[x3:y2])

        x3 = f.find('Amount') + 7
        g = f[x3:]
        x4 = g.find(':') + x3 + 1
        y3 = g.find('Amount') + x3
        gh.append(f[x4:y3])

        x4 = f.find('IFSC Code') + 10
        g = f[x4:]
        y4 = g.find('through') + x4
        gh.append(f[x4:y4])

        x5 = f.find('Account Number') + 14
        g = f[x5:]
        y5 = g.find('with') + x5
        gh.append(f[x5:y5])

        x5 = f.find('Value Date :') + 12
        g = f[x5:]
        y5 = g.find('Amount') + x5
        gh.append(f[x5:y5])

        gh = [sub.replace('  ', '') for sub in gh]
        s1.cell(row=1, column=1).value = 'Sr. No.'
        s1.cell(row=1, column=2).value = 'INTIMATION NO'
        for i in range(0, len(gh)):
            s1.cell(row=t + 2, column=1).value = t + 1
            s1.cell(row=t + 2, column=2).value = cli
            s1.cell(row=t + 2, column=i + 3).value = gh[i]

    print("Done")
    wbk.save(wbkName)
    wbk.close()
    subprocess.run(["python", "make_master.py", 'small_star', op, '', wbkName])
    ###########################################################
    move_master_to_master_insurer('')
    print(f'processed {wbkName}')

except:
    log_exceptions()
    pass
