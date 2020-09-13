import os
import subprocess
import sys

import html2text
import openpyxl
import pdftotext

from make_log import log_exceptions
from movemaster import move_master_to_master_insurer

try:
    pdfpath = sys.argv[1]
    onlyfiles = [os.path.split(pdfpath)[1]]
    mypath = os.path.dirname(pdfpath)+'/'

    hosp_name = ''

    with open(pdfpath, "rb") as f:
        pdf = pdftotext.PDF(f)
    with open('temp_files/output.txt', 'w') as f:
        f.write(" ".join(pdf))
    with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
        f = myfile.read()
    if 'wanted_text' not in f:
        sys.exit(f'{pdfpath} wrong pdf recieved, so not processed')
    else:
        if 'Balaji Medical' in f:
            op = 'Tpappg@maxhealthcare.com May@2020 outlook.office365.com Max PPT'
            hosp_name = 'Max'
        else:
            op = 'mediclaim@inamdarhospital.org Mediclaim@2019 imap.gmail.com inamdar hospital'
            hosp_name = 'inamdar'
    ###########################################################
    wbkName = 'temp_files/' + 'health_heritage' + hosp_name + '.xlsx'
    t, wq =0, 0
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')
    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    for t in range(0, len(onlyfiles)):
        sh1 = ['Sr No.', 'Claim ID', 'Patient Name', 'I-Card No.', 'Policy No.', 'DOA', 'DOD', 'Illness',
               'Amount Claimed', 'Net Amount', 'NEFT No.', 'Date of Transfer', 'TDS Deducted', 'insurance company',
               ' primary beneficiary', 'bank', 'settle amt']
        sh2 = ['Sr No.', 'Claim ID', 'deductions amount', 'deductions Reasons']

        for i in range(0, len(sh1)):
            s1.cell(row=1, column=i + 1).value = sh1[i]
        for i in range(0, len(sh2)):
            s2.cell(row=1, column=i + 1).value = sh2[i]

        html = open('temp_files/attachments_' + hosp_name + str(t + 1) + ".html")
        f = str(html.read())
        w = open("temp_files/out.txt", "w")
        w.write(html2text.html2text(f))
        html.close()
        w.close()
        with open('temp_files/out.txt', 'r') as myfile:
            f = myfile.read()

        w = f.find('Details of deductions:') + 22
        u = f.find('Sincerely')
        g = f[w:u]
        g = g.replace('\n', '$$')
        g = g.replace('$$Rs.', '$$\n')
        g = g.replace('$$', ' ')
        g = g.replace('   ', '')
        g = g.replace('**', '')
        sy = g.split('\n')
        sy.remove('')

        hg = []
        w = f.find('Patient Name') + 13
        g = f[w:]
        u = g.find('|') + w
        hg.append(f[w:u])

        w = f.find('I-Card No.') + 12
        g = f[w:]
        u = g.find('|') + w
        hg.append(f[w:u])

        w1 = f.find('Policy No.') + 12
        g = f[w1:]
        u1 = g.find('|') + w1
        hg.append(f[w1:u1])

        w2 = f.find('DOA') + 6
        g = f[w2:]
        u2 = g.find('|') + w2
        hg.append(f[w2:u2])

        w9 = f.find('DOD') + 6
        g = f[w9:]
        u9 = g.find('\n') + w9
        hg.append(f[w9:u9])

        w2 = f.find('Illness') + 10
        g = f[w2:]
        u2 = g.find('|') + w2
        hg.append(f[w2:u2])

        w9 = f.find('Amount Claimed') + 18
        g = f[w9:]
        u9 = g.find('|') + w9
        hg.append(f[w9:u9])

        w9 = f.find('Net Amount') + 11
        g = f[w9:]
        u9 = g.find('\n') + w9
        hg.append(f[w9:u9])

        w9 = f.find('NEFT No.:') + 9
        g = f[w9:]
        u9 = g.find('Date of Transfer') + w9
        hg.append(f[w9:u9])

        w2 = f.find('Date of Transfer') + 16
        g = f[w2:]
        u2 = g.find('Settled Amount') + w2
        hg.append(f[w2:u2])

        w9 = f.find('TDS Deducted :') + 15
        g = f[w9:]
        u9 = g.find('Deducted') + w9
        hg.append(f[w9:u9])

        w9 = f.find('insurer') + 8
        g = f[w9:]
        u9 = g.find(',') + w9
        hg.append(f[w9:u9])

        w9 = f.find('Employee Name') + 14
        g = f[w9:]
        u9 = g.find('\n') + w9
        hg.append(f[w9:u9])
        if f.find('Drawn On') != -1:
            w9 = f.find('Drawn On') + 9
            g = f[w9:]
            u9 = g.find('\n') + w9
            hg.append(f[w9:u9])
        else:
            hg.append(' ')
        w9 = f.find('settled for') + 16
        g = f[w9:]
        u9 = g.find('on') + w9
        hg.append(f[w9:u9])

        w9 = f.find('CCN :') + 5
        g = f[w9:]
        u9 = g.find(')') + w9
        ccn = f[w9:u9]
        if ccn.find(' ') != -1:
            u9 = g.find(' ') + w9
            ccn = f[w9:u9]
        hg = [sub.replace('  ', '') for sub in hg]
        hg = [sub.replace(':', '') for sub in hg]
        hg = [sub.replace('*', '') for sub in hg]
        hg = [sub.replace('\n', ' ') for sub in hg]
        s1.cell(row=t + 2, column=1).value = t + 1
        s1.cell(row=t + 2, column=2).value = ccn
        for i in range(0, len(hg)):
            s1.cell(row=t + 2, column=i + 3).value = hg[i]
        # print(hg)
        for i in sy:
            max_row = s2.max_row + 1
            k = i.find('.')
            jk = i[:k + 3]
            kj = i[k + 3:]
            wq += 1
            s2.cell(row=max_row, column=1).value = wq
            s2.cell(row=max_row, column=2).value = ccn
            s2.cell(row=max_row, column=3).value = jk
            s2.cell(row=max_row, column=4).value = kj


    print("Done")
    wbk.save(wbkName)
    wbk.close()
    subprocess.run(["python", "make_master.py", 'health_heritage', op, '', wbkName])
    ###########################################################
    move_master_to_master_insurer('')
    print(f'processed {wbkName}')
except:
    log_exceptions()
    pass