import os
import subprocess
import PyPDF2
import openpyxl
import sys
import camelot
import pdftotext
import xlrd
from make_log import log_exceptions
from backend import mark_flag
from movemaster import move_master_to_master_insurer

try:
    pdfpath = sys.argv[1]
    onlyfiles = [os.path.split(pdfpath)[1]]
    mypath = os.path.dirname(pdfpath)+'/'

    hosp_name = ''

    with open(pdfpath, "rb") as f:
        pdf = pdftotext.PDF(f)
    with open('temp_files/output.txt', 'w', encoding='utf-8') as f:
        f.write(" ".join(pdf))
    with open('temp_files/output.txt', 'r',  encoding='utf-8') as myfile:
        f = myfile.read()
    if 'wanted_text' not in f:
        sys.exit(f'{pdfpath} wrong pdf recieved, so not processed')
    else:
        if 'Balaji Medical' in f:
            op = 'Tpappg@maxhealthcare.com May@2020 outlook.office365.com Max PPT'
            hosp_name = 'Max'
        else:
            op = 'mediclaim@inamdarhospital.org Mediclaim@2019 imap.gmail.com inamdar hospital'
            hosp_name = 'inamdar'
    ###########################################################
    wbkName = 'temp_files/' + 'health_india' + hosp_name + '.xlsx'
    t, wq =0, 0
    wbk = openpyxl.Workbook()
    wbk.create_sheet('1')

    s1 = wbk.worksheets[0]
    s2 = wbk.worksheets[1]
    sh1 = ['Sr No.', 'Claim Number', 'Insurance Company', 'Policy Number', 'Corporate/Retail', 'Type of Claim',
           'Employee Code', 'Employee Name', 'Patient Name', 'Hospital Name', 'Hospital City',
           'Date of Admission - Discharge', 'Ailment', 'UTR No', 'UTR Date', 'Claim Amount', 'Deduction Amount',
           'Discount Amount', 'Approved Amount', 'TDS Amount', 'NEFT/Paid Amount']
    sh2 = ['Sr No.', 'Claim ID', 'Bill No.', 'Bill Date', 'Bill Amt', 'Payable Amt', 'Disallowance amount',
           'Disallowance Reasons', 'Deduction Category']
    for t in range(0, len(onlyfiles)):
        for i in range(0, len(sh1)):
            s1.cell(row=1, column=i + 1).value = sh1[i]
        for i in range(0, len(sh2)):
            s2.cell(row=1, column=i + 1).value = sh2[i]
        tables = camelot.read_pdf(mypath + onlyfiles[t], pages='all')
        tables.export('temp_files/foo1.xls', f='excel')
        loc = ("temp_files/foo1.xls")
        wb = xlrd.open_workbook(loc)
        s = []
        d = []
        sheet_3 = wb.sheet_by_index(0)
        sheet_3.cell_value(0, 0)

        for i in range(1, sheet_3.nrows - 1):
            d.append(sheet_3.cell_value(i, 1))
            s.append(sheet_3.cell_value(i, 2))
        ccn = s[0]
        s = [sub.replace('\t', ' ') for sub in s]
        d = [sub.replace('\t', ' ') for sub in d]
        d = [sub.replace('\n', ' ') for sub in d]
        s = [sub.replace('Rs.', '') for sub in s]
        s = [sub.replace('/-', '') for sub in s]
        # print(s)
        s1.cell(row=t + 2, column=1).value = t + 1
        for i in range(0, len(d)):
            k = sh1.index(d[i])
            s1.cell(row=t + 2, column=k + 1).value = s[i]
        sheet_1 = wb.sheet_by_index(1)
        sheet_1.cell_value(0, 0)
        d = []
        s = []
        p = []
        r = []
        e = []
        ro = []
        po = []
        for i in range(2, sheet_1.nrows):
            d.append(sheet_1.cell_value(i, 1))
            s.append(sheet_1.cell_value(i, 2))
            p.append(sheet_1.cell_value(i, 3))
            r.append(sheet_1.cell_value(i, 4))
            e.append(sheet_1.cell_value(i, 5))
            ro.append(sheet_1.cell_value(i, 6))
            po.append(sheet_1.cell_value(i, 7))
        if (tables.n == 3):
            sheet_2 = wb.sheet_by_index(2)
            sheet_2.cell_value(0, 0)
            for i in range(1, sheet_2.nrows):
                d.append(sheet_2.cell_value(i, 1))
                s.append(sheet_2.cell_value(i, 2))
                p.append(sheet_2.cell_value(i, 3))
                r.append(sheet_2.cell_value(i, 4))
                e.append(sheet_2.cell_value(i, 5))
                ro.append(sheet_2.cell_value(i, 6))
                po.append(sheet_2.cell_value(i, 7))
        # print(p)
        num_ko = s1.max_row
        for i in range(0, len(d)):
            if (d[i] != ''):
                if po[i].find('Discount') != -1:

                    for ko in range(1, num_ko + 1):
                        # print(e[i],ccn,s1.cell(row=ko, column=2).value)
                        if (ccn == s1.cell(row=ko, column=2).value):
                            s1.cell(row=ko, column=18).value = e[i]
                        # print(e[i])
                row_num = s2.max_row + 1
                wq += 1
                s2.cell(row=row_num, column=1).value = wq
                s2.cell(row=row_num, column=2).value = ccn
                s2.cell(row=row_num, column=3).value = d[i]
                s2.cell(row=row_num, column=4).value = s[i]
                s2.cell(row=row_num, column=5).value = p[i]
                s2.cell(row=row_num, column=6).value = r[i]
                s2.cell(row=row_num, column=7).value = e[i]
                s2.cell(row=row_num, column=8).value = ro[i]
                s2.cell(row=row_num, column=9).value = po[i]


    print("Done")
    wbk.save(wbkName)
    wbk.close()
    subprocess.run(["python", "make_master.py", 'health_india', op, '', wbkName])
    ###########################################################
    move_master_to_master_insurer(sys.argv[2], pdfpath=pdfpath)
    mark_flag('X', sys.argv[2])
    print(f'processed {wbkName}')
except SystemExit as e:
    v = e.code
    if 'exit' in v:
        a =1
        os._exit(0)
except:
    log_exceptions()
    pass